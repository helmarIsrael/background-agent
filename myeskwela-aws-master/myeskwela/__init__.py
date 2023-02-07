# -*- coding: utf-8 -*-

from unicodedata import name
from flask import Flask, jsonify, request, url_for, send_file, make_response, render_template, Response
from flask_httpauth import HTTPBasicAuth
import sys, hashlib, flask, os, types, re, io, urllib.request
from flask_mail import Mail, Message
from datetime import date
from PIL import Image
#from cloudpic import fileServer, megaServer
from io import StringIO
#from model import DBconn, spcall
#from utils import *
from openpyxl import Workbook
from flask_jsglue import JSGlue
from werkzeug.utils import secure_filename
import random, string

import pubnubex as pub
import json

#### cloudpic.py ####
import cloudinary
import cloudinary.uploader
import cloudinary.api

import re

class megaServer:
    def __init__(self):
        mega = Mega()
        self.m = mega.login('orven.llantos@gmail.com', '3ZeyucxJsW')

    def upload(self, par_file):
        try:
            file = self.m.upload(par_file)
            link = self.m.get_upload_link(file)
            #do something
            return {"status": "ok", "file": link}#,"result": result}

        except:
            return {"status":"error", "message":str(sys.exc_info()[0]) +
                " " + str(sys.exc_info()[1])}

    def getFile(self, par_fileurl):
        file = self.m.download_url(par_fileurl, 'uploads/')
        return file


class fileServer:
    def __init__(self):
        cloudinary.config(
            cloud_name = "myeskwela",
            api_key = "194739911629951",
            api_secret="BI9aqoHU-j2rYP-a-FFAl6cs_wk"
        )

    def upload(self, file, par_personid):
        try:
            result = cloudinary.uploader.upload(
                file,
                public_id=par_personid,
                crop='limit',
                width=150,
                height=200,
                eager = [{
                    'width':200, 'height':200,
                    'crop':'thumb', 'gravity':'face',
                    'radius': 20, 'format':'jpg'
                }],
                tags=['for_homepage']
            )

            return {"status": "ok","result": result}

        except:
            return {"status":"error", "message":str(sys.exc_info()[0]) +
                " " + str(sys.exc_info()[1])}


    def getImage(self,  par_personid):
        return '<img src="' + spcall("getcloudinaryurl", (par_personid,))[0][0] +'" />'

    def getImageUrl(self, par_personid):
        return  spcall("getcloudinaryurl", (par_personid,))[0][0]


            #cloudinary.CloudinaryImage(par_personid + ".jpg")\
            #.image(alt = "User Image")
        # '<img src=' +\
        #       "'http://res.cloudinary.com/demo/image/upload/v1/" + par_personid + ".jpg' />"

    def getImgSrcUrl(self, par_cloudinaryurl):
        return '<img src="' + par_cloudinaryurl +'" />'

#### cloudpic.py ####

#### dosql
import sys
import psycopg2
import os
#from utils import constr

class doSql(object):
    #attributes
    _cxn = ""
    _cur = ""
    errmsg =""

    #methods
    def __init__(self, ptype): #constructor
        self._cxn = psycopg2.connect(constr(ptype))
        self._cur = self._cxn.cursor()

    def __del__(self): #destructor
        self._cur.close()
        self._cxn.close()

    def execqry(self, sql, apply_):
        rows = []
        try:
            self._cur.execute(sql)
            rows = self._cur.fetchall()
            if apply_:
                self._cxn.commit()
                if self._cur.rowcount == 0:
                    rows.append(['None'])
        except:
            errmsg =  str(sys.exc_info()[1])
            rows.append([errmsg])
        return rows

### dosql

### utils
from datetime import datetime
import re
from openpyxl.styles import Alignment
import os

def age(when, on=None):
    if on is None:
        on = datetime.today()
    was_earlier = (on.month, on.day) < (when.month, when.day)
    return on.year - when.year - (was_earlier)


def autoswitch():
    yearnow = datetime.today()

    if yearnow.month <= 6:
        now = datetime.strptime("06/01/" + str(yearnow.year), '%m/%d/%Y')
    else:
        now = yearnow
    return now

def cleandat(d, pattern, rep):
    ret = d
    if len(d) > 0:
        ret = d.replace(pattern, rep)
    return ret



def restrequest(spname, spparams, spcall, spgroup, spcommit, jsonify):
    #for use in SPs that returns JSON, or singleton

    res = spcall(spname, spparams, spgroup, spcommit)[0][0]
    if 'Error' in str(res):
        return jsonify({
            'status': "error",
            'message': str(res).split('CONTEXT:')[0].split('DETAIL:')[0].split('LINE')[0]
        })

    return jsonify(res)



#for setup: postgre is either owner or superuser
def sqlAlconn(ptype):
    #this assumes that a DATABASE_URL is properly set-up at the
    #localhost machine as well as the cloud machine
    #sample DATABASE_URL = postgresql://username:password@ipaddress:ipport/dbname
    #_osconf = os.environ['DATABASE_URL'].split('@')[1].split("/")[0].split(":")
    #_dbadd =  _osconf[0]#"localhost"#osconf[0] #'sl-us-south-1-portal.27.dblayer.com'#_osconf[0] #
    #_dbport = _osconf[1]#_osconf[1] #'5432'#'47445'#'5433'#

    _connstr = {"super": { 'user':'postgre',
                           'password':'UflY9ib43U'
                         },
                "faculty":{'user':'teacher',
                            'password':'T5XP9MbHoR'
                          },
                "admin":  {'user':'trator',
                           'password':'k2oDBalkTs'
                          },
                "principal": {'user': 'trator',
                          'password': 'k2oDBalkTs'
                          },
                "students": {'user':'aaral',
                             'password':'YxYKWg48l7'
                            },
                "staff":   {'user':'assist',
                           'password':'cEBAdKWFcL'
                           },
                "registrar":
                           {'user':'regis',
                          'password':'gNNVaY6MKZ'
                           },
                "parents":{
                          'user':'roots',
                          'password':'FrXnD4UGCF'
                          },
                "super10":{
                           'user':'super10dent',
                           'password': 'uz7En2vAmE'
                           },
                "bisor":
                    {
                        'user': 'subisor',
                        'password': 'q0owSP3FLa'


                },
                "ispeval":
                {
                        'user': 'ispbisor',
                        'password': 'mtFq5EEZca'
                },
                "subadmin":
                    {
                        'user': 'subadminagent',
                        'password': 'SwR18ci7fE'
                    },
                "mande":
                    {
                        'user': 'mandemore',
                        'password': 'qC7NpLsdqr'
                    },

                "datamart":{
                          'user':'postgres',
                          'password':'12345'
                }
                }

    
    _dbname = 'my2' #local:'my2'

    _dbadd = 'localhost' #'sl-us-south-1-portal.27.dblayer.com'
    _dbport = '5432' #'47445'

    #_dbadd = 'localhost'
    #_dbport = '5433'

    return "postgresql://"+ _connstr[ptype]["user"] + \
            ":" + _connstr[ptype]["password"] + "@"+ \
            _dbadd + ":" + _dbport + "/" + _dbname
#print age(datetime.strptime("12/1/1978", '%m/%d/%Y'), autoswitch())

#def age(dateString):
#  #code originally from http://stackoverflow.com/a/10119380
#    yearnow = datetime.today()

#    if yearnow.month <= 6:
#        now = datetime.strptime("06/01/" + str(yearnow.year), '%m/%d/%Y')
#    else:
#        now = yearnow



#    yearNow = now.year
#    monthNow = now.month
#    dateNow = now.day

#    dob = datetime.strptime(dateString, '%m/%d/%Y')

#    yearDob = dob.year
#    monthDob = dob.month
#    dateDob = dob.day

#    yearAge = yearNow - yearDob

#    if monthNow >= monthDob:
#        monthAge = monthNow - monthDob
#    else:
#        yearAge = yearAge - 1
#        monthAge = 12 + monthNow - monthDob

#    if dateNow >= dateDob:
#        dateAge = dateNow - dateDob
#    else:
#        monthAge = monthAge - 1
#        dateAge = 31 + dateNow - dateDob
#        if monthAge < 0:
#            monthAge = 11
#            yearAge = yearAge - 1
#    return yearAge

#print sqlAlconn("registrar")


def checkpassword(password):
    errorm = ''
    if len(password) < 8:
        errorm += ' Must be at least 8 characters.'

    if len(re.compile(r'[0-9]').findall(password)) == 0:
        errorm += ' Must contain at least one digit.'

    if len(re.compile(r'[a-z]').findall(password)) == 0:
        errorm += ' Must contain at least one small character.'

    if len(re.compile(r'[A-Z]').findall(password)) == 0:
        errorm += ' Must contain at least one capital character.'

    return errorm


def updatepassword(par_username):
    password = spcall('random_string',(10,))[0][0]
    res = spcall('updatepassword', (par_username, password), 'super', True)
    if res[0][0] == 'OK':
        return par_username + '->' + password
    else:
        return res[0][0]

def centerCellText(ws,address):
    currentCell = ws.cell(address)
    currentCell.alignment = Alignment(horizontal='center')

def rightCellText(ws,address):
    currentCell = ws.cell(address)
    currentCell.alignment = Alignment(horizontal='right')

def leftCellText(ws,address):
    currentCell = ws.cell(address)
    currentCell.alignment = Alignment(horizontal='left')

def appendBlanks(ws, n):
    for i in range(n):
        ws.append([''])

def mergeAlign(ws, row,range, alignfcn):
    ws.append(row)
    ws.merge_cells(range)
    alignfcn(ws, range.split(":")[0])

def mergeAlignOnly(ws, range, alignfcn):
    ws.merge_cells(range)
    alignfcn(ws, range.split(":")[0])

#thanks to: https://stackoverflow.com/a/48698261
def letter_range(start, stop="{", step=1):
    """Yield a range of lowercase letters."""
    for ord_ in range(ord(start.lower()), ord(stop.lower()), step):
        yield chr(ord_)


def finalize(res):
    if 'Error' in res[0][0]:
        return {"status": "error", "message": res[0][0]}


    return res[0][0]

def cleanhtml(raw_html): #thanks to: https://stackoverflow.com/a/12982689
  cleanr = re.compile('<.*?>')
  cleantext = re.sub(cleanr, '', raw_html)
  return cleantext

def errordict(errmsg):
    return {
            'status': 'error',
            'msg': str(errmsg)
        }

def lengthesize(v,l):
    if len(v) <= l:
        return v
    else:
        return v[0:l]

def forcemax(token):
    return lengthesize(token, 128)
### utils


#### model.py ###
from sqlalchemy import create_engine
from sqlalchemy.pool import NullPool
#from doSql import doSql
import os
#from utils import sqlAlconn
import sys

class DBconn:
    def __init__(self, ptype):
        #dbadd = os.environ['OPENSHIFT_POSTGRESQL_DB_HOST']
        #dbport = os.environ['OPENSHIFT_POSTGRESQL_DB_PORT']
        #dbloc = dbadd + ":" + dbport
        #engine = create_engine("postgresql://myeskwela2:letmein@"+dbloc+"/myeskwela", echo=False)
        self.engine = create_engine(sqlAlconn(ptype),
                                    poolclass=NullPool)

        self.conn = self.engine.connect()
        self.trans = self.conn.begin()

    def getcursor(self):
        cursor = self.conn.connection.cursor()
        return cursor

    def dbcommit(self):
        self.trans.commit()

    def close(self):
        self.conn.close()
        self.engine.dispose()

class DBConn2:
    def __init__(self, type):
        self.db = doSql(type)

    def execQry(self, qry, commit=False):
        return self.db.execqry(qry, commit)


def spcall(qry, param, ptype="super",commit=False):
    try:
        dbo = DBconn(ptype)
        cursor = dbo.getcursor()
        cursor.callproc(qry, param)
        res = cursor.fetchall()
        if commit:
            dbo.dbcommit()
        dbo.close()
        return res
    except:
        res = [("Error: " + str(sys.exc_info()[0]) +
                " " + str(sys.exc_info()[1]),)]
    return res
#### model.py



#thanks to: https://pynative.com/python-generate-random-string/
def get_random_string(length):
    letters = string.ascii_lowercase
    result_str = ''.join(random.choice(letters) for i in range(length))
    return result_str


def get_random_alphanumeric_string(length):
    letters_and_digits = string.ascii_letters + string.digits
    result_str = ''.join((random.choice(letters_and_digits) for i in range(length)))
    return result_str


from psycopg2.extras import Json #to pass Json from py to pl

#template_dir = os.path.dirname(os.path.dirname(os.path.abspath(os.path.dirname(__file__))))

app = Flask(__name__)
jsglue = JSGlue(app)

app.config['MAIL_SERVER']='smtp.gmail.com'
app.config['MAIL_PORT'] = 465
app.config['MAIL_USERNAME'] = 'myeskwela.app@gmail.com'
app.config['MAIL_PASSWORD'] = 'wc8VQ2sfNYkH7P9'
app.config['MAIL_USE_TLS'] = False
app.config['MAIL_USE_SSL'] = True

mail = Mail(app)
auth = HTTPBasicAuth()
#port = int(os.getenv('VCAP_APP_PORT'))
#local
#app.config['UPLOAD_FOLDER'] = "/Users/orvenllantos/OneDrive/uone/extension/myeskwela/code/myeskwela-aws/myeskwela/uploads"
#production
app.config['UPLOAD_FOLDER'] = "/var/www/myeskwela/myeskwela/uploads"
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 2024
#command to tell the browser not to cache pages
#so that it will fetch all the files once refreshed
#this is also to make sure that any changes from updating
#the app will automatically reflect in the client browsers
#once refreshed
#thanks to: https://stackoverflow.com/a/54422901
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

def digest(word):
    return hashlib.sha1(word.encode('utf-8')).hexdigest()

@auth.get_password
def getpassword(username):
    #return 'Pokemon2go'
    udig = str(digest(username))
    password = spcall("get_password", (udig,))[0][0]
    return password


@app.route("/")
def hello():


    disable = spcall("disablesite", (), "super")[0][0]
    if disable:
        return render_template('migration.html')

    announce = spcall("getsystemannouncements", (), "super")[0][0]

    box = ''
    if len(announce) > 0:
        parts = announce.split(",")
        box = {'header':parts[0], 'contents': parts[1], 'footer':parts[2]}

    message = {'message':box}
    return render_template('index.html', message=message)
    #password = spcall("get_password", (str(digest('orven.ebarle.llantos')),))[0][0]
    #return render_template('index.html', message={'message':{'header':'hello', 'contents': password, 'footer':'sdfs'}})


def putcomma(par_value):
    if len(par_value) == 0:
        return ''
    return par_value + ","


def ztok(par_val):
    if par_val == 0:
        return 'Kinder'
    return par_val


@app.route("/profile/<string:idnum>/<string:schoolid>/<string:semid>", methods=['GET'])
@auth.login_required
def fetchprofile(idnum, schoolid, semid):
    username = auth.username()
    group = spcall("getuser_group", (username,))[0][0]
    # getidnumbyuser(par_diguser text, par_ptype text)

    if group == 'students':
        useridnum = spcall("getidnumbyuser", (digest(username),
                                              group), group)[0][0]
        if useridnum != idnum:
            return jsonify({"status": "error",
                            "message": "Access restricted."})
    notmykid = True
    if group == 'parents':
        parentid = spcall("getpersonidbyusername", (username,))[0][0]
        kids = spcall("getmykids", (parentid, 'students'), group)
        for kid in kids:
            if kid[0] == idnum:
                notmykid = False
                break
        if notmykid:
            return jsonify({"status": "error",
                            "message": "Access restricted."})

    details = formatres(spcall("getpersondetails",
                               (idnum, schoolid, semid, False)))

    if details["status"] == 'Error':
        return jsonify(details)
    # return jsonify(details)

    details = details["item"][0].split("@")
    stuname = details[0].split("*")
    father = details[11].split("*")
    mother = details[12].split("*")
    guardian = details[13].split("*")
    sex = details[1].upper()

    if schoolid == 'NOT SET':
        schoolid = details[14]

    lvlsxn = spcall("getlevelsection",
                    (idnum, schoolid, semid))
    f = fileServer()
    imgurl = f.getImageUrl(spcall("getpersonidbyidnum", (idnum,))[0][0])
    # if sex == 'F':
    imgsrc = imgurl
    # else:
    #    imgsrc = "../dist/img/user1-128x128.jpg"

    if len(lvlsxn) > 0:
        dets = lvlsxn[0]
        moredetails = [
            {"col": ['Grade', ztok(dets[2])]},
            {"col": ['Section', dets[1]]},
            {"col": ['Adviser', dets[3]]}]
    else:
        moredetails = [
            {"col": ['Grade', "-"]},
            {"col": ['Section', "-"]},
            {"col": ['Adviser', "-"]}]

    healthinfo = spcall("getstudentinfo", (idnum,))[0][0].split('*')

    stuload = spcall("getpersonload", (idnum, semid))

    subjects = []

    if len(stuload) > 0:
        for load in stuload:
            subjects.append({
                "offeringid": load[0],
                "subject": load[1]
            })

    return jsonify({
        "status": "ok",
        "img": imgsrc,
        "name": stuname[1] + " " +
                stuname[2] + " " +
                stuname[0],
        "lrn": idnum,
        "moredetails": moredetails,
        "subjectload": subjects,
        "aboutme": [
            {
                'icon': "fa fa-ambulance",
                'label': "Nutritional Status",
                'value': healthinfo[0]
            },
            {
                'icon': "fa fa-ruble",
                'label': "Conditional Transfer",
                'value': healthinfo[1]
            },
            {
                'icon': "fa fa-wheelchair",
                'label': "Disability",
                'value': healthinfo[2]
            },
            {
                "icon": "fa fa-home margin-r-5",
                "label": "Address",
                "value": putcomma(details[6]) + "<br />" +
                         putcomma(details[7]) + "<br />" +
                         putcomma(details[8]) + "<br />" +
                         putcomma(details[9]) + "<br />" +
                         details[10]

            },
            {
                "icon": "fa fa-bank margin-r-5",
                "label": "Father",
                "value": father[1] + " " +
                         father[2] + " " +
                         father[0]
            },
            {
                "icon": "fa fa-heart margin-r-5",
                "label": "Mother",
                "value": mother[1] + " " +
                         mother[2] + " " +
                         mother[0]
            },
            {
                "icon": "fa fa-user-secret margin-r-5",
                "label": "Guardian",
                "value": guardian[1] + " " +
                         guardian[2] + " " +
                         guardian[0]
            }

        ],
        "btns": [
            {
                "bclass": "btn btn-success btn-xs",
                "label": "Enroll",
                "fcn": "model.enroll('" + idnum + "', 'btnprofenroll')",
                "id": "btnprofenroll"
            },
            {
                "bclass": "btn btn-warning btn-xs",
                "label": "Edit",
                "fcn": "view.regautofill('" + idnum + "','btnprofedit')",
                "id": "btnprofedit"
            },
            {
                "bclass": "btn btn-danger btn-xs",
                "label": "Remove",
                "fcn": "model.removestu('" + idnum + "', 'btnprofremove')",
                "id": "btnprofremove"
            }],
        "historydata": {
            "labels": ["January", "February", "March", "April", "May", "June", "July"],
            "datasets": [
                {
                    "label": "Electronics",
                    "fillColor": "rgba(210, 214, 222, 1)",
                    "strokeColor": "rgba(210, 214, 222, 1)",
                    "pointColor": "rgba(210, 214, 222, 1)",
                    "pointStrokeColor": "#c1c7d1",
                    "pointHighlightFill": "#fff",
                    "pointHighlightStroke": "rgba(220,220,220,1)",
                    "data": [65, 59, 80, 81, 56, 55, 40]
                },
                {
                    "label": "Digital Goods",
                    "fillColor": "rgba(60,141,188,0.9)",
                    "strokeColor": "rgba(60,141,188,0.8)",
                    "pointColor": "#3b8bba",
                    "pointStrokeColor": "rgba(60,141,188,1)",
                    "pointHighlightFill": "#fff",
                    "pointHighlightStroke": "rgba(60,141,188,1)",
                    "data": [28, 48, 40, 19, 86, 27, 90]
                }
            ]
        },
        "donutdata": [
            {
                "value": 700,
                "color": "#f56954",
                "highlight": "#f56954",
                "label": "Chrome"
            },
            {
                "value": 500,
                "color": "#00a65a",
                "highlight": "#00a65a",
                "label": "IE"
            },
            {
                "value": 400,
                "color": "#f39c12",
                "highlight": "#f39c12",
                "label": "FireFox"
            },
            {
                "value": 600,
                "color": "#00c0ef",
                "highlight": "#00c0ef",
                "label": "Safari"
            },
            {
                "value": 300,
                "color": "#3c8dbc",
                "highlight": "#3c8dbc",
                "label": "Opera"
            },
            {
                "value": 100,
                "color": "#d2d6de",
                "highlight": "#d2d6de",
                "label": "Navigator"
            }
        ],
        "schoolid":schoolid
    })


ALLOWED_EXTENSIONS = ['jpg', 'jpeg', 'JPG', 'JPEG']


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1] in ALLOWED_EXTENSIONS


@app.route("/image", methods=['GET'])
# @auth.login_required
def dispimage():
    try:
        f = fileServer()
        url = f.getImageUrl('S2016super-9')
        # with Image.open(StringIO(urllib.urlopen(url).read())) as img:
        #    def wsgi_app(environ, start_response):
        #        start_response('200 OK', [('Content-type', 'image/jpeg')])
        #        return img.read()
        #    return make_response(wsgi_app)

        #request = urllib2.Request(url) deprecared from python 2.7
        img = StringIO.StringIO(urllib.request.urlopen(url).read())
        return send_file(img, mimetype='image/jpg')
    except:
        return jsonify({"status": "error", "message": "Error: " + str(sys.exc_info()[0]) +
                                                      " " + str(sys.exc_info()[1])})

@app.route("/file", methods=['POST'])
@auth.login_required
def filesupload():
    #return jsonify({"status": "error", "message": "File Upload is not available."})
    username = request.form["username"]
    group = request.form["group"] #spcall("getuser_group", (username,))[0][0]
    token = request.form["token"]

    files = request.files.getlist('file')
    #mega = megaServer()
    links = []
    fileexpos = []
    for file in files:
        fileexpos.append(file.filename.rsplit('.', 1)[0][0:5] + "." + file.filename.rsplit('.', 1)[1])
        filename = os.path.join(app.config['UPLOAD_FOLDER'],
                                "%s.%s" %
                                ('my.e.' + digest(username + '_' + file.filename.rsplit('.', 1)[0]),
                                 file.filename.rsplit('.', 1)[1]))
        file.save(filename)
        #links = mega.upload(filename)
        #print(links)
        #return jsonify({'status':'OK', 'links': links})
        fileid = 'my.e.' + digest(username + '_' + file.filename.rsplit('.', 1)[0]) + '.'  +file.filename.rsplit('.', 1)[1]
        links.append(fileid)
        #os.remove(filename)

    #print(username)
    #print(group)
    #print(token)

    return jsonify({"status": "OK", "links": links, "filenames": fileexpos})


@app.route("/file", methods=['GET'])
#@auth.login_required
def filesdownload():
    params = request.args
    link = params["link"]
    #print(link)
    #mega = megaServer()
    #file = mega.getFile("https://mega.co.nz/#!" + link)
    filename = os.path.join(app.config['UPLOAD_FOLDER'], link)
    return send_file(filename)

@app.route("/image", methods=['POST'])
#@auth.login_required
def uploadimage():
    username = request.form["username"]
    group = request.form["group"] #spcall("getuser_group", (username,))[0][0]

    lrn = request.form["lrn"]
    token = request.form["token"]




    if len(lrn) == 0:
        resuserid = spcall("getpersonidbyusrtoken", (username, token, group), group)
        if formatres(resuserid)["status"] == "Error":
            return jsonify({"status": "error", "message": formatres(resuserid)["message"]})
        qrypersonid = resuserid[0][0]
    else:
        reslrnid = spcall("getpersonidbyidnum2ken", (lrn, username, token, group), group)
        qrypersonid = reslrnid[0][0]

    v = re.search('NO', qrypersonid)

    if type(v) != type(None):
        return jsonify({"status": "error", "message": "DATA ERROR!"})

    personid = qrypersonid

    file = request.files['file']
    if file and allowed_file(file.filename):
        # now = datetime.now()
        filename = os.path.join(app.config['UPLOAD_FOLDER'],
                                "%s.%s" %
                                (file.filename.rsplit('.', 1)[0],
                                 file.filename.rsplit('.', 1)[1]))

        #fname = secure_filename(filename)
        fname = filename
        file.save(fname)
        f = fileServer()
        result = f.upload(fname, personid)
        os.remove(fname)
        # return jsonify()

        if result["status"] == 'error':
            return jsonify(
                {"status":"error",
                 "message": result["message"]})

        secureurl = result["result"]["secure_url"]
        result = spcall("insupcloudinary", (personid, secureurl), "super", True)
        return jsonify(
            {"status": "ok", "message": secureurl})  # result["secure_url"]})#, 'filename':filename, "lrn":lrn})

    return jsonify({"status": "Error"})
    # return jsonify({"status": "ok", "image": image.filename})
    # f = fileServer()
    # result = f.upload(file, 'testing')
    # return jsonify(result)


@app.route("/student/<string:idnum>/<string:schoolid>", methods=['GET'])
@auth.login_required
def getstudetails(idnum, schoolid):
    details = spcall("getpersondetails",
                     (idnum, schoolid))[0][0]  # .split("@")
    # return jsonify({"satus":"error", "message":details})

    if "error" in details or "Error" in details:
        return jsonify({
            "status": "error",
            "message": details
        })

    moreinfo = spcall("getstudentinfo", (idnum,))[0][0].split('*')

    # "LASTNAME*FIRSTNAME*MI@M@2001-08-12@mt@IP@re@WEEW@TIB@FGW@EE@2420@WER*T*ERT@ERTE*RET*RET@ERT*WRT24*TW"
    details = details.split("@")
    stuname = details[0].split("*")
    father = details[11].split("*")
    mother = details[12].split("*")
    guardian = details[13].split("*")

    return jsonify({"status": "ok",
                    "student": {"lname": stuname[0],
                                "fname": stuname[1],
                                "mi": stuname[2]},
                    "sex": details[1],
                    "birthdate": details[2],
                    "tongue": details[3],
                    "ip": details[4],
                    "religion": details[5],
                    "address": {
                        "streetno": details[6],
                        "brgy": details[7],
                        "city": details[8],
                        "state": details[9],
                        "zip": details[10]
                    },
                    "father": {
                        "lname": father[0],
                        "fname": father[1],
                        "mi": father[2]
                    },
                    "mother": {
                        "lname": mother[0],
                        "fname": mother[1],
                        "mi": mother[2]
                    },
                    "guardian": {
                        "lname": guardian[0],
                        "fname": guardian[1],
                        "mi": guardian[2]

                    },
                    "moreinfo": {
                        'nutritionalstatus': moreinfo[0],
                        'conditionalcashtransfer': moreinfo[1],
                        'disabilities': moreinfo[2]
                    }
                    })


@app.route("/student/<string:lastname>", methods=['GET'])
@auth.login_required
def searchlist(lastname):
    params = request.args
    token = params["token"]
    return searchoperson(lastname, auth.username(), token,'students')


def searchoperson(lastname, username, token, searchgroup):
    result = spcall("searchperson", (lastname, username, token,searchgroup))

    if len(result) == 0:
        return jsonify({
            "status": "ok",
            "size": 0
        })

    if 'Error' in result[0][0]:
        return jsonify({
            'status':'error',
            'message':result[0][0]
        })


    ret = []
    for item in result:
        stu = dict({
            "lrn": item[0],
            "fullname": item[1],
            "picurl": item[2]
        })
        ret.append(stu)
    return jsonify({
        "status": "ok",
        "size": len(ret),
        "list": ret
    })


@app.route("/faculty/<string:lastname>", methods=['GET'])
@auth.login_required
def searchlistfaculty(lastname):
    params = request.args
    token = params["token"]
    return searchoperson(lastname, auth.username(), token,'faculty')


@app.route("/classlist/<string:idnum>/<string:schoolid>/<string:semid>/<int:level>/<string:section>",
           methods=['DELETE'])
@auth.login_required
def removestudent(idnum, schoolid, semid, level, section):
    # removestudent('44', '128081', getcurrsem(), 2, 'aquamarine')
    username = auth.username()
    usergroup = spcall("getuser_group", (username,))[0][0]
    return jsonify(
        formatres(
            spcall("removestudent",
                   (idnum, schoolid, semid, level, section),
                   usergroup, True)
        )
    )

def stringifyQuarter(par_num):
    quarter = " QUARTER"
    if par_num == 1:
        return "FIRST" + quarter
    elif par_num == 2:
        return "SECOND" + quarter
    elif par_num == 3:
        return "THIRD" + quarter
    else:
        return "FOURTH" + quarter

@app.route("/classlist/spreadsheet", methods=["GET"])
@auth.login_required
def getclasslistspreadsheet():
    params = request.args
    username = auth.username()
    token = params["token"]
    section = params["section"]
    level = params["level"]
    schoolid = params["schoolid"]
    semid = params["semid"]
    teachername = params["teachername"] #id=name-leftbadge
    schoolname = params["schoolname"] #id=schoolname
    quarter = params["quarter"]
    subjectdesc = params["subjectdesc"]#apputils.getSubjectDescfromOffID($("#name-rightbadge").data("sectionid"))

    genderize = spcall("getclasslistbygender",
                       (
                           username, "faculty", token, section,
                           level, schoolid, semid
                       ), "faculty", True)

    if 'Error' in genderize[0][0]:
        return jsonify(
            {
                "status":"error",
                "message":genderize[0][0]
             }
        )

    wb = Workbook()
    ws = wb.active
    path = "static/downloads"
    file = section+level+schoolid + semid + "ECRECORD" + ".xlsx"

    #try:
    #    fp = open(os.path.join(path, file))
    #    return jsonify({"status": "ok"})
    #except Exception, e:
    #    pass
    file_location = os.path.join(path, file)
    if os.path.exists(file_location):
        os.remove(file_location)


    mergeAlign(ws, ["Class Record"], "A1:AJ1", centerCellText) #A1:AJ1 Class Record
    ws.append([""])
    ws.append([""])
    ws.append(["", "", "REGION", genderize[0][0]["region"],
               "DIVISION", genderize[0][0]["division"],
               "DISTRICT", genderize[0][0]["district"]])
    ws.append(["", "", "SCHOOL NAME", schoolname,
               "SCHOOL ID", schoolid, "SCHOOL YEAR", semid
               ])
    ws.append([""])

    ws.append(["", "", stringifyQuarter(int(quarter)), "", "", "",
                   "GRADE & SECTION:", level + "-" + section, "", "", "", "","", "", "", "","",
                   "TEACHER", teachername, "", "", "", "","", "", "", "","", "", "", "",
                   "SUBJECT", subjectdesc
                   ])
    ws.append(['', "LEARNERS' NAMES", '', '', '',
               "WRITTEN WORKS"
               ] + [""] * 12 +
              ["PERFORMANCE TASKS"] + [""] * 12 +
              ["QUARTERLY ASSESSMENT", "", ""] +
              ["INITIAL", "QUARTER"]
              )
    i = 1
    row = []
    for cols in list(letter_range('f', 'p')):
        row.append(i)
        i = i + 1
    row.append("Total")
    row.append("PS")
    row.append("WS")
    i = 1
    for cols in list(letter_range('s', '{')):
        #mergeAlign(ws, [i], cols.upper() + "9", centerCellText)
        row.append(i)
        i = i + 1
    row = row + ["9","10", "Total", "PS", "WS", "1", "PS", "WS", "Grade", "Grade"]
    ws.append([""] * 5 + row)
    mergeAlignOnly(ws, "F8:R8", centerCellText)
    mergeAlignOnly(ws, "S8:AE8", centerCellText)
    mergeAlignOnly(ws, "AF8:AH8", centerCellText)
    ws.append(["", "HIGHEST POSSIBLE SCORE"])
    ws.append(["", "MALE"])
    i = 1
    row = 12
    for mstu in genderize[0][0]["malelist"]:
        ws.append([i, mstu["name"]])
        mergeAlignOnly(ws, "B" + str(row) + ":" + "E" + str(row), leftCellText)
        i = i + 1
        row = row + 1
    ws.append(["", "FEMALE"])
    row = row + 1
    for fstu in genderize[0][0]["femalelist"]:
        ws.append([i, fstu["name"]])
        mergeAlignOnly(ws, "B" + str(row) + ":" + "E" + str(row), leftCellText)
        i = i + 1
        row = row + 1

    wb.save(path + "/" + file)

    return jsonify({"status": "ok"})
    #current


@app.route("/faculty/<string:fromfacultyid>/<string:tofacultyid>/<string:offeringid>", methods=['POST'])
@auth.login_required
def loadfaculty(fromfacultyid, tofacultyid, offeringid):
    params = request.get_json()
    token = params["token"]
    quarter = params["quarter"]
    usergroup = params["group"]
    username = auth.username()
    #usergroup = spcall("getuser_group", (username,))[0][0]


    #print (offeringid, fromfacultyid, tofacultyid, username, token, usergroup)
    res = formatres(
        spcall("designate_subject",
               (offeringid, fromfacultyid, tofacultyid, username, token, usergroup, quarter),usergroup, True)
    )

    if res["status"] == 'Error':
        return jsonify(res)

    update = "yes"

    if res["item"][0].split('->')[0] == 'OKO':
        update = "no"

    return jsonify({
        "status": "ok",
        "update": update,
        "pic": res["item"][0].split('->')[1]
    })


@app.route("/faculty/<string:facultyid>/<string:offeringid>", methods=['DELETE'])
@auth.login_required
def removeloadfaculty(facultyid, offeringid):
    params = request.get_json()
    token = params["token"]
    usergroup = params["group"]
    quarter = params["quarter"]
    username = auth.username()
    res = formatres(
        spcall("removeload",
               (username, facultyid, offeringid, token, usergroup, quarter),
               usergroup, True)
    )

    if res["status"] == "Error":
        return jsonify(res)

    return jsonify(
        {
            "status": "ok"
        })


def coteacher(username, offeringid):
    usergroup = spcall("getuser_group", (username,))[0][0]
    adviserid = spcall("getidnumfromusername", (username, usergroup),
                       usergroup)[0][0]
    teachers = spcall("getcoteacher",
                      (adviserid, offeringid),
                      usergroup)
    pic = fileServer()

    coteachers = []
    for teacher in teachers:
        coteachers.append({
            "pic": teacher[2],
            "fullname": teacher[0],
            "idnum": teacher[1]
        })

    return {
        'status': 'ok',
        'coteachers': coteachers
    }


@app.route("/classlist/<string:idnum>/<string:schoolid>/<string:semid>/<int:level>/<string:section>/<string:reason>",
           methods=['DELETE'])
@auth.login_required
def dropstudent(idnum, schoolid, semid, level, section, reason):
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    username = auth.username()
    usergroup = group#spcall("getuser_group", (username,))[0][0]

    receiveridnum = formatres(
        spcall("getpersonidbyidnum",
               (idnum,)))

    if receiveridnum["status"] == 'Error':
        return jsonify(receiveridnum)

    result = formatres(
        spcall("removestudent",
               (idnum, schoolid, semid,
                level, section.upper(), username, token, usergroup),
               usergroup, True))

    if result["status"] == 'Error':
        return jsonify(result)


    result = savetimeline(username,
                          idnum,
                          reason,
                          semid,
                          "drop",
                          2, schoolid)

    if result["status"] == 'Error':
        return jsonify(result)

    #print result
    return jsonify(result)


@app.route(
    "/classlist/<string:idnum>/<string:schoolid>/<string:semid>/<int:fromlevel>/<string:fromsection>/<int:tolevel>/<string:tosection>/<string:reason>",
    methods=['PUT'])
@auth.login_required
def reassignstudent(idnum, schoolid,
                    semid, fromlevel, fromsection,
                    tolevel, tosection, reason):
    params = request.get_json()
    token = params["token"]
    group = params["group"]

    try:
        username = auth.username()
        usergroup = group
        #usergroup = spcall("getuser_group", (username,))[0][0]
        receiveridnum = formatres(
            spcall("getpersonidbyidnum",
                   (idnum,)))

        tosection = tosection.upper()

        if receiveridnum["status"] == 'Error':
            return jsonify(receiveridnum)


        result = formatres(spcall("removestudent",
                                  (idnum, schoolid, semid,
                                   fromlevel, fromsection, username, token, usergroup),
                                  usergroup, True))

        if result["status"] == 'Error':
            return jsonify(result)



        result = savetimeline(username,
                              idnum,
                              reason,
                              semid,
                              're-assign',
                              2, schoolid)

        if result["status"] == 'Error':
            return jsonify(result)

        return jsonify(
            formatres(
                spcall("enrollstudent",
                       (
                           idnum, schoolid,
                           semid, tolevel,
                           tosection.upper(), username,
                           token, usergroup),
                       usergroup, True)))
    except:
        return jsonify({
            "status": "error",
            "message": str(sys.exc_info()[0]) + " " + str(sys.exc_info()[1])
        })


@app.route(
    "/classlist/<string:idnum>/<string:schoolid>/<string:semid>/<int:level>/<string:section>/<string:toschool>/<string:reason>",
    methods=['PUT'])
@auth.login_required
def transferstudent(idnum,
                    schoolid,
                    semid,
                    level,
                    section,
                    toschool,
                    reason):
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    username = auth.username()
    usergroup = group #spcall("getuser_group", (username,))[0][0]

    result = formatres(
        spcall("removestudent",
               (idnum, schoolid, semid,
                level, section.upper(), username, token, usergroup),
               usergroup, True))

    if result["status"] == 'Error':
        return jsonify(result)

    return jsonify(savetimeline(username,
                                idnum,
                                reason,
                                semid,
                                'transfer',
                                2, schoolid))


def getclasslistbygender(username, schoolid, semid, grade, section):
    try:
        usergroup = spcall("getuser_group", (username,))[0][0]
        idnum = spcall("getidnumbyuser", (digest(username), usergroup))[0][0]
        classlist = spcall("getclasslist",
                           (section, grade, schoolid, semid))

        total = len(classlist)

        if total == 0:
            return {"status": "ok", "size": 0,
                    "male": 0, "female": 0}

        male = 0
        female = 0
        malelist = []
        femalelist = []
        f = fileServer()
        for item in classlist:
            fullname = item[2].split("*")

            stu = dict(
                {
                    "name": fullname[0] + ", " + fullname[1] + " " + fullname[2],
                    "id": item[0]
                    # "color":color,
                    # "imgsrc": imgsrc
                })
            res = f.getImageUrl(spcall("getpersonidbyidnum", (item[0],))[0][0])
            if item[3] == 'F':
                female = female + 1
                color = 'bg-maroon-active'
                imgsrc = res
                stu["color"] = color
                stu["imgsrc"] = imgsrc
                femalelist.append(stu)
            else:
                color = 'bg-aqua-active'
                male = male + 1
                imgsrc = res
                stu["color"] = color
                stu["imgsrc"] = imgsrc
                malelist.append(stu)

        return dict({"status": "ok",
                     "size": total,
                     "malelist": malelist,
                     "femalelist": femalelist,
                     "male": male,
                     "female": female
                     })

    except:
        return {
            "status": "error",
            "message": str(str(sys.exc_info()[0]) + " " \
                           + str(sys.exc_info()[1]))
        }


def gendercolor(gender):
    if gender == 'F':
        color = 'bg-maroon-active'
    else:
        color = 'bg-aqua-active'
    return color


def getclasslist(username, schoolid, semid, grade, section, token):
    try:
        return dict(
            spcall('getclasslistjson',
                   (section, grade, schoolid, semid, username, token)
                   )[0][0]
        )

    except:
        return {
            "status": "error",
            "message": str(str(sys.exc_info()[0]) + " " \
                           + str(sys.exc_info()[1]))
        }


@app.route("/classlist/<string:schoolid>/<string:semid>/<int:grade>/<string:section>/<string:token>", methods=['GET'])
@auth.login_required
def getclasslistjson(schoolid, semid, grade, section, token):
    return jsonify(
        getclasslist(
            auth.username(),
            schoolid,
            semid,
            grade,
            section,
            token
        ))


@app.route("/classlist/<string:sectionid>/<string:token>", methods=['GET'])
@auth.login_required
def getclasslistgender(sectionid, token):
    username = auth.username()
    lst = spcall("getclasslistgendersectioncode", (sectionid, username, token))[0][0]
    if 'Error' in lst:
        return jsonify({
            'status':"error",
            'message': lst
        })

    return jsonify(lst)
    # group = spcall("getuser_group", (username,))[0][0]
    # offeringdetails = spcall("getofferingdetails", (sectionid,), group)
    #
    # semid = offeringdetails[0][0]
    # schoolid = offeringdetails[0][1]
    # level = offeringdetails[0][2]
    # section = offeringdetails[0][3]
    #
    # res = getclasslistbygender(username, schoolid, semid, level, section)
    #
    # if res["status"] != 'ok':
    #     return jsonify({
    #         "status": res["status"],
    #         "message": res["message"],
    #         "listmale": [],
    #         "listfemale": []
    #     })
    #
    # if res["status"] == 'ok' and res["size"] > 0:
    #     listmale = res["malelist"]
    #     listfemale = res["femalelist"]
    #
    # return jsonify({
    #     "status": res["status"],
    #     "listmale": listmale,
    #     "listfemale": listfemale
    # })

@app.route("/classlist/<string:schoolid>/<string:semid>/<string:level>/<string:section>", methods=['POST'])
@auth.login_required
def enrollfromxlsx(schoolid, semid, level, section):
    params = request.get_json()
    username = auth.username()
    token = params["token"]
    group = params["group"]
    lrn = str(params["lrn"])
    lastname= params["lastname"]
    firstname= params["firstname"]
    middlename= params["middlename"]
    sex= params["sex"]
    birthdate= params["birthdate"]
    age= params["age"]
    mothertongue= params["mothertongue"]
    ip= params["ip"]
    religion= params["religion"]
    housenum= params["housenum"]
    brgy= params["brgy"]
    mun= params["mun"]
    province= params["province"]
    zipcode= params["zipcode"]
    flname= params["flname"]
    ffname= params["ffname"]
    fmname= params["fmname"]
    mlname= params["mlname"]
    mfname= params["mfname"]
    mmname= params["mmname"]
    glname= params["glname"]
    gfname= params["gfname"]
    gmname= params["gmname"]
    condcashtrans= params["condcashtrans"]
    nutionalstatus= params["nutionalstatus"]
    disabilities= params["disabilities"]
    # print token, group, lrn, lastname,firstname,middlename,sex,birthdate,age,\
    #     mothertongue,ip,religion,housenum,brgy,mun,province,zipcode,\
    #     flname,ffname,fmname,mlname,mfname,mmname,glname,gfname,\
    #     gmname,condcashtrans,nutionalstatus,disabilities, schoolid, semid, level, section
    result = spcall("enrollfromxls", (username,token, group, schoolid, semid, lrn, section, level,
                                      lastname, firstname, middlename, sex, birthdate, age, mothertongue,
                                      ip, religion, housenum, brgy, mun, province, zipcode, flname, ffname, fmname,
                                      mlname, mfname, mmname, glname, gfname, gmname, condcashtrans, nutionalstatus, disabilities
                                      ), group, True)
    return jsonify(formatres(result))


@app.route("/classlist/<string:schoolid>/<string:semid>/<int:level>/<string:section>/<string:lrn>", methods=['POST'])
@auth.login_required
def enrollstudent(schoolid, semid, level, section, lrn):
    # select enrollstudent('xcabit', '128081', getcurrsem(), 2, 'AQUAMARINE');
    params = request.get_json()
    username = auth.username()
    token = params["token"]
    group = params["group"]
    #group = spcall("getuser_group", (username,))[0][0]
    result = spcall("enrollstudent",
                    (lrn, schoolid, semid, level, section, username, token, group),
                    group, True)
    level = ztok(level)

    if level != 'Kinder':
        level = "Grade " + str(level)

    timeres = spcall("stampevent", (username, lrn,
                                    "Enrolled in " +
                                    level +
                                    " Section " + section + " @ " + spcall("getschoolname", (schoolid,))[0][0],
                                    'enroll', 2, semid, schoolid), "super", True)
    return jsonify(formatres(result))


@app.route("/student/<string:fields>/<string:schoolid>", methods=['POST'])
@auth.login_required
def register(fields, schoolid):
    username = auth.username()
    group = spcall("getuser_group", (username,))[0][0]
    loc_argsful = fields.split("*")
    params = request.get_json()
    oldlrn = params["oldlrn"]

    # return jsonify({"ret": fields})
    if len(loc_argsful) != 29:
        return jsonify({'status': 'error',
                        'message': 'Unexpected Input',
                        'item': 'None'})

    condcashtransfer = loc_argsful.pop(24).upper()
    healthstatus = loc_argsful.pop(24).upper()
    pwd = loc_argsful.pop(24).upper()

    '''
     return jsonify({'status': 'error',
                     'message': condcashtransfer + ' ' +
                                healthstatus + ' ' +
                                pwd})
     '''

    username = auth.username()
    usergroup = spcall("getuser_group", (username,))[0][0]

    # passed value
    # lrn no.*lastname*firstname*m.*Male*12/01/1978*
    # automatic*Cebuano*Visaya*Roman Catholic*house no*
    # brgy*municipal*provice*9200*flname*ffname*
    # fmi*mlname*mfname*mmi*glname*gfname*gmi""

    age = loc_argsful.pop(6)
    loc_argsful.insert(4, loc_argsful.pop(4)[0:1].upper())  # change Male to M
    # remove age
    # age will be computed again inside the
    # stored proc
    # age was passed as part of the parameter
    # because it was part when form is
    # processed

    # processed value
    # 0-lrn no.*1-lastname*2-firstname*3-m.*4-Male*5-12/01/1978
    # *6-Cebuano*7-Visaya*8-Roman Catholic*9-house no*
    # 10-brgy*11-municipal*12-provice*13-9200*14-flname*15-ffname*
    # 16-fmi*17-mlname*18-mfname*19-mmi*20-glname*21-gfname*22-gmi""
    loc_argsful.insert(3, cleandat(loc_argsful.pop(3), ".", "").upper())
    loc_argsful.insert(16, cleandat(loc_argsful.pop(16), ".", "").upper())
    loc_argsful.insert(19, cleandat(loc_argsful.pop(19), ".", "").upper())
    loc_argsful.insert(22, cleandat(loc_argsful.pop(22), ".", "").upper())
    loc_argsful.insert(5, cleandat(loc_argsful.pop(5), "-", "/"))
    # return jsonify({"d":loc_argsful})
    grade = loc_argsful.pop(23)
    section = loc_argsful.pop(23)
    loc_argsful.append('students')
    loc_argsful.append('')
    loc_argsful.append(username)
    loc_argsful.append(schoolid)
    loc_argsful.append('students')
    loc_argsful.append(grade)
    loc_argsful.append(section)
    loc_argsful.append(spcall("getschoolcourse", (schoolid,), usergroup)[0][0])

    '''
    return jsonify({'status': 'error',
                     'message': loc_argsful})
    idnum = loc_argsful[0]
    condcashtransfer = loc_argsful.pop(24).upper()
    healthstatus = loc_argsful.pop(24).upper()
    pwd = loc_argsful.pop(24).upper()
    '''

    idnum = loc_argsful[0]

    if len(oldlrn) > 0:
        updatelrn = spcall("updatepersonnum",
                           (oldlrn, idnum, "students"),
                           group, True)
        # if formatres(updatelrn)["status"] <> "Error":
        #    idnum =

    # return jsonify({"items":loc_argsful})
    items = []
    items.append(
        tuple(
            map(lambda x: x, loc_argsful)
        ))  # returns an array of tuple
    #return jsonify({"status":"error","items": items[0]})
    #insup2studentinfo(par_lrn, par_semid, par_nutionalstatus, par_condcashtrans, par_disabilities);
    result = spcall("inspersondetails", items[0], usergroup, True)

    if formatres(result)["status"] == "Error":
        return jsonify({"status": "error", "message": formatres(result)["message"]})

    result = spcall("insup2studentinfo", (idnum,
                                          spcall("getcurrsem", (),
                                                 usergroup)[0][0],
                                          healthstatus,
                                          condcashtransfer,
                                          pwd), usergroup, True)

    stuid = spcall("idgenerator", ('students', username))[0][0]
    if 'Error' in result[0][0] or \
            'function' in result[0][0] or \
            'error' in result[0][0]:

        status = 'Error'
        message = result[0][0]
    else:
        status = 'ok'
        message = 'ok'
        timeres = spcall("stampevent", (username, idnum,
                                        "Details encoded  @ " + spcall("getschoolname", (schoolid,))[0][0],
                                        'register', 2, spcall("getcurrsem", (), group)[0][0], schoolid), "super", True)

    return jsonify({'status': status,
                    'message': message,
                    'item': result[0],
                    'freeid': stuid})


def getreligionoptions():
    religions = spcall("getreligion", ())

    out = []
    for religion in religions:
        out.append(religion[0])

    return out


def composer(fields, rows):
    result = []
    for row in rows:
        dic = dict()
        i = 0
        for field in fields:
            dic[field] = row[i]
            i = i + 1
        result.append(dic)
    return result


def grade2kinder(grade):
    if grade == 0:
        return "Kinder"
    else:
        return grade

def syslogcred():
    try:
        username = auth.username()
        #return jsonify({'status':'error', 'message': username})
        credentials = spcall("login_credentials", (username,), "super", True)
        #return formatres(credentials)
        jsonifycred = formatres(credentials)
        pic = fileServer()
        if jsonifycred["status"] != 'ok':
            return {"status": "error", "message": jsonifycred["item"]}

        credentials = jsonifycred["item"][0]
        # print(f'\n\nCredentials: {credentials}\n\n')
        userdetails = credentials[u"userdetails"].split("*")
        # print(f'\n\nUser Details: {userdetails}\n\n')
        schoolassin = credentials[u"userschool"].split("*")
        # print(f'\n\nSchool Assign: {schoolassin}\n\n')
        #print credentials[u"designated"]

        channel = credentials[u"vroomid"]
        channels = []
        channels.clear()
        if channel != 'NONE':
            channels.append(channel)

        designated = []

        
        
        if credentials[u"designated"] != 'none':
            for c in credentials[u"designated"].split(",")[:-1]:
                #print c
                designatedarr = c.split("*")
                designated.append({"grade": designatedarr[0].replace('none', ''), "section": designatedarr[1]})
        
        # religionoptions is just an array
        load = []
        if credentials[u'load'] != 'none':
            for loaditem in credentials[u'load'].split(',')[:-1]:
                loadinfo = loaditem.split("*")

                coteachers = []
                if len(loadinfo[7]) > 0:
                    for coteacheritem in loadinfo[7].split('@'):
                        if len(coteacheritem) > 0:
                            coteacherinfo = coteacheritem.split("-")
                            coteachers.append({
                                "fullname": coteacherinfo[1].replace("^", ","),
                                "idnum": coteacherinfo[2],
                                "pic": coteacherinfo[0]
                            })

                subjectinfo = [
                    loadinfo[0].replace('none', '') + ' ' + loadinfo[5],
                    loadinfo[1],
                    loadinfo[2],
                    loadinfo[3],
                    loadinfo[4],
                    loadinfo[5],
                    loadinfo[6],
                    {
                        "coteachers": coteachers,
                        "status": "ok"
                    }]

                load.append(subjectinfo)

        nutritionalstatus = []
        for nitems in credentials[u"nstatus"]:
            nutritionalstatus

        mystu = []

        if credentials[u"usertype"] == 'parents':

            kids = credentials[u"mystu"].split("$")[:-1]
            for kid in kids:
                kid = kid.split("*")
   
                channels.append(kid[5])
                mystu.append({
                    "id": kid[0],
                    "name": kid[2],
                    "gender": kid[3],
                    "color": gendercolor(kid[3]),
                    "imgsrc": pic.getImage(kid[1]),
                    "btn": "{;$.cookie('schoolid', '" + kid[4] + "');" +
                           "model.profile('" + kid[0] + "', 'btn" + kid[0] + "', ''); " +
                           "$('#name-rightbadge').data('timelrn', '" + kid[0] + "');" +
                           "$('#name-rightbadge').data('offset', 0);" +
                           "}"
                })
        todayis = date.today().strftime("%d %B %Y")
       
        
        personnumid = credentials[u"personnumid"]
        clean_cred =  {"status": "ok", "token": credentials[u"token"], "usertype": credentials[u"usertype"],
                "userdetails": {"name": userdetails[0], "position": userdetails[1]},
                "userschool": {"id": schoolassin[0],
                               "name": schoolassin[1],
                               "region": schoolassin[2],
                               "division": schoolassin[3],
                               "district": schoolassin[4]
                               },
                "designated": designated,
                "religions": credentials[u"religions"].split(","),
                "imgsrc": pic.getImgSrcUrl(credentials[u"imgsrc"]),
                "date": todayis,
                "semid": credentials[u"semid"],
                "freeid": credentials[u"freeid"],
                "load": load,
                "nstatus": credentials[u"nstatus"].split(",")[:-1],
                "quarter": credentials[u"quarter"],
                "lrn": credentials[u"lrn"],
                "mystu": mystu
                }
        if credentials[u"usertype"] == 'admin' or credentials[u"usertype"] == 'faculty':
            channels.append(schoolassin[0])
            if credentials[u"usertype"] == 'faculty':
                channels.append(personnumid)

       
        # print(f'\n\nCredentials: {clean_cred}\n\n')

        return {"status": "ok", "token": credentials[u"token"], "usertype": credentials[u"usertype"],
                "userdetails": {"name": userdetails[0], "position": userdetails[1]},
                "userschool": {"id": schoolassin[0],
                               "name": schoolassin[1],
                               "region": schoolassin[2],
                               "division": schoolassin[3],
                               "district": schoolassin[4]
                               },
                "designated": designated,
                "religions": credentials[u"religions"].split(","),
                "imgsrc": pic.getImgSrcUrl(credentials[u"imgsrc"]),
                "date": todayis,
                "semid": credentials[u"semid"],
                "freeid": credentials[u"freeid"],
                "load": load,
                "nstatus": credentials[u"nstatus"].split(",")[:-1],
                "quarter": credentials[u"quarter"],
                "lrn": credentials[u"lrn"],
                "mystu": mystu,
                "virtualroomid": channels,
                "personnumid": personnumid
                }
    except:
        return {
            "status": "error",
            "message": str(str(sys.exc_info()[0]) + " " +
                           str(sys.exc_info()[1]) + " " +
                           str(sys.exc_info()[2]))
        }


def validsess(token):
    # a,b,c
    sysver = syslogcred()
    return token == sysver["token"]


@app.route("/auth", methods=['GET'])
@auth.login_required
def user_auth():
    #return({'status':'error', 'message':auth.username()})
    return jsonify(syslogcred())


@app.route("/auth/<string:section>/<string:grade>/<string:schoolyear>", methods=['POST'])
@auth.login_required
def enabale_all_account(section, grade, schoolyear):
    username = auth.username()
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    schoolnumber = params["schoolid"]
    wb = Workbook()
    ws = wb.active
    path = "static/downloads"
    file = section+grade+schoolnumber+schoolyear + '.xlsx'

    #create file only once, the rest will be encoded in the system.
    #print os.path.join(path, file)
    #try:
    #    fp = open(os.path.join(path, file))
    #    return jsonify({"status": "ok", "accounts": {"accounts":[]}})
    #except Exception, e:
    #    pass #print "File Access Error:" + str(e)

    result = spcall("enable_all_accounts",
                    (username,
                     token,
                     group,
                     int(grade),
                     section,
                     schoolyear,
                     schoolnumber),
                    "super", True)


    if 'Error' in result[0][0]:
        return jsonify({
            'status':"error",
            "message": result[0][0]
        })

    return jsonify({"status": "ok", "accounts": result[0][0]})
    '''
    mergeAlign(ws,
               ["Auto Generated Usernames "
                "and Passwords for Grade "
                + grade +" Section " +
                section],
               "A1:H1",
               centerCellText
               )
    ws.append(["Student", "", "Father", "", "Mother", "", "Guardian", ""])
    mergeAlignOnly(ws, "A2:B2", centerCellText)
    mergeAlignOnly(ws, "C2:D2", centerCellText)
    mergeAlignOnly(ws, "E2:F2", centerCellText)
    mergeAlignOnly(ws, "G2:H2", centerCellText)
    ws.append(["Username", "Password"] * 4)



    for account in result[0][0]["accounts"]:
        student = account["student"].split("->")
        father = account["father"].split("->")
        mother = account["mother"].split("->")
        guardian = account["guardian"].split("->")

        if len(student) == 1:
            student.append("")

        if len(father) == 1:
            father.append("")

        if len(mother) == 1:
            mother.append("")

        if len(guardian) == 1:
            guardian.append("")

        ws.append(student + father + mother + guardian)

    wb.save(path + "/" + file)
    '''



@app.route("/auth", methods=['POST'])
@auth.login_required
def activate_account():
    username = auth.username()
    params = request.get_json()
    personnumid = params["lrn"]
    # email = params["email"]
    semid = params["semid"]
    lastname = params["lastname"]
    firstname = params["firstname"]
    middlename = params["middlename"]
    group = params["group"]
    token = params["token"]

    result = spcall("create_user_account", (username,
                                            personnumid,
                                            semid,
                                            lastname,
                                            firstname,
                                            middlename,
                                            group, token), "super", True)

    result = formatres(result)
    status = result["status"].upper()
    credentials = {
        "username": '',
        "password": ''
    }

    if status == 'OK':
        credentials = {
            "username": result["item"][0]["credentials"]["username"],
            "password": result["item"][0]["credentials"]["password"]
        }

    return jsonify({
        "status": status,
        "credentials": credentials,
        "message": result["message"]
    })


@app.route("/auth", methods=["PUT"])
@auth.login_required
def change_password():
    params = request.get_json()
    p1 = params["password1"]
    p2 = params["password2"]
    token = params["token"]
    group = params["group"]

    res1 = checkpassword(p1)

    if len(res1) > 0:
        return jsonify({
            "status": "error",
            "message": res1
        })

    res2 = checkpassword(p2)

    if len(res2) > 0:
        return jsonify({
            "status": "error",
            "message": res2
        })

    if p1 != p2:
        return jsonify({
            "status": "error",
            "message": "Password do not match."
        })

    username = auth.username()
    #group = spcall("getuser_group", (username,))[0][0]
    respass = spcall("updatepassword", (username, p1, token, group), group, True)

    if formatres(respass)["status"] != "ok":
        return jsonify({"status": "error", "message": respass})

    return jsonify({"status": "OK", "message": "Password Successfully Changed"})


@app.route("/exit", methods=['GET'])
@auth.login_required
def sysout():
    username = auth.username()

    return restrequest("sysout",
                       (
                           digest(username),
                       ),
                       spcall,
                       "super",
                       True,
                       jsonify
                       )




def savetimeline(username, receiverid, message, semid, tltype, publicity, schoolid):
    group = spcall("getuser_group", (username,))[0][0]
    initiator = formatres(spcall('getpersonidbyusername', (username,)))
    if initiator["status"] == 'Error':
        return jsonify(initiator)

    initiatorid = initiator["item"][0]

    receiver = formatres(
        spcall(
            "getpersonidbyidnum",
            (receiverid,)))

    if receiver["status"] == 'Error':
        return receiver

    receiverid = receiver["item"][0]

    if receiverid == 'NO':
        return receiver

    return formatres(
        spcall("insert2timeline",
               (
                   initiatorid,
                   receiverid,
                   message,
                   tltype,
                   publicity,
                   spcall("now", ())[0][0],
                   semid,
                   schoolid
               ), group, True)
    )


# @app.route("/timeline/<string:receiverid>/<string:message>/<string:tltype>/<string:publicity>",  methods=['POST'])
# @auth.login_required
# def savetimelineapi(receiverid, message, semid, tltype, publicity):
#    username = auth.username()
#    return jsonify(
#        savetimeline(
#            username,
#            receiverid,
#            message,
#            semid,
#            tltype,
#            publicity))

# def getload(username, quarter, schoolid, semid, token):
#     group = spcall("getuser_group", (username,))[0][0]
#     activecurriculum = formatres(
#         spcall("getactivecurriculum", (), group)
#     )
#
#     if activecurriculum["status"] == 'Error':
#         return activecurriculum
#
#     personnumid = formatres(
#         spcall("getidnumbyuser", (digest(username), group), group)
#     )
#
#     if personnumid["status"] == 'Error':
#         return personnumid
#
#     load = spcall("getload", (personnumid["item"][0],
#                               schoolid,
#                               str(quarter), semid,
#                               activecurriculum["item"][0]), group)
#
#     # return {"status": "ok", "load": load}
#     # if formatres(load)["status"] == "Error":
#     #    return formatres(load)
#
#     load = map(lambda x: list(x), load)  # convert tuple to list in order to add co-teacher
#     if len(load) > 0:
#         i = 0
#         for iload in load:
#             result = coteacher(username, iload[4])
#             # if result["status"] == 'ok':
#             load[i].append(result)
#             i = i + 1
#
#     # quarter = '1' #formatres(spcall("getactivequarter",
#     #             (section,level, semid, schoolid), group, True))
#
#     # if quarter["status"] == "Error":
#     #    return quarter
#
#     advisee = spcall("getadvisee", (digest(username), schoolid, semid), group)
#     if advisee != []:
#         level = advisee[0][0]
#         section = advisee[0][1]
#     else:
#         level = '1'
#         section = ''
#
#     res = getclasslistbygender(username, schoolid, semid, level, section)
#
#     quarter = formatres(spcall("getactivequarter",
#                                (section, level, semid, schoolid), group, False))
#
#     if quarter["status"] == "Error":
#         return quarter
#
#     if res["status"] == 'ok' and res["size"] > 0:
#         listmale = res["malelist"]
#         listfemale = res["femalelist"]
#     else:
#         listmale = []
#         listfemale = []
#
#     return {
#         'status': 'ok',
#         'load': load,
#         'size': len(load),
#         'quarter': quarter["item"][0],
#         'listmale': listmale,
#         'listfemale': listfemale
#     }

@app.route("/load/<string:schoolid>/<string:semid>/<string:token>", methods=['GET'])
@auth.login_required
def getrestload2(schoolid, semid,token):
    username = auth.username()
    params = request.args
    sem = params["semester"]
    load = spcall("getloadjsonfromidnumber", (username, schoolid, semid, token, sem))[0][0]
    if 'Error' in load:
        return jsonify({"status":"error", "message": load})

    processedload = []

    if load["size"] > 0:
        for item in load["load"]:
            strippenitem = item.strip('{}').split(",")
            coteachers_count = int(strippenitem[-1])
            coteachers_array = []
            makearray = []
            i = 1
            for stritem in strippenitem:
                if i <= 6:
                    makearray.append(stritem.strip('"'))
                    i = i +1

            adder = 0
            for co_i in range(1, coteachers_count+1):
                if strippenitem[adder + 10] == u'""':
                    picurl = ''
                else:
                    picurl = strippenitem[adder + 13]

                coteachers_array.append({
                    "fullname": strippenitem[adder + 10] + "," + strippenitem[adder + 11],
                    "idnum": strippenitem[adder + 12],
                    "pic": picurl
                })
                adder = adder + 4

            makearray.append({
                "coteachers": coteachers_array
            })
            processedload.append(makearray)


    return jsonify(
        {
            "status": load["status"],
            "load": processedload,
            "size": load["size"]
        })

'''
@app.route("/load/<int:quarter>/<string:section>/<int:yearlevel>/<string:schoolid>/<string:semid>/<string:token>", methods=['GET'])
@auth.login_required
def getloadpersection(quarter, section, yearlevel, schoolid, semid,token):
    
    #getloadjsonyearsection(
    #par_username text,
    #par_quarter int,
    #par_section text,
    #par_level int,
    #par_schoolid text,
    #par_semid text,
    #par_token text)
    
    username = auth.username()
    load = spcall("getloadjsonyearsection", (username, quarter, section,
                                             yearlevel, schoolid, semid, token))[0][0]
    if 'Error' in load:
        return jsonify({"status":"error", "message": load})

    processedload = []

    if load["size"] > 0:
        for item in load["load"]:
            strippenitem = item.strip('{}').split(",")
            makearray = []
            i = 1
            for stritem in strippenitem:
                if i <= 6:
                    makearray.append(stritem.strip('"'))
                    i = i +1
            makearray.append({
                "coteachers": [{
                    "fullname": strippenitem[6],
                    "idnum": strippenitem[7] + "," + strippenitem[8],
                    "pic": strippenitem[9]
                }]
            })
            processedload.append(makearray)


    return jsonify(
        {
            "status": load["status"],
            "load": processedload,
            "size": load["size"]
        })

'''
@app.route("/load/<int:quarter>/<string:schoolid>/<string:semid>/<string:token>", methods=['GET'])
@auth.login_required
def getrestload(quarter, schoolid, semid,token):
    '''
    Use this route only on current user, other routes will be used for anybody wishing to
    retrieve other load.
    :param section:
    :param level:
    :param quarter:
    :param schoolid:
    :return:
    '''
    username = auth.username()

    # if group == 'faculty':
    #    classlist = getclasslist(username, schoolid, semid, level, section)
    # else:
    #    classlist = []

    load = spcall("getloadjson", (username, quarter, schoolid, semid, token))[0][0]
    if 'Error' in load:
        return jsonify({"status":"error", "message": load})

    processedload = []

    if load["size"] > 0:
        for item in load["load"]:
            strippenitem = item.strip('{}').split(",")
            makearray = []
            i = 1
            for stritem in strippenitem:
                if i <= 6:
                    makearray.append(stritem.strip('"'))
                    i = i +1
            makearray.append({
                "coteachers": [{
                    "fullname": strippenitem[6],
                    "idnum": strippenitem[7] + "," + strippenitem[8],
                    "pic": strippenitem[9]
                }]
            })
            processedload.append(makearray)


    return jsonify(
        {
            "status": load["status"],
            "load": processedload,
            "size": load["size"]
        })

def getusergroup(username):
    return spcall("getuser_group", (username,))[0][0]


@app.route("/gradebook/<string:par_offeringid>/" +
           "<string:par_entryname>/<float:par_maxscore>/" +
           "<string:par_semid>/<string:par_periodid>/<string:par_categoryid>",
           methods=['POST'])
@auth.login_required
def postentry(par_offeringid, par_entryname, par_maxscore, par_semid, par_periodid, par_categoryid):
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    #group = getusergroup(auth.username())
    if par_maxscore != 9999.00:
        sp = "insupentry"
        comm = True
    else:
        sp = "getentry"
        comm = False

    # return jsonify({"status":"error", "sp":sp})
    resentryid = formatres(
        spcall(sp, (par_offeringid,
                    par_entryname,
                    par_maxscore,
                    par_semid,
                    par_periodid,
                    par_categoryid, auth.username(), token, group),
               group, comm)
    )

    # return jsonify({"status":"error", "sp":resentryid})
    if resentryid["status"] == "Error":
        return jsonify(resentryid)

    entries = spcall("getentries",
                     (resentryid["item"][0].split("@")[0],), group)
    #print(resentryid)
    en3s = []
    for item in entries:
        en3s.append(
            (item[0], str(item[1]), str(item[2]))    #map(lambda x: str(x), item)
        )
    #print (entries)
    if sp == "insupentry":
        timeres = spcall("stampnostuconv", (auth.username(), spcall("getdefault", ())[0][0],
                                            "Set Max Score to " + str(par_maxscore) + " for " +
                                            spcall("getsubjectoff", (par_offeringid,))[0][0] +
                                            " " + par_categoryid + " Entry No " + par_entryname +  " for period " +
                                            str(par_periodid),
                                            'grade', 2), "super", True)


    return jsonify(
        {
            "status": "ok",
            "entryid": resentryid["item"][0].split("@")[0],
            "entries": en3s,
            "maxscore": resentryid["item"][0].split("@")[1]
        }

    )


def inputgrade(par_studentid, par_entryid, par_score, par_mult, par_username, par_gradeinfo, par_token, par_group, par_semid, par_sectioncode):
    #group = getusergroup(par_username)
    group = par_group
    gradeinfo = par_gradeinfo.split("*")
    #  gradeinfo[0],  # section
    #  gradeinfo[1],  # level
    #  gradeinfo[2],  # quarter
    #  gradeinfo[3],  # semid
    #  gradeinfo[4],  # schoolid
    #  gradeinfo[5],  # categoryid (ie., WRITTEN WORKS, etc.)
    #  gradeinfo[6].split(' ' + gradeinfo[0])[0],  # subject
    result = formatres(spcall("entrystuscore",
                              (par_studentid,
                               par_entryid,
                               par_score,
                               par_mult, par_username,
                               par_token, par_group,
                               par_semid, par_sectioncode,
                               gradeinfo[2], #quarter
                               gradeinfo[5]  #categoryid
                               ),
                              group, True))

    if result["status"] != 'Error':
        gradeinfo = par_gradeinfo.split("*")
        #OUT: [u'ELITE', u'12', u'1', u'20182019', u'XXX', u'Written Works', u'CONTEMPORARY 12 ELITE']

        timeres = spcall("stampevent", (par_username, par_studentid,
                                        "Encoded with a score of " + str(par_score) + ' to ' +
                                        spcall("getentryinfo", (par_entryid,))[0][0],
                                        'grade', 2, gradeinfo[3], gradeinfo[4]), "super", True)

        # computestugradepercat(par_lrn text, par_section text, par_level int, par_quarter_ text, par_semid text,
        #                 par_schoolid text, par_categoryid text, par_subject text, par_currid text)

        # graderes = spcall("computestugradepercat", (par_studentid,
        #                                             gradeinfo[0],  # section
        #                                             gradeinfo[1],  # level
        #                                             gradeinfo[2],  # quarter
        #                                             gradeinfo[3],  # semid
        #                                             gradeinfo[4],  # schoolid
        #                                             gradeinfo[5],  # categoryid (ie., WRITTEN WORKS, etc.)
        #                                             gradeinfo[6].split(' ' + gradeinfo[0])[0],  # subject
        #                                             spcall("getcurriculumfromsem",
        #                                                    (gradeinfo[3],))[0][0]),
        #                   group, True)
        # print graderes
        #feelingres = addfeeling(par_username, par_studentid, par_entryid, par_score)
        # return {"k":feelingres}
    return result


def addfeeling(par_username, par_studentid, par_entryid, par_score):
    initiatorid = spcall("getpersonidbyusername", (par_username,))[0][0]

    receiverid = spcall("getpersonidbyidnum", (par_studentid,))[0][0]
    timestamp = spcall("getrecenttlts", (receiverid,))[0][0]
    recordentryid = digest(str(timestamp) + str(initiatorid) + str(receiverid) + str(par_entryid) + str(par_score))

    # return {'D':timestamp, 'i':initiatorid, 'r':receiverid, 'rid':recordentryid, 'eid':par_entryid}

    result = \
    spcall("createrecordentry", (recordentryid, receiverid, initiatorid, par_entryid, str(timestamp),), "super", True)[
        0][0]

    return result


@app.route("/gradebook/<string:par_studentid>/" +
           "<string:par_entryid>/<float:par_score>/<float:par_mult>/" +
           "<string:par_gradeinfo>", methods=['POST'])
@auth.login_required
def studentscoreentry(par_studentid, par_entryid, par_score, par_mult, par_gradeinfo):
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    semid = params["semid"]
    sectioncode = params["sectioncode"]
    return jsonify(
        inputgrade(par_studentid, par_entryid, par_score, par_mult, auth.username(),
                   par_gradeinfo, token, group, semid, sectioncode))


@app.route("/gradebook/<string:par_entries>/<string:par_entryid>/<string:par_gradeinfo>",
           methods=['POST'])
@auth.login_required
def batchscoreentry(par_entries, par_entryid, par_gradeinfo):
    # lrn*score@
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    semid = params["semid"]
    sectioncode = params["sectioncode"]
    result = []
    #group = getusergroup(auth.username())
    for entry in par_entries.split("@"):
        stuid = entry.split("*")[0]
        score = entry.split("*")[1]
        item = inputgrade(stuid, par_entryid, score, 1.00, auth.username(),
                          par_gradeinfo, token, group, semid, sectioncode)

        result.append({
            "lrn": stuid,
            "result": item["item"][0]
        })

    return jsonify({
        "results": result,
        "status": "ok"
    })


def getgradeentries(par_subject, par_schoolid, par_semid,
                    par_quarter_, par_section, par_level, par_username):
    group = getusergroup(par_username)
    #result = formatres(spcall("computequartergrade", (par_subject, par_schoolid,
    #                                                  par_semid, par_quarter_,
    #                                                  spcall("getcurriculumfromsem", (par_semid))[0][0],
    #                                                  par_section,
    #                                                  par_level), group, True))
    #print 'result'
    #print result

    # if result["status"] == 'Error':
    #    return result
    xres = spcall("getsummativequartergrade",
                  (par_subject, par_semid, par_section,
                   par_level, par_schoolid, par_quarter_),
                  group, False)

    if len(xres) == 0:
        return {"status": "ok", "grades": []}

    result = formatres(xres)

    if result["status"] == 'Error':
        return result

    objectified = []
    for item in xres:
        objectified.append({
            "lrn": item[0],
            "grade": str(item[1]),
            "finalized": item[3],
            "published":item[4]
        })

    return {"status": "ok",
            "grades": objectified}


@app.route("/gradebook/<string:par_subject>/<string:par_schoolid>/<string:par_semid>/" +
           "<string:par_quarter_>/<string:par_section>/<int:par_level>",
           methods=['GET'])
@auth.login_required
def restgetgradeentries(par_subject, par_schoolid, par_semid,
                        par_quarter_, par_section, par_level):
    return jsonify(getgradeentries(par_subject.split(' ' + par_section)[0], par_schoolid,
                                   par_semid, par_quarter_, par_section,
                                   par_level, auth.username()))


@app.route("/gradebook/<string:par_subject>/<int:par_level>/<string:par_schoolid>/<string:par_semid>/" +
           "<string:par_quarter_>/<string:par_section>/<string:par_currid>",
           methods=['GET'])
@auth.login_required
def restgetgradetable(par_subject, par_schoolid, par_semid,
                      par_quarter_, par_section, par_level, par_currid):
    categories = spcall("getcategorysequence", (par_currid,))
    username = auth.username()
    for category in categories:
        classlist = getclasslistbygender(username, par_schoolid, par_semid, par_level, par_section)


def saveindivquartergrade(par_studentid, par_subject, par_section, par_level, par_semid,
                          par_schoolid, par_grade, par_quarter_, username, token, group):
    subject = par_subject.split(' ' + par_section)[0]
    return formatres(spcall(
        "insupsummativegrade",
        (par_studentid, subject, par_section, par_level,
         par_semid, par_schoolid, par_grade, par_quarter_, username,
         token, group, False),
        group, True))


# @app.route("/gradebook/<string:par_studentid>/<string:par_subject>/<string:par_section>/" +
#            "<int:par_level>/<string:par_semid>/<string:par_schoolid>/<float:par_grade>/" +
#            "<string:par_quarter_>", methods=['POST'])
# @auth.login_required
# def savestuquartergrade(par_studentid, par_subject, par_section, par_level, par_semid,
#                         par_schoolid, par_grade, par_quarter_):
#     item = saveindivquartergrade(par_studentid, par_subject, par_section,
#                                  par_level, par_semid, par_schoolid, par_grade,
#                                  par_quarter_, auth.username())
#
#     if item["status"] != 'Error':
#         timeres = spcall("stampevent", (auth.username(), par_studentid,
#                                         "Recorded Grade Entry " + str(par_grade) + " for subject "
#                                         + par_subject,
#                                         'grade', 2, par_semid, par_schoolid), "super", True)
#
#     return jsonify(item)

@app.route("/gradebook/lock/<string:par_offeringid>", methods=["PUT"])
@auth.login_required
def lockclassrecord(par_offeringid):
    params = request.get_json()
    token = params["token"]
    quarter = params["quarter"]
    group = params["group"]

    lockres = spcall("lockgrades",
                     (
                         par_offeringid,
                         quarter,
                         auth.username(),
                         group,
                         token,
                         True
                     ),
                     "faculty", True)

    if 'Error' in lockres[0][0]:
        return jsonify(
            {
                "status":"error",
                "message": lockres[0][0]
            })

    return jsonify(lockres[0][0])

@app.route("/gradebook/publish/<string:par_offeringid>", methods=["PUT"])
@auth.login_required
def publishclassrecord(par_offeringid):
    params = request.get_json()
    token = params["token"]
    quarter = params["quarter"]
    group = params["group"]
    name = params["name"]
    username = auth.username() 

    
    
    lockres = spcall("publishgrades",
                     (
                         par_offeringid,
                         quarter,
                         auth.username(),
                         group,
                         token,
                         True
                     ),
                     "faculty", True)

    

    if 'Error' in lockres[0][0]:
        return jsonify(
            {
                "status":"error",
                "message": lockres[0][0]
            })
    else:
        print(lockres[0][0])
        
       
        if lockres[0][0]['message'] != 'This is already published!': 
            messageTextOnly = f"{name} publish grade action: {lockres[0][0]['message']}"
            poster = f"{name} has graded the class {lockres[0][0]['subdetails']}. You can check it Now."
            msg_type = 'grade'
            usernum = spcall("getpersonidbyusername", (username,),)[0][0]
            channels = lockres[0][0]['studentvroomid']
            initiatorid = lockres[0][0]['initiatorid']
            receivers = ['']
            timestamps = lockres[0][0]['ts']

            # print(receivers)

            notif = pub.notifications(username=username,
            poster=poster, msg_payload=messageTextOnly, type=msg_type, user_type=group, 
            channels=channels, initiator_id=initiatorid, section=None,
            receiver_id=receivers, tstamp=timestamps, due_date=None, start_date=None,
            name=name, action_initiator=usernum)

            notif.notify()

    return jsonify(lockres[0][0])
    # return jsonify({'status': 'Ok', 'message': 'Goods'})



@app.route("/gradebook/unlock/<string:par_offeringid>", methods=["PUT"])
@auth.login_required
def unlockclassrecord(par_offeringid):
    pass


@app.route("/gradebook/<string:par_subject>/<string:par_section>/" +
           "<int:par_level>/<string:par_semid>/<string:par_schoolid>/" +
           "<string:par_quarter_>/<string:par_entries>/<string:par_offeringid>", methods=['POST'])
@auth.login_required
def savegroupquartergrade(par_subject, par_section, par_level, par_semid,
                          par_schoolid, par_quarter_, par_entries, par_offeringid):
    username = auth.username()
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    entries = par_entries.split("@")
    # par_entries = stuid*grade@...

    results = []
    for entry in entries:
        detail = entry.split("*")
        #print entry
        item = saveindivquartergrade(detail[0], par_subject, par_section,
                                     par_level, par_semid, par_schoolid,
                                     detail[1], par_quarter_, username, token, group)
        #group = spcall("getuser_group", (username,))[0][0]
        if item["status"] != 'Error':
            submitres = spcall("insupgradesubmissions", (par_offeringid,
                                                         par_semid,
                                                         spcall("getpersonidbyusername", (username,))[0][0],
                                                         par_quarter_),
                               group, True)

            # return jsonify(formatres(lockres))

        results.append({"lrn": detail[0],
                        "result": item["item"][0]
                        })

    #lockres = spcall("lockgrades", (par_subject, par_section,
    #                                par_level, par_schoolid, par_quarter_, par_semid, True),
    #                 group, True)

    return jsonify({"status": "ok",
                    "results": results})


@app.route("/gradebook", methods=['POST'])
@auth.login_required
def entryfromexcel():
    params = request.get_json()
    entries = params["records"]
    semid = params["semid"]
    schoolid = params["schoolid"]
    quarter = params["quarter"]
    gradelevel = params["grade"]
    section = params["section"]
    token = params["token"]
    group = params["group"] #spcall("getuser_group", (auth.username(),))[0][0]

    result = []
    sub = ''
    for entry in entries:
        if sub != entry["subjectno"]:
            subjectno = spcall("getsubjectfromoffid", (entry["sectionid"],))[0][0]
            sub = entry["subjectno"]

        for detail in entry["details"]:
            #print detail["student"]
            lrn = spcall("getpersonnumidbysanitizedname", (detail["student"], 'students'))[0][0]
            #print entry["category"]
            qryres = formatres(spcall("entryscorefromexcel", (str(semid),
                                                              str(schoolid),
                                                              str(quarter),
                                                              int(gradelevel),
                                                              str(section),
                                                              str(entry["sectionid"]),
                                                              detail["student"],
                                                              str(entry["category"]),
                                                              str(detail["entrycount"]),
                                                              float(detail["maxscore"]),
                                                              float(detail["score"]),
                                                              auth.username(),
                                                              token,
                                                              group
                                                              ),
                                      group,
                                      True
                                      ))
            '''
            if qryres["status"] != 'Error':
                timeres = spcall("stampevent", (auth.username(),
                                                lrn,
                                                "Encoded with a score of " + str(detail["score"]) + ' to ' +
                                                str(entry["sectionid"]).split(semid)[0] + ' via Spreadsheet',
                                                'grade', 2, semid, schoolid), "super", True)
                # select computestugradepercat('0078', 'MAKOPA', 1, '3', getcurrsem(), '128081', 'WRITTEN WORKS', 'FIL 1', 'K-12')
            
            graderes = spcall("computestugradepercat", (lrn,
                                                        section, int(gradelevel), quarter, semid, schoolid,
                                                        entry["category"], subjectno, spcall("getcurriculumfromsem", (semid))
                                                        [0][0]),
                              group, True)
            '''

            result.append({
                "sectionid": entry["sectionid"],
                "category": entry["category"],
                "studentname": detail["student"],
                "lrn": lrn,
                "result": qryres["status"],
                'entrycount': str(detail["entrycount"]),
                'scoreentry': qryres["message"]
            })

        #quartergraderes = spcall("computequartergrade", (subjectno, schoolid, semid, quarter,
        #                                                 spcall("getcurriculumfromsem", (semid))[0][0],
        #                                                 section, gradelevel), group, True)


    return jsonify({"result": result})


@app.route("/gradebook/<string:par_section>", methods=['GET'])
@auth.login_required
def getclassrecord(par_section):
    params = request.args
    semid = params["semid"]
    schoolid = params["schoolid"]
    quarter = params["quarter"]
    gradelevel = params["grade"]
    section = params["section"]
    subjectx = params["subject"]
    token = params["token"]


    classrecord = spcall("retrievegradebook", (par_section, auth.username(),
                                               semid, schoolid, quarter, gradelevel, section,
                                               subjectx.split(' ' + section)[0], token, "faculty"), "faculty", True)

    return jsonify(finalize(classrecord))

    # group = 'faculty'
    # subject = subjectx.split(' ' + section)[0]
    # activecurr = spcall("getactivecurriculum",())[0][0]
    # categories = spcall("getcategorysequence",(activecurr,))
    # out = []
    # defheader = {'lrn':'', 'name':''}
    # classlist = spcall("getclasslist2", (str(section), int(gradelevel), str(schoolid), str(semid)), group)
    # classrecord = []
    # for student in classlist:
    #     lrn = student[0]
    #     studentrow = {"lrn":lrn, "name":student[2]}
    #     inigrade = 0
    # #
    #     for category in categories:
    #         f = category[0].split(' ')
    #         tabidroot = (f[0][:1] + f[1][:1]).lower()
    #         #entries = spcall("getentrieswithfullname",(par_section,),group) #returns list of stu scoress
    #         entries = spcall("get_entry_header", (par_section, quarter, category[0], semid), group)
    #         total = 0
    #         totalstu = 0
    #         if len(entries) == 0:
    #             out.append({
    #                 "status":"ok",
    #                 "category":category[0],
    #                 "size":0
    #             })
    #             continue
    # #
    #         for entry in entries:
    #             entryname = tabidroot + entry[2]
    #             defheader[ entryname ] = str(entry[1])
    #             total = total + entry[1]
    #             stu_scores = spcall("getentryforstudent",(entry[0],lrn),group)
    # #
    #             if len(stu_scores) == 0:
    #                 continue
    #             studentrow[entryname] = stu_scores[0][0]
    #             totalstu += float(stu_scores[0][0])
    # #         #get the grade
    # #
    #         stugrade = spcall("getstugradesdistrib", (lrn, subject, category[0],
    #                                                   semid, schoolid, quarter,
    #                                                   activecurr, section,
    #                                                   gradelevel), group)
    # #         #return jsonify({"dfdf":stugrade})
    #         defheader[ tabidroot+'total' ] = str(total)
    #         studentrow[tabidroot + 'total'] = str(totalstu)
    #         if len(stugrade) == 0:
    #             continue
    # #
    # #
    #         studentrow[tabidroot + "ps"] = str(round(stugrade[0][0],2))
    #         studentrow[tabidroot + "ws"] = str(round(stugrade[0][0] * (stugrade[0][2]/100),2))
    #         inigrade += round(stugrade[0][0] * (stugrade[0][2]/100),2)
    #         defheader[ tabidroot+'ws' ] = str(stugrade[0][2])
    # #     #gettotalgrade
    # #     #select * from getstudentsummativequartergrade('ARTS 1', '0078',getcurrsem(), 'MAKOPA', 1, '128081', '3');
    # #
    #     quartergrade = spcall("getstudentsummativequartergrade",
    #                           (subject, lrn, semid, section, gradelevel, schoolid, quarter),
    #                           group)
    #     if len(quartergrade) == 0:
    #         studentrow["grade"] = '-'
    #         studentrow["igrade"] = '-'
    #     else:
    #         studentrow["grade"] = str(quartergrade[0][0])
    #         studentrow["igrade"] = str(inigrade)
    #     classrecord.append(studentrow)
    # #
    # defheader[tabidroot + 'ps' ] = '100.00'
    # classrecord.insert(0, defheader)
    # #
    # #
    # return jsonify({"status":"OK", "size":len(classrecord), "classrecord":classrecord})


def getclassrecord1(par_section):
    sectioncode = par_section
    params = request.args
    semid = params["semid"]
    schoolid = params["schoolid"]
    quarter = params["quarter"]
    gradelevel = params["grade"]
    section = params["section"]
    subject = params["subject"]
    token = params["token"]
    username = auth.username()
    group = spcall("getuser_group", (username,))[0][0]
    #retrievegradebook('ARPAN 320162017&ORCHIDS128081', 'irene.quilang', '20162017', '128081', '1', '3', 'ORCHIDS',
    #                  'ARPAN 3', '5c1c271623bcf752d3e7b6d4fdbd414918f0c1c9');
    # classrecord = spcall("retrievegradebook", (
    #     sectioncode,
    #     auth.username(),
    #     semid,
    #     schoolid,
    #     quarter,
    #     gradelevel,
    #     section,
    #     subject,
    #     token))[0][0]
    #
    # if 'Error' in classrecord:
    #     return jsonify({
    #         "status":"error",
    #         "message":classrecord
    #     })


    #token = params["token"]
    activecurr = spcall("getactivecurriculum", ())[0][0]
    categories = spcall("getcategorysequence", (activecurr,))
    out = []
    defheader = {'lrn': '', 'name': ''}
    classlist = spcall("getclasslist2", (str(section), int(gradelevel), str(schoolid), str(semid)), group)
    classrecord = []
    for student in classlist:
        lrn = student[0]
        studentrow = {"lrn": lrn, "name": student[2]}
        inigrade = 0

        for category in categories:
            f = category[0].split(' ')
            tabidroot = (f[0][:1] + f[1][:1]).lower() #gets first letters of words like ww for written works, pt for performance tasks in lower case
            # entries = spcall("getentrieswithfullname",(par_section,),group) #returns list of stu scoress
            entries = spcall("get_entry_header", (par_section, quarter, category[0], semid), group)
            total = 0
            totalstu = 0
            if len(entries) == 0:
                out.append({
                    "status": "ok",
                    "category": category[0],
                    "size": 0
                })
                continue

            for entry in entries:
                entryname = tabidroot + entry[2]
                defheader[entryname] = str(entry[1])
                total = total + entry[1]
                stu_scores = spcall("getentryforstudent", (entry[0], lrn), group)

                if len(stu_scores) == 0:
                    continue
                studentrow[entryname] = stu_scores[0][0]
                totalstu += float(stu_scores[0][0])
            # get the grade

            stugrade = spcall("getstugradesdistrib", (lrn, subject, category[0],
                                                      semid, schoolid, quarter,
                                                      activecurr, section,
                                                      gradelevel), group)
            # return jsonify({"dfdf":stugrade})
            defheader[tabidroot + 'total'] = str(total)
            studentrow[tabidroot + 'total'] = str(totalstu)
            if len(stugrade) == 0:
                continue

            studentrow[tabidroot + "ps"] = str(round(stugrade[0][0], 2))
            studentrow[tabidroot + "ws"] = str(round(stugrade[0][0] * (stugrade[0][2] / 100), 2))
            inigrade += round(stugrade[0][0] * (stugrade[0][2] / 100), 2)
            defheader[tabidroot + 'ws'] = str(stugrade[0][2])
        # gettotalgrade
        # select * from getstudentsummativequartergrade('ARTS 1', '0078',getcurrsem(), 'MAKOPA', 1, '128081', '3');

        quartergrade = spcall("getstudentsummativequartergrade",
                              (subject, lrn, semid, section, gradelevel, schoolid, quarter),
                              group)

        if len(quartergrade) == 0:
            studentrow["grade"] = '-'
            studentrow["igrade"] = '-'
        else:
            studentrow["grade"] = str(quartergrade[0][0])
            studentrow["igrade"] = str(inigrade)
        classrecord.append(studentrow)

    defheader[tabidroot + 'ps'] = '100.00'
    classrecord.insert(0, defheader)

    return jsonify({"status": "OK", "size": len(classrecord), "classrecord": classrecord})#jsonify(classrecord)


@app.route("/quarter/<string:section>/<int:level>/<string:quarter>/<string:schoolid>/<string:semid>", methods=['POST'])
@auth.login_required
def setquarter(section, level, quarter, schoolid, semid):
    params = request.get_json()
    #print params
    token = params["token"]
    group = params["group"]
    #isadviser = params["adviser"]
    username = auth.username()
    verify = spcall("verifyuser", (username,token, "faculty"))[0][0]

    if 'Error' in verify[0][0]:
        quarter = formatres(spcall("getactivequarter",
                                   (section, level, semid, schoolid), group, False))
        curqua = {}
        curqua["quarter"] = quarter["item"][0]

        return jsonify(curqua)


    #if isadviser == 'false':
    #    section = username

    success = formatres(spcall("setquarter",
                               (section, level, semid,
                                schoolid, quarter, True),
                                "faculty", True))

    if success["status"] == 'Error':
        quarter = formatres(spcall("getactivequarter",
                                   (section, level, semid, schoolid), group, False))

        success["quarter"] = quarter["item"][0]

        return jsonify(success)

    success["quarter"] = quarter #quarter from parameters


    return jsonify(success)


@app.route("/attendance/<string:par_studentid>/<string:par_semid>/<string:par_status>/" +
           "<string:par_att_date>/<float:par_points>/<string:par_daypart>/" +
           "<string:par_schoolid>/<string:par_section>/<int:par_level>", methods=['POST'])
@auth.login_required
def add2attendance(par_studentid, par_semid, par_status, par_att_date, par_points, par_daypart, par_schoolid,
                   par_section, par_level):
    result = formatres(
        spcall("add2attendance", (par_studentid, par_semid, par_status, par_att_date,
                                  par_points, par_daypart, par_schoolid, par_section, par_level),
               getusergroup(auth.username()), True)
    )

    if result["status"] != 'Error':
        timeres = spcall("stampevent", (auth.username(), par_studentid,
                                        "Recorded attendance on " + str(par_att_date) + " " +
                                        par_daypart + " session as " + par_status,
                                        'attendance', 2, par_semid, par_schoolid), "super", True)

    return jsonify(result)


@app.route("/attendance/<string:par_entries>/<string:par_section>/" +
           "<int:par_level>/<string:par_schoolid>/<string:par_semid>/<string:par_att_date>", methods=['POST'])
@auth.login_required
def batchattandanceentry(par_entries, par_section, par_level, par_schoolid, par_semid, par_att_date):
    # lrn*status*points*daypart@
    result = []
    group = getusergroup(auth.username())
    for entry in par_entries.split("@"):
        stuid = entry.split("*")[0]
        status = entry.split("*")[1]
        points = entry.split("*")[2]
        daypart = entry.split("*")[3]
        item = formatres(
            spcall("add2attendance", (stuid, par_semid, status, par_att_date,
                                      points, daypart, par_schoolid, par_section, par_level),
                   group, True)
        )

        if item["status"] != 'Error':
            timeres = spcall("stampevent", (auth.username(), stuid,
                                            "Recorded attendance on " + str(par_att_date) + " " +
                                            daypart + " session as " + status,
                                            'attendance', 2, par_semid, par_schoolid), "super", True)

        result.append({
            "lrn": stuid,
            "result": item["item"][0]
        })

    return jsonify({
        "results": result,
        "status": "ok"
    })


@app.route("/attendance/<string:par_section>/" +
           "<int:par_level>/<string:par_schoolid>/" +
           "<string:par_semid>/<string:par_att_date>/<string:par_daypart>", methods=['GET'])
@auth.login_required
def getattendance(par_section, par_level, par_schoolid, par_semid, par_att_date, par_daypart):
    # getclassattendance('makopa', 1, '128081', getcurrsem(), '02/22/2016', 'AM')
    items = []
    for item in spcall("getclassattendance", (par_section,
                                              par_level, par_schoolid,
                                              par_semid,
                                              par_att_date.replace("-", "/"),
                                              par_daypart),
                       getusergroup(auth.username())):
        items.append(dict({"id": item[0],
                           "status": item[1]}))
    return jsonify({"status": "ok",
                    "items": items})


def getstugrade(par_lrn, par_semid, par_curriculumid, par_schoolid, group):
    # getcard('FRIEND', getcurrsem(), '128081', 'K-12')
    cardtext = formatres(spcall("getcard", (par_lrn.upper(),
                                            par_semid,
                                            par_schoolid,
                                            par_curriculumid), group))

    if cardtext["status"] == 'Error':
        return {"classcard": [cardtext["message"]]}

    quarters = 1
    classcard = []
    for quarter in cardtext["item"][0].split("*")[:-1]:
        grades = quarter.split('#quarter')[:-1][0].split(',')[:-1]
        gradecompose = []
        for grade in grades:
            decograde = grade.split("=")
            gradecompose.append({
                "subject": decograde[0],
                "grade": decograde[1]
            })
        classcard.append({
            "quarter": quarters,
            "subjectgrade": gradecompose
        })
        quarters = quarters + 1
    return classcard


@app.route("/card/<string:par_lrn>/<string:par_semid>",
           methods=['GET'])
@auth.login_required
def restgetstugrade(par_lrn, par_semid):
    params = request.args
    #schoolid = params["schoolid"]
    group = getusergroup(auth.username())
    classcards = []
    if par_semid == '*':
        # select * from getstudentsemesters('0078')
        stusemesters = spcall("getstudentsemesters", (par_lrn,), group)
        if len(stusemesters) == 0:
            return jsonify({
                "status": "ok",
                "size": 0,
                "classcard": []
            })
        idx = 0
        sems = []
        for stusemester in stusemesters:
            #print(stusemester[0])
            classcards.append(getstugrade(par_lrn, stusemester[0], stusemester[1], stusemester[2], group))
            sems.append(stusemester[0])
            idx = idx + 1

        for jidx in range(0, idx):
            classcards[jidx].insert(0, {
                "quarter": 1,
                "subjectgrade": [{
                    "subject": "School Year:",
                    "grade": spcall("getsemdesc", (sems[jidx],), group)[0][0]
                }]})
        #print(classcards)
    else:
        activecurriculum = formatres(
            spcall("getcurriculumfromsem", (par_semid,),
                   group))["item"][0]
        #print par_lrn, par_semid, activecurriculum, schoolid, group
        classcards.append(getstugrade(par_lrn, par_semid, activecurriculum, '', group))

    return jsonify({
        "status": "ok",
        "size": len(classcards),
        "classcard": classcards
    })


@app.route("/timeline",
           methods=['GET'])
@auth.login_required
def gettimelines():
    params = request.args
    page = params["offset"]
    lrn = params["lrn"]
    token = params["token"]
    getchild = params["getchild"]
    timelines = formatres(spcall("poptimelines", (lrn, auth.username(), page, token, getchild == 'getchild'), "super", True))


    if timelines["status"] == "Error":
        return jsonify({"status": "error", "message": timelines["item"][0]})

    tsize = timelines[u"item"][0][u"size"]

    if tsize == 0:
        return jsonify({"status": "ok",
                        "size": 0})

    timelines = timelines[u"item"][0][u"timelines"]
    pic = fileServer()
    timelinelists = []
    for timeline in timelines:
        ##this is for showing the feelings with its count
        # recordentryid = ""
        # recentfeeling = ""
        # if timeline[2] == 'grade' and group == 'students':
        # owner = timeline[0].split('$')[2]
        # ts = str(timeline[1])
        # feelings = spcall("showfeelings", (), group)
        # user = spcall("getpersonidbyusername", (auth.username(),), group)[0][0]

        # recordentryid = spcall("getrecordentryidbyts", (str(timeline[3]),))[0][0]
        # recordentryid = getrecordentryid(timeline[0].split('$')[0].split('msg:')[1], timeline[0].split('$')[1].split(':')[1], user, auth.username(), str(timeline[3]))
        # return jsonify({"FF":timeline[0].split('$')[0].split('msg:')[1]})
        # entryid = spcall("getentryidbyrecordentryid", (recordentryid,), group)[0]
        # recentfeeling = spcall("recentfeeling", (user, entryid,), group)
        # return jsonify({"FF":recentfeeling})

        timelineitem = dict()
        timelineitem["body"] = timeline[u"body"].split('$')[0].split('msg:')[1]
        timelineitem["pic"] = pic.getImgSrcUrl(timeline[u"pic"])
        # timelineitem["pic"] = timeline[u"pic"]
        timelineitem["tstamp"] = str(timeline[u'tstamp'])
        timelineitem["tltype"] = timeline[u'tltype']
        timelineitem["owner"] = timeline[u"body"].split('$')[2]
        # timelineitem["feelings"] = feelings
        # timelineitem["recentfeeling"] = recentfeeling
        # timelineitem["recordentryid"] = recordentryid
        timelineitem["initiatorid"] = timeline[u"initiatorid"]
        timelineitem["receiverid"] = timeline[u"receiverid"]
        timelineitem["tlts"] = timeline["tlts"]
        timelineitem["reactions"] = dict({
            'okc':timeline["okc"],
            'happyc': timeline['happyc'],
            'sadc': timeline["sadc"],
            'angryc': timeline["angryc"],
            'surprisedc': timeline["surprisedc"]
        })                           #timeline[u"comments"]
        timelineitem["comments"] = spcall("get_comments_relaxed", (timeline[u"initiatorid"], timeline[u"receiverid"], timeline["tlts"], 0, auth.username()), "super", True)[0][0]
        timelinelists.append(timelineitem)

    
    return jsonify(
        {
            "status": "ok",
            "size": tsize,
            "timelines": timelinelists
        }
    )


def getrecordentryid(par_message, par_initiatorid, par_receiverid, par_username, par_ts):
    # "You were Encoded with a score of 2.0 to ARPAN 3 WRITTEN WORKS Entry Number:2 MAX SCORE:5.0 by QUILANG, IRENE B.."
    # return par_receiverid
    message = par_message.split(" ")
    offeringid = str(message[9]) + " " + str(message[10])
    try:
        entrynumber = str(message[14].split(':')[1])
    except:
        entrynumber = '0'
    # return entrynumber
    try:
        category = str(message[11]) + " " + str(message[12])
    except:
        category = 'Default'
    # return category
    username = digest(str(par_username))
    ptype = spcall("getuser_group", (str(par_username),))[0][0]
    # return {"e":ptype, "t":username}
    personnumid = spcall("getidnumbyuser", (username, ptype,))[0][0]
    # return {"d":personnumid}
    semid = spcall("getcurrsem", ())[0][0]
    schoolid = spcall("getpersonschool", (personnumid, semid,))[0][0]
    levelsection = spcall("getlevelsection", (personnumid, schoolid, semid,))[0]
    section = levelsection[1]
    level = levelsection[2]
    quarter = spcall("getactivequarter", (section, level, semid, schoolid,))[0][0]
    return quarter
    #entryid = str(offeringid) + str(semid) + '&' + str(section) + str(schoolid) + str(semid) + str(entrynumber) + str(quarter) + str(category)
    #return entryid

    #result = spcall("getrecordentryidbyentryid", (par_initiatorid, par_receiverid, entryid, par_ts))[0][0]

    #return result


# @app.route("/feeling/<string:tlid>/<string:feelingid>/<string:username>", methods=['POST'])
# # @auth.login_required
# def record_feeling(tlid, feelingid, username):
#     # return jsonify({"e":username})
#     personid = spcall("getpersonidbyusername", (username,), "students")[0][0]
#     # return jsonify({"d":personid})
#     # group = spcall("getuser_group", (username,))
#     # return jsonify({"k":group})
#     group = "students"
#     entryid = spcall("getentryidbyrecordentryid", (tlid,), group)[0]
#     # return jsonify({"d":entryid, "t":tlid, "f":feelingid})
#     # initiatorid = entrydetails[0]
#     # receiverid = entrydetails[1]
#
#     recentfeeling = spcall("recentfeeling", (personid, entryid,), group)
#     # return jsonify({'e':recentfeeling})
#     result = spcall("recordfeeling", (tlid, int(feelingid), personid, entryid), group, True)[0][0]
#     # return jsonify({"res":result})
#
#     try:
#         rf = int(recentfeeling[0][0])
#     except:
#         rf = 0
#
#     # return jsonify({"p":result, "l":rf});
#     if result == 'OK':
#         # return jsonify({"p":result, "l":feelingid});
#         if int(feelingid) == 1:
#             res = spcall("upd8angryfeeling", (entryid, rf,), group, True)[0][0]
#         elif int(feelingid) == 2:
#             res = spcall("upd8sadfeeling", (entryid, rf,), group, True)[0][0]
#         elif int(feelingid) == 3:
#             res = spcall("upd8happyfeeling", (entryid, rf,), group, True)[0][0]
#         elif int(feelingid) == 4:
#             res = spcall("upd8surprisedfeeling", (entryid, rf,), group, True)[0][0]
#         else:
#             res = "Uh-oH"
#
#         # feelings = spcall("showfeelings", (ownerid, thispersonid, ts,), group)[0][0]
#         return jsonify({"status": "ok", "resstatus": res})
#     else:
#         return jsonify({"status": "error", "message": "Something went wrong.", "result": result})
#

# @app.route("/feeling/<string:par_entryid>/<string:par_username>", methods=['GET'])
# # @auth.login_required
# def getreportpersubjectentry(par_entryid, par_username):
#     initiatorid = spcall("getpersonidbyusername", (par_username,))[0][0]
#
#     # recordentryid = spcall("getrecordentryidbyfaculty", (initiatorid, par_entryid,), 'faculty')[0][0]
#
#     result = spcall("getfeelingcounts", (par_entryid, initiatorid,), "faculty")
#     # angrycount = feelingcount[0][0]
#     # sadcount = feelingcount[
#     # happycount = spcall("gethappyfeelingcount", (par_entryid, initiatorid,), "faculty")[0][0]
#     # surprisedcount = spcall("getsurprisedfeelingcount", (par_entryid, initiatorid,), "faculty")[0][0]
#
#     # return jsonify({"l":records})
#     feelingcounts = dict()
#
#     try:
#         feelingcounts["hcount"] = result[0][2]
#         feelingcounts["acount"] = result[0][0]
#         feelingcounts["scount"] = result[0][1]
#         feelingcounts["ucount"] = result[0][3]
#     except:
#         feelingcounts["hcount"] = 0
#         feelingcounts["acount"] = 0
#         feelingcounts["scount"] = 0
#         feelingcounts["ucount"] = 0
#
#     return jsonify(feelingcounts)


# frequent_itemset = dict()
# svm_model = dict()

# @app.route("/reports", methods=["GET"])
# @auth.login_required
# def initialize_for_reports():
#     subjects = ['arpan', 'arts', 'cad', 'cid',
#                 'cvd', 'eng', 'epp', 'esp',
#                 'fil', 'hlt', 'hwmd', 'ld',
#                 'math', 'mt', 'mus', 'pe',
#                 'sci', 'tle'
#                 ]
#     frequent_itemset = dict()
#     svm_model = dict()
#     for subject in subjects:
#         res = dict()
#         fp = feelingfpgrowth()
#         res['1'] = fp.generate_itemsets(1, subject)
#         res['2'] = fp.generate_itemsets(2, subject)
#         res['3'] = fp.generate_itemsets(3, subject)
#         res['4'] = fp.generate_itemsets(4, subject)
#         # global frequent_itemset
#         frequent_itemset[subject] = res
#         gd = getdata()
#         svm = feelingssvm()
#         dataset = gd.get_data(subject)
#         # global svm_model
#         svm_model[subject] = svm.create_model(dataset)
#
#     return {"frequent_itemset": frequent_itemset, "svm_model": svm_model}


# @app.route("/reports/<string:subject>")
# # @auth.login_required
# def get_reports(subject):
#     init = initialize_for_reports()
#     # return jsonify({"init":init})
#     group = spcall("getuser_group", (auth.username(),))[0][0]
#
#     params = request.args
#     # return jsonify({"message":params})
#     gradingperiod = params["gradingperiod"]
#     # return jsonify({"mes":gradingperiod})
#     section = params["section"]
#     grade_level = params["grade_level"]
#     schoolid = params["schoolid"]
#     semid = params["semid"]
#
#     latest_grades_of_students = get_latest_grades_of_students(gradingperiod,
#                                                               section,
#                                                               grade_level,
#                                                               schoolid,
#                                                               semid,
#                                                               subject)
#     # return jsonify({"message":latest_grades_of_students})
#     pfp = PredictByFP()
#     svm = FeelingsSVM()
#     # return jsonify({"mes": svm})
#     result = []
#     for each in latest_grades_of_students:
#         student = each["student_detail"]
#         fp_growth_result = pfp.predict_grade(init["frequent_itemset"][str(subject)][gradingperiod], each["data_for_fp"],
#                                              str(subject))
#         svm_result = svm.predict(init["svm_model"][str(subject)], each["data_for_svm"])
#
#         lrn = student[0]
#         row = dict()
#         row["fp_growth"] = fp_growth_result
#         row["svm"] = svm_result
#         row["details"] = student
#         result.append(row)
#
#     return jsonify({"status": "ok", "result": result})


# def get_latest_grades_of_students(gradingperiod, section, gradelevel, schoolid, semid, subject):
#     classlist = spcall("getclasslist", (section, gradelevel, schoolid, semid))
#     # return classlist
#     result = []
#
#     for each in classlist:
#         forfp = spcall("getdataforfpquery", (each[0], gradingperiod, subject))
#         processedforfp = process_for_fp(forfp)
#
#         forsvm = spcall("getdataforsvmquery", (each[0], subject + str(gradelevel), gradingperiod))
#         processedforsvm = process_for_svm(forsvm)
#         row = dict()
#         row["student_detail"] = each
#         row["data_for_fp"] = processedforfp
#         row["data_for_svm"] = processedforsvm
#         result.append(row)
#
#     return result
#
#
# def process_for_fp(data):
#     result = []
#     return data
#     row = dict()
#     for each in data:
#         try:
#             if row[each[0]] == '':
#                 row[each[0]] = row[each[0]] + str(each[1]) + ',' + \
#                                str(grade) + ',' + str(each[3])
#         except:
#             row[each[0]] = ''
#             grade = (float(each[2]) / float(each[1])) * 100
#             row[each[0]] = row[each[0]] + str(each[1]) + ',' + \
#                            str(grade) + ',' + str(each[3])
#     temp = []
#     for key, value in row:
#         temp.append(value)
#     result.append(temp)
#     return result
#
#
# def process_for_svm(data):
#     result = []
#
#     if data is not None:
#         row = []
#         for each in data:
#             row.append(float(each[2]))
#
#         if len(data) < 4:
#             additional = 4 - len(data)
#             for i in range(0, additional):
#                 default = float(80.00)
#                 row.append(default)
#
#         result.append(row)
#
#     return result


def buttonact(bclass, label, fcn, id):
    return '<button id="' + id + '" class="btn btn-'+ bclass+'"' + ' onclick="'+fcn+'" >'+label+'</button>';


@app.route("/virtualroom/monitor", methods=["GET"])
@auth.login_required
def virtualclassroomslist():
    username = auth.username()
    params = request.args
    token = params["token"]
    schoolid = params["schoolid"]
    semid = params["semid"]
    date_ = params["date_"]
    group = params["group"]

    if group != 'admin':
        return jsonify({"status": "error", "message": "User type not allowed!"})

    onlineclasses = spcall("getonlineclass", (
        username,
        token,
        schoolid,
        semid,
        date_
    ), group, False)[0][0]

    if 'Error' in onlineclasses:
        return jsonify({"status": "error", "message": onlineclasses[0][0]})

    return jsonify({'status': 'OK', 'onlineclasses': onlineclasses})


@app.route("/virtualroom/join", methods=["POST"])
@auth.login_required
def joinvirtualroom():
    #available for students only
    params = request.get_json()
    semid = params["semid"]
    level = params["level"]
    section = params["section"]
    schoolid = params["schoolid"]
    group = params["group"]
    token = params["token"]
    username = auth.username()

    virtualroomid = spcall("getvirtualroom",
                           (username, token, group,
                            level, section,
                            schoolid, semid, ''
                            ), group, True)[0][0]

    if virtualroomid == 'NONE':
        return jsonify({"status": "error", "message": "You are not yet enrolled!"})

    return jsonify({'status': 'OK', 'virtualroomid': virtualroomid})


@app.route("/virtualroom", methods=['POST'])
@auth.login_required
def bcastvirtualclassroom():
    params = request.get_json()
    semid = params["semid"]
    level = params["level"]
    section = params["section"]
    schoolid = params["schoolid"]
    group = params["group"]
    token = params["token"]
    offid = params["offid"]
    username = auth.username()
    #getvirtualroom(par_username text, par_token text, par_group text, par_level int, par_section text, par_schoolid text, par_semid text) returns text AS

    if schoolid == 'NOT SET':
        appending = username.split('.')[0]
        virtualroomid = appending + get_random_alphanumeric_string(15)
    else:
        virtualroomid = spcall("getvirtualroom",
                               (username, token, group,
                                level, section,
                                schoolid, semid, offid
                                ), group, True)[0][0]

        if virtualroomid == 'NONE':
            appending = username.split('.')[0]
            virtualroomid = appending + get_random_alphanumeric_string(15)

    if 'Error' in virtualroomid:
        return jsonify({"status": "error", "message": virtualroomid[0][0]})

    message = "Join the  " + buttonact("success", "Video Conference", "model.joinvideocall('"+virtualroomid+"')", "btnjoin" + virtualroomid)

    res = spcall("post_timeline2",
                 (username, token, '',
                  '01/02/2018', message,
                  group, 'Bulletin Board',
                  True, 3, semid, schoolid
                  ), group, True)


    channels = []

    channels.append(virtualroomid)
    channels.append(schoolid)

    msg_type = 'BCast'  
    # messageTextOnly = 'Posted a video Conference'
    person_id = spcall("getpersonidbyusername", (username,),)[0][0]
    person = spcall("getpersonname", (person_id,),)[0][0]
    person = person.split("*")
    poster = f'{person[1]} {person[0]} has posted a Video Conference!'
    name = f'{person[1]} {person[0]}'
    messageTextOnly = f'{poster} has posted a Video Conference'
    receivers = []
    initiatorid = res[0][0]['initiatorid']
    timestamps = res[0][0]['responses'][0]['ts']
    if len(res[0][0]['responses']) > 1:
        for item in res[0][0]['responses']:
            receivers.append(item['receiverid'])
    else:
        receivers.append(res[0][0]['responses'][0]['receiverid'])

    print(receivers)

    usernum = spcall("getpersonidbyusername", (username,),)[0][0]
    notif = pub.notifications(username=username,
            poster=poster, msg_payload=messageTextOnly, type=msg_type, 
            user_type=group, channels=channels, 
            initiator_id=initiatorid, receiver_id=receivers, tstamp=timestamps,
            due_date=None, start_date=None, section=None,
            name=name, action_initiator=usernum)
    

    

    if 'Error' in res[0][0]:
        return jsonify({"status": "error", "message": res[0][0]})
    else:
        notif.notify()
    return jsonify({'status': 'OK', 'virtualroomid': virtualroomid})



@app.route("/bulletin", methods=['POST'])
@auth.login_required
def bulletinpost():
    params = request.get_json()
    schoolid = params["schoolid"]
    semid = params["semid"]
    message = params["message"]
    token = params["token"]
    group = params["group"]
    vroomid = params["vroomid"]
    notif_msg = params["notif_msg"]
    name = params["name"]
    username = auth.username() 
    

    messageTextOnly = notif_msg
    msg_type = 'Bulletin Board'
    
    print(semid)
    if len(message) == 0:
        return jsonify({"status": "error", "message": "empty message"})

    res = spcall("post_timeline2",
                 (username, token,'',
                  '01/02/2018', message,
                  group, 'Bulletin Board',
                  True, 3, semid, schoolid
                  ), group, True)

    poster = f'{name} has posted!'
    channels = vroomid
    receivers = []
    timestamps = res[0][0]['responses'][0]['ts']
    initiatorid = res[0][0]['initiatorid']

    if group == 'admin':
        receivers = ['']
    elif len(res[0][0]['responses']) > 1:
        for item in res[0][0]['responses']:
            receivers.append(item['receiverid'])
   
    else:
        receivers.append(res[0][0]['responses'][0]['receiverid'])


    # print(f'\n\n{channels}\n\n')
    usernum = spcall("getpersonidbyusername", (username,),)[0][0]
    notif = pub.notifications(username=username,
                poster=poster, msg_payload=messageTextOnly, type=msg_type, user_type=group, 
                channels=channels, initiator_id=initiatorid, section=None,
                receiver_id=receivers, tstamp=timestamps, due_date=None, start_date=None,
                name=name, action_initiator=usernum)
    

   


    if 'Error' in res[0][0]:
        return jsonify({"status": "error", "message": res[0][0]})
    else:
        notif.notify()
    return jsonify(res[0][0])


@app.route("/assignment/faculty", methods=["GET"])
@auth.login_required
def getsubjectassignment():
    username = auth.username()
    params = request.args
    token = params["token"]
    group = params["group"]
    offeringid = params["offeringid"]
    rowskip = params["rowskip"]

    """
    select getsubjectassignment('cardo.magtanggol', 'c571e224658d280844000d664e6bec84141976ec', 
                'faculty', 
                'ESP 620202021&XCLNTXXX',
                'assignment',
                0
                );
    """


    res = spcall("getsubjectassignment", (username, token, group, offeringid, 'assignment', rowskip),
                 group, True)

    if 'Error' in res[0][0]:
        return jsonify(
            {
                "status": "error",
                "message": res[0][0]
            })

    result = res[0][0]

    tsize = result[u"size"]

    if tsize == 0:
        return jsonify({"status": "ok", "size": 0})

    timelines = result[u"timelines"]
    pic = fileServer()
    timelinelists = []

    for timeline in timelines:
        timelineitem = dict()
        timelineitem["body"] = timeline[u"body"]
        timelineitem["pic"] = pic.getImgSrcUrl(timeline[u"pic"])
        timelineitem["tstamp"] = str(timeline[u'tstamp'])
        timelineitem["tltype"] = timeline[u'tltype']
        timelineitem["initiatorid"] = timeline[u"initiatorid"]
        timelineitem["tlts"] = timeline["tlts"]
        timelineitem["reactions"] = dict({
            'okc':timeline["okc"],
            'happyc': timeline['happyc'],
            'sadc': timeline["sadc"],
            'angryc': timeline["angryc"],
            'surprisedc': timeline["surprisedc"]
        })                           #timeline[u"comments"]
        timelineitem["comments"] = timeline[u"submissions"]
        timelineitem["classsize"] = timeline[u"classsize"]
        timelinelists.append(timelineitem)

    return jsonify(
        {
            "status": "ok",
            "size": tsize,
            "timelines": timelinelists
        }
    )


@app.route("/assignment/faculty/submissions/student", methods=["GET"])
@auth.login_required
def getsubmissions():

    params = request.args
    username = auth.username()
    token = params["token"]
    group = params["group"]
    ts = params["ts"]
    studentid = params["lrn"]

    res = spcall("get_comments_simplify", (
        username, token, group,  ts, studentid), group, True)

    """
    get_comments_simplify(
        par_username text,
        par_token text, 
        par_group text, 
        par_ts timestamp without time zone,
        par_studentid text)
    """


    if 'Error' in res[0][0]:
        return jsonify(
            {
                "status": "error",
                "message": res[0][0]
            })

    return jsonify({"status": "ok", "size": res[0][0]["size"] ,"results": res[0][0]["comments"]})


@app.route("/assignment/faculty/link/score", methods=["POST"])
@auth.login_required
def savelinkstuscore():

    username = auth.username()
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    lrn = params["studentid"]
    entryname = params["entryname"]
    score = params["score"]
    periodid = params["periodid"]
    categoryid = params["categoryid"]
    semid = params["semid"]
    ts = params["ts"]
    offeringid = params["sectionid"]
    """
    savestuscorelinkpost(
  par_username text,
    par_token text,
    par_group text,
    par_semid text,
    par_sectioncode text,
  par_studentid text,
    par_entryname text, -- 1, 2,3,4
    par_score numeric,
    par_quarter text,
    par_categoryid text, -- 'WRITTEN WORKS'
  par_ts timestamp WITHOUT time ZONE
)
    :return:
    """

    res = spcall("savestuscorelinkpost", (
                                                 username, token, group, semid,
                                                 offeringid, lrn,
                                                 entryname, score,
                                                 periodid, categoryid,
                                                 ts
    ),
                 group, True)

    #print(res)

    if 'Error' in res[0][0]:
        return jsonify(
            {
                "status": "error",
                "message": res[0][0]
            })

    return jsonify({"status": "ok", "results": res[0][0]})




@app.route("/assignment/faculty/link", methods=["POST"])
@auth.login_required
def settlineentry():
    username = auth.username()
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    ts = params["ts"]
    offeringid = params["sectionid"]
    entryname = params["entryname"]
    maxscore = params["maxscore"]
    periodid = params["periodid"]
    categoryid = params["categoryid"]
    semid = params["semid"]

    """
    ion linkentrypost( par_username text,
        par_token text, 
        par_group text, 
        par_ts timestamp without time zone,
        par_offeringid text,
        par_entryname text,
        par_maxscore numeric,
        par_periodid text, 
        par_categoryid text,
        par_semid text)
    :return: 
    """
    res = spcall("linkentrypost", (
        username, token, group, ts, offeringid,
        entryname, maxscore, periodid, categoryid, semid),
                 group, True)

    #print(res)

    if 'Error' in res[0][0]:
        return jsonify(
            {
                "status": "error",
                "message": res[0][0]
            })

    timeres = spcall("stampnostuconv", (auth.username(), spcall("getdefault", ())[0][0],
                                        "Set Max Score to " +  str(maxscore) + " for " +
                                        spcall("getsubjectoff", (offeringid,))[0][0] +
                                        " " + categoryid + " Entry No " + entryname + " for period " +
                                        str(periodid),
                                        'grade', 2), "super", True)
    """
            timeres = spcall("stampnostuconv", (auth.username(), spcall("getdefault", ())[0][0],
                                            "Set Max Score to " + str(par_maxscore) + " for " +
                                            spcall("getsubjectoff", (par_offeringid,))[0][0] +
                                            " " + par_categoryid + " Entry No " + par_entryname +  " for period " +
                                            str(par_periodid),
                                            'grade', 2), "super", True)

    """

    return jsonify({"status": "ok", "results": res[0][0]})


@app.route("/assignment/faculty/submissions", methods=["GET"])
@auth.login_required
def getsubmissiondetails():

    params = request.args
    username = auth.username()
    token = params["token"]
    group = params["group"]
    offeringid = params["sectionid"]
    duedate = params["duedate"]
    initiatorid = params["initiator"]
    ts = params["ts"]


    res = spcall("getstudentsubmission", (
        username, token, group, offeringid, duedate,
        initiatorid, ts, "assignment"), group, True)

    if 'Error' in res[0][0]:
        return jsonify(
            {
                "status": "error",
                "message": res[0][0]
            })

    return jsonify({"status": "ok", "results": res[0]})





@app.route("/assignment/student", methods=["GET"])
@auth.login_required
def getassignmentstudents():
    username = auth.username()
    params = request.args
    token = params["token"]
    group = params["group"]
    schoolid = params["schoolid"]
    semid = params["semid"]
    subjectid = params["subjectidsection"]
    rowskip = params["rowskip"]

    res = spcall("getstudentassignment", (username, token, group, schoolid, semid, 'assignment', subjectid, rowskip),
                 group, True)

    if 'Error' in res[0][0]:
        return jsonify(
            {
                "status": "error",
                "message": res[0][0]
            })

    result = res[0][0]

    tsize = result[u"size"]

    if tsize == 0:
        return jsonify({"status": "ok", "size": 0})

    timelines = result[u"timelines"]
    pic = fileServer()
    timelinelists = []

    for timeline in timelines:
        timelineitem = dict()
        timelineitem["body"] = timeline[u"body"].split('$')[0].split('msg:')[1]
        timelineitem["pic"] = pic.getImgSrcUrl(timeline[u"pic"])
        timelineitem["tstamp"] = str(timeline[u'tstamp'])
        timelineitem["tltype"] = timeline[u'tltype']
        timelineitem["owner"] = timeline[u"body"].split('$')[2]
        timelineitem["initiatorid"] = timeline[u"initiatorid"]
        timelineitem["receiverid"] = timeline[u"receiverid"]
        timelineitem["tlts"] = timeline["tlts"]
        timelineitem["reactions"] = dict({
            'okc':timeline["okc"],
            'happyc': timeline['happyc'],
            'sadc': timeline["sadc"],
            'angryc': timeline["angryc"],
            'surprisedc': timeline["surprisedc"]
        })                           #timeline[u"comments"]
        #print(timeline[u"initiatorid"], timeline[u"receiverid"], timeline["tlts"])
        timelineitem["comments"] = spcall("get_comments_relaxed", (timeline[u"initiatorid"], timeline[u"receiverid"], timeline["tlts"], 0, auth.username()), "super", True)[0][0]
        timelinelists.append(timelineitem)

    return jsonify(
        {
            "status": "ok",
            "size": tsize,
            "timelines": timelinelists
        }
    )

@app.route("/assignment", methods=["POST"])
@auth.login_required
def assignmentpost():
    username = auth.username()
    params = request.get_json()
    message = params["message"]
    duedate = params["duedate"]
    section = params["section"]
    token = params["token"]
    group = params["group"]
    semid = params["semid"]
    schoolid = params["schoolid"]
    notif_section = params["notif_section"]
    vroomid = params["vroomid"]
    name = params["name"]

    CLEANR = re.compile('<.*?>') 
    messageTextOnly = re.sub(CLEANR, '', message)
    msg_type = 'assignment'
    # print(f'section: {section}')

    res = spcall("post_timeline2", (
        username, token, section, duedate,
        message,
        group, "assignment",
        False,
        3,
        semid,
        schoolid
    ), group, True)
    
    
    receivers = []
    initiatorid = res[0][0]['initiatorid']

    if res[0][0]['responses'][0]['ts'] != None:
        timestamps = res[0][0]['responses'][0]['ts']
    if len(res[0][0]['responses']) > 1:
        for item in res[0][0]['responses']:
            receivers.append(item['receiverid'])
    else:
        receivers.append(res[0][0]['responses'][0]['receiverid'])
    
    poster = f'{name} has posted an assignment!'
    assignment_section = notif_section

    getSection = assignment_section.split()

    vroom = spcall("getvirtualroomidbysection",(getSection[-1],),)[0][0]

    channels = vroom
    # channels = ['a934fae687b6d918841b', 'myeskwela-testchan']
    usernum = spcall("getpersonidbyusername", (username,),)[0][0]
    notif = pub.notifications(username=username,
        poster=poster, msg_payload=messageTextOnly, 
        type=msg_type, user_type=group,
        section=assignment_section,
        channels=channels, initiator_id=initiatorid, receiver_id=receivers, 
        tstamp=timestamps, due_date=duedate, start_date=None,
        name=name, action_initiator=usernum
        )

    



    if 'Error' in res[0][0]:
        return jsonify(
            {
                "status": "error",
                "message": res[0][0]
            })
    else:
        notif.notify()
    return jsonify({"status": "ok", "results": res[0]})


@app.route("/event", methods=["POST"])
@auth.login_required
def eventpost():
    params = request.get_json()
    username = auth.username()
    schoolid = params["schoolid"]
    semid = params["semid"]
    message = params["message"]
    begindate = params["begindate"]
    enddate = params["enddate"]
    token = params["token"]
    group = params["group"]
    vroomid = params["vroomid"]
    notif_msg = params["notif_msg"]
    name = params["name"]
    #default = spcall("getdefault", ())[0][0]



    msg_type = 'event'

    CLEANR = re.compile('<.*?>') 
    messageTextOnly = re.sub(CLEANR, '', notif_msg)


    if len(message) == 0:
        return jsonify({"status": "error", "message": "empty message"})


    res = spcall("post_timeline2", (username, token, '', '01/02/2018',
                                   message + '@begin:' + begindate +
                                   "#@end:" + enddate, group, 'event', True, 3, semid, schoolid
                                   ), group, True)

    receivers = []
    initiatorid = res[0][0]['initiatorid']
    timestamps = res[0][0]['responses'][0]['ts']

    
    if group == 'admin':
        receivers = ['']
    elif len(res[0][0]['responses']) > 1:
        for item in res[0][0]['responses']:
            receivers.append(item['receiverid'])
   
    else:
        receivers.append(res[0][0]['responses'][0]['receiverid'])
    
    usernum = spcall("getpersonidbyusername", (username,),)[0][0]
    poster = f'{name} has posted an event'
    channels = vroomid
    notif = pub.notifications(username=username,
            poster=poster, msg_payload=messageTextOnly, 
            type=msg_type, user_type=group, channels = channels,
            initiator_id=initiatorid, receiver_id=receivers, tstamp=timestamps,
            due_date=enddate, 
            start_date=begindate, section=None,
            name=name, action_initiator=usernum
        )
    

    if 'Error' in res[0][0]:
        return jsonify({'status': 'error', 'message': res[0][0]})
    else:
        notif.notify()
    return jsonify(res[0][0])

    # return jsonify(savetimeline
    #                (auth.username(),
    #                 default,
    #                 message + '@begin:' + begindate +
    #                 "#@end:" + enddate,
    #                 semid,
    #                 "event",
    #                 3,
    #                 schoolid
    #                 )
    #                )


@app.route("/calendar", methods=["GET"])
@auth.login_required
def getcalendarentries():
    username = auth.username()
    group = spcall("getuser_group", (username,))[0][0]
    params = request.args
    notmykid = True
    personid = spcall("getpersonidbyusername", (username,), group)[0][0].upper()
    if group == 'parents':

        idnum = params["lrn"]
        #print idnum
        kids = spcall("getmykids", (personid, 'students'), group)
        for kid in kids:
            if kid[0] == idnum:
                notmykid = False
                break

        if notmykid:
            return jsonify({"status": "error",
                            "message": "Access restricted."})
        personid = spcall("getpersonidbyidnum", (idnum,), group)[0][0]

    entries = spcall("getcalentry", (personid,), group)
    #print entries
    result = spcall("inpm", (personid, 'view calendar'), group, True)
    if len(entries) == 0:
        return jsonify({"status": "ok", "size": 0})

    if formatres(entries)["status"].lower() == "error":
        return jsonify({"status": "error", "message": formatres(entries)["message"]})

    if formatres(entries)["status"] != 'ok':
        return jsonify({"status": formatres[entries]["status"],
                        "message": formatres[entries]["message"]})
    # return jsonify({"t":entries})
    calentries = []
    for entry in entries:
        #print entry
        message = entry[0].split('.$id:')[0]
        calentry = dict()
        if entry[2] == 'assignment':
            calentry["title"] = message.split('@due:')[0].split('msg:')[1]
            calentry['start'] = message.split('@due:')[1].split(' to ')[0]
            calentry["allDay"] = "true"
            calentry["backgroundColor"] = "#f39c12"
            calentry["borderColor"] = '#030300'

        elif entry[2] == 'event':
            calentry["title"] = message.split('@begin:')[0].split('msg:')[1]
            calentry['start'] = message.split('@begin:')[1].split('#@end:')[0]
            calentry['end'] = message.split('@begin:')[1].split('#@end:')[1].split(".")[0]
            calentry["allDay"] = "true"
            calentry["backgroundColor"] = "#adbfd2"
            calentry["borderColor"] = '#030300'
        else:
            if group == 'parents' and 'You posted' in message.split('msg:')[1]:
                continue
            calentry["title"] = message.split('msg:')[1]
            calentry['start'] = entry[1]
            calentry["allDay"] = "true"
            calentry["backgroundColor"] = "#00c0ef"
            calentry["borderColor"] = '#030300'

        calentries.append(calentry)

    return jsonify({'status': 'ok', "size": len(calentries), 'items': calentries})


@app.route("/schoolyear/<string:par_semid>/<string:par_schoolid>/<string:par_token>", methods=["POST"])
@auth.login_required
def changesemester(par_semid, par_schoolid, par_token):
    username = auth.username()
    designated = spcall("changesemester", (par_semid, par_schoolid, username, par_token))[0][0]

    if 'Error' in designated:
        return jsonify({"status": "error", "message": designated})


    load = designated.pop(u"load", [])

    processedload = []

    if load["size"] > 0:
        for item in load["load"]:
            strippenitem = item.strip('{}').split(",")
            makearray = []
            i = 1
            for stritem in strippenitem:
                if i <= 6:
                    makearray.append(stritem.strip('"'))
                    i = i + 1
            makearray.append({
                "coteachers": [{
                    "fullname": strippenitem[6],
                    "idnum": strippenitem[7] + "," + strippenitem[8],
                    "pic": strippenitem[9]
                }]
            })
            processedload.append(makearray)


    return jsonify({"status": "OK", "designated": designated, "load":processedload})


@app.route("/gradebook", methods=['GET'])
@auth.login_required
def getstugradedetails():
    params = request.args
    offeringid = params["offeringid"]
    studentid = params["studentid"]
    schoolid = params["schoolid"]
    quarter = params["quarter"]
    # gradelevel = params["grade"]
    semid = params["semid"]
    group = spcall("getuser_group", (auth.username(),))[0][0]
    currid = spcall("getpersoncurriculum", (studentid, semid), group)[0][0]

    if currid == 'NONE':
        return jsonify({"status": "error",
                        "message": "Not Yet Enrolled"})

    categories = spcall("getcategorysequence", (currid,), group)
    detailscard = {"status": "ok", "category": []}
    # select * from get_entry_header('PE 120152016MAKOPA128081', '4', 'WRITTEN WORKS', getcurrsem());

    for category in categories:
        entries = spcall("get_entry_header", (offeringid, quarter, category[0], semid), group)
        categoryscores = []
        for entry in entries:
            score = spcall("getentryforstudent", (entry[0], studentid), group)
            if len(score) == 0:
                score = ''
            else:
                score = score[0][0]

            categoryscores.append({
                "entryname": entry[2],
                "maxscore": entry[1],
                "score": score
            })

        # select getstugradedistrib2('098765','MATH 120152016MAKOPA128081','1','K-12', 'WRITTEN WORKS')
        stugrade = spcall("getstugradedistrib2", (studentid, offeringid, quarter,
                                                  currid, category), group)
        detailscard["category"].append(
            {
                "name": category[0],
                "entries": categoryscores,
                "stugrade": stugrade
            })

    return jsonify(detailscard)


@app.route("/attendance", methods=['GET'])
@auth.login_required
def getstuattenddetails():
    params = request.args
    studentid = params["studentid"]
    semid = params["semid"]
    schoolid = params["schoolid"]
    month = params["month"]
    year = params["year"]
    group = spcall("getuser_group", (auth.username(),))[0][0]
    personlevel = spcall("getpersonlevel", (studentid, semid, schoolid), group)
    attend_details = spcall("get_attendancedatespersemdetailsmonth", (studentid,
                                                                      personlevel[0][0],
                                                                      personlevel[0][1],
                                                                      semid,
                                                                      schoolid,
                                                                      month + '/01/' + year
                                                                      ), group)

    d = []
    for attend_detail in attend_details:
        d.append({
            "date": attend_detail[0],
            "points": attend_detail[1],
            "daypart": attend_detail[2],
            "status": attend_detail[3]
        })

    # select * from get_studentattendancemonth('1234567', 1, 'MAKOPA', '20152016', '128081', 'July/01/2016');
    total = spcall("get_studentattendancemonth", (studentid,
                                                  personlevel[0][0],
                                                  personlevel[0][1],
                                                  semid,
                                                  schoolid,
                                                  month + '/01/' + year
                                                  ), group)[0][0].split("/")[1]
    return jsonify({
        "status": "ok",
        "attendance": d,
        "total": total
    })

@app.route("/rawscore", methods=['GET'])
@auth.login_required
def getrawscores():
    params = request.args
    username = auth.username()
    token = params["token"]
    sectioncode = params["sectioncode"]
    quarter = params["quarter"]
    group = params["group"]

    res = spcall('getrawscores',
                 (username, token,
                  group, sectioncode,
                  quarter),
                 'faculty', True)[0][0]

    if 'Error' in res:
        return jsonify({
            'status':"error",
            'message': res
        })

    return jsonify(res)


@app.route("/answerkey", methods=['GET'])
@auth.login_required
def getanswerkey():
    params = request.args
    username = auth.username()
    token = params["token"]
    sectioncode = params["sectioncode"]
    quarter = params["quarter"]
    group = params["group"]

    res = spcall('getanswerkey',
                 (username, token,
                  group, sectioncode,
                  quarter),
                 'faculty', True)[0][0]

    if 'Error' in res:
        return jsonify({
            'status':"error",
            'message': res
        })

    return jsonify(res)


@app.route("/answerkey", methods=['POST'])
@auth.login_required
def saveanswerkey():
    username = auth.username()
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    sectioncode = params["sectioncode"]
    quarter = params["quarter"]
    answers = params["answers"]
    totalitems = params["totalitems"]

    res = spcall("saveanswerkeytext", (username, token, group,
                                       sectioncode, quarter,
                                       answers, totalitems), "faculty", True)[0][0]
    if 'Error' in res:
        return jsonify({
            'status':"error",
            'message': res
        })

    return jsonify({"status":"OK", 'message':res})

@app.route("/answerkey", methods=['PUT'])
@auth.login_required
def lockanswerkey():
    username = auth.username()
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    sectioncode = params["sectioncode"]
    quarter = params["quarter"]
    # lockanswerkey(
    # par_username
    # par_token
    # par_group
    # par_sectioncode
    # par_quarter
    res = spcall("lockanswerkey",
                 (username, token, group, sectioncode, quarter),
                 "faculty", True)[0][0]

    if 'Error' in res:
        return jsonify({
            'status':"error",
            'message': res
        })

    return jsonify({"status": "OK", 'message': res})


@app.route("/rawscore/answer", methods=['GET'])
@auth.login_required
def getstudentanswers():
    # getstudentansweria(par_username
    # par_token
    # par_group
    # par_lrn
    # par_sectioncode
    # par_quarter

    params = request.args
    username = auth.username()
    token = params["token"]
    sectioncode = params["sectioncode"]
    quarter = params["quarter"]
    group = params["group"]
    lrn = params["lrn"]

    res = spcall("getstudentansweria",
                 (
                     username, token, group, lrn,
                     sectioncode, quarter),
                 'faculty', True)[0][0]

    if 'Error' in res:
        return jsonify({
            'status': "error",
            'message': res
        })

    return jsonify(res)
@app.route("/rawscore/answer", methods=['POST'])
@auth.login_required
def savestudentanswer():
    username = auth.username()
    params = request.get_json()
    token = params["token"]
    sectioncode = params["sectioncode"]
    quarter = params["quarter"]
    group = params["group"]
    lrn = params["lrn"]
    answers = params["answers"]


    # computescore(par_username
    # par_token
    # par_group
    # par_lrn
    # par_sectioncode
    # par_answers
    # par_quarter

    res = spcall("computescore", (username,
                                  token, group, lrn,
                                  sectioncode, answers,
                                  quarter), 'faculty', True)[0][0]

    if 'Error' in str(res):
        return jsonify({
            'status': "error",
            'message': str(res)
        })

    return jsonify({"status": "OK", "score": str(res)})

def stringizequarter(par_quarter):
    digq = int(par_quarter)
    if digq == 1:
        return 'First'
    elif digq == 2:
        return 'Second'
    elif digq == 3:
        return 'Third'
    elif digq == 4:
        return 'Fourth'
    else:
        return ''


@app.route("/rawscore/reports", methods=["GET"])
@auth.login_required
def getitemanalysismps():
    # getreportspersection(par_username
    # par_token
    # par_group
    # par_sectioncode
    # par_quarter
    params = request.args
    username = auth.username()
    token = params["token"]
    sectioncode = params["sectioncode"]
    quarter = params["quarter"]
    group = params["group"]

    res = spcall("getreportspersection",
                 (username, token, group, sectioncode, quarter),
                 'faculty', True)[0][0]

    if 'Error' in res:
        return jsonify({
            'status': "error",
            'message': res
        })

    wb = Workbook()
    ws = wb.active
    path = "static/downloads"
    file = "IAMPS" + sectioncode + quarter  + ".xlsx"

    file_location = os.path.join(path, file)
    if os.path.exists(file_location):
        os.remove(file_location)#return jsonify({"status": "OK", "report": res})
    #mergeAlign(ws, ["Class Record"], "A1:AJ1", centerCellText)
    mergeAlign(ws, [res["schoolname"]], "A1:K1", centerCellText)
    mergeAlign(ws, ["District of " + res["district"]], "A2:K2", centerCellText)
    mergeAlign(ws, ["Division of " + res["division"]], "A3:K3", centerCellText)
    mergeAlign(ws, ["Region " + res["region"]], "A4:K4", centerCellText)
    ws.append([""])
    mergeAlign(ws, ["SY: " + res["semdesc"]], "A6:K6", centerCellText)
    ws.append([""])
    mergeAlign(ws, ["Periodical Test Result"], "A8:K8", centerCellText)
    mergeAlign(ws, ["For " + res['subject']], "A9:K9", centerCellText)
    mergeAlign(ws, ["Section " + res['section']], "A10:K10", centerCellText)
    ws.append([""])
    mergeAlign(ws, [stringizequarter(quarter) + " Grading"], "A12:K12", centerCellText)
    ws.append([""])
    ws.append(["Teacher: " + res["teacher"], "", "", "", "Grade " + str(res["grade"]) + "-" + res["section"]])
    ws.append([""])
    mergeAlign(ws, ["Item Analysis"], "A16:L16", centerCellText)
    ws.append(["Item", "A", "B", "C", "D", "ANSWER", "CORRECT","Difficulty Index", "Interpretation","Discrimination Index", "Evaluation"])
    for items in res["item_analysis"]:
        ws.append([
            items["item"],
            items["a"],
            items["b"],
            items["c"],
            items["d"],
            items["correct_answer"],
            items["correct_answer_count"],
            items["difficulty_index"],
            items["difficulty_interp"],
            items["discrimination_index"],
            items["item_evaluation"]
        ])

    ws.append([""])
    ws.append([""])
    mergeAlign(ws, ["Raw Score and Percentage Score"], "A30:D30", centerCellText)
    ws.append(["LRN", "Name", "Score", "PS"])
    for rawscore in res["stups"]:
        ws.append([rawscore["lrn"], rawscore["name"], rawscore["score"], rawscore["ps"]])
    ws.append(["Exam Takers:", res["examtakers"]])
    ws.append(["Total Raw Score: ", res["total_raw_score"]])
    ws.append(["Total Test Item: ", res["total_test_item"]])
    ws.append(["No. of Achieving 75% PL", res["achievers"]])
    ws.append(["% Achieving 75% PL", res['achievers_percent']])
    ws.append(['Class MPS: ', res['class_mps']])
    ws.append(['Average Class PS: ', res['average_mps']])
    ws.append(['Mean Raw Score: ', res['raw_score_mean']])

    wb.save(path + "/" + file)

    return jsonify({"status": "OK", "report": res})


@app.route("/report/submission/status", methods=["GET"])
@auth.login_required
def getreportmonitoring():
    params = request.args
    username = auth.username()
    token = params["token"]
    group = params["group"]
    quarter = params["quarter"]
    gradelevel = params["gradelevel"]
    semid = params["semid"]
    schoolid = params["schoolid"]

    return restrequest("getreportsubmissionstatus",
                   (
                       username,
                       token,
                       group,
                       quarter,
                       gradelevel,
                       semid,
                       schoolid
                   ),
                   spcall,
                   group,
                   True,
                   jsonify)

@app.route("/report/submission/details", methods=["GET"])
@auth.login_required
def getreportdetails():
    params = request.args
    username = auth.username()
    token = params["token"]
    group = params["group"]
    quarter = params["quarter"]
    teacherid = params["teacherid"]
    semid = params["semid"]
    schoolid = params["schoolid"]

    return restrequest("getreportdetails",
                   (
                       username,
                       token,
                       group,
                       teacherid,
                       semid,
                       quarter,
                       schoolid
                   ),
                   spcall,
                   group,
                   True,
                   jsonify)


@app.route("/report/submission/faculty/search", methods=["GET"])
@auth.login_required
def getfacultysearch():
    params = request.args
    #token=607cba82cfa2edd2aad00ca7b70323beb04c40bb&
    #group=admin&
    # lastname=mag&
    #quarter=1&gradelevel=12&semid=20182019&schoolid=XXX
    username = auth.username()
    token = params["token"]
    group = params["group"]
    lastname = params["lastname"]
    schoolid = params["schoolid"]
    semid = params["semid"]
    gradelevel = params["level"]
    quarter = params["quarter"]

    return restrequest("searchteacher",
                   (
                       username,
                       token,
                       group,
                       lastname,
                       schoolid,
                       semid,
                       gradelevel,
                       quarter
                   ),
                   spcall,
                   group,
                   True,
                   jsonify)

@app.route("/report/submission/deadline", methods=["POST"])
@auth.login_required
def setdeadline():
    params = request.get_json()
    username = auth.username()
    token = params["token"]
    group = params["group"]
    quarter = params["quarter"]
    dlntype = params["dlntype"]
    dlndate = params["date"]
    semid = params["semid"]
    schoolid = params["schoolid"]



    return restrequest("setdeadline2",
                       (
                           username,
                           token,
                           group,
                           semid,
                           schoolid,
                           quarter,
                           dlntype,
                           dlndate
                       ),
                       spcall,
                       group,
                       True,
                       jsonify)

  


@app.route("/report/submission/faculty/remind", methods=["PUT"])
@auth.login_required
def remindfaculty():
    '''
            token: $.cookie('token'),
            group: $.cookie('usertype'),
            quarter: $("#cbodlnquarter").val(),
            semid: $.cookie("semid"),
            schoolid: $.cookie("schoolid"),
            facultyid: par_facultyid,
            sectionid:par_sectionid
    :return:
    '''
    username = auth.username()
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    quarter = params["quarter"]
    semid = params["semid"]
    schoolid = params["schoolid"]
    facultyid = params["facultyid"]
    subject = params['subject']
    section = params['section']
    title = params['title']
    ptype = params["type"]

    return restrequest("sendreminder2",
                       (
                           username,
                           token,
                           group,
                           quarter,
                           semid,
                           schoolid,
                           facultyid,
                           subject,
                           section,
                           title,
                           ptype
                       ),
                       spcall,
                       group,
                       True,
                       jsonify)




@app.route("/report/submission/faculty/remind/all", methods=["PUT"])
@auth.login_required
def remindfacultyall():
    username = auth.username()
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    facultyid = params["facultyid"]
    semid = params["semid"]
    schoolid = params["schoolid"]
    quarter = params["quarter"]

    return restrequest("sendreminderall2",
                       (
                           username,
                           token,
                           group,
                           facultyid,
                           semid,
                           schoolid,
                           quarter
                       ),
                       spcall,
                       group,
                       True,
                       jsonify
                       )


@app.route("/report/submission/status/principal", methods=["GET"])
@auth.login_required
def getschoolreportstatus():
    params = request.args
    username = auth.username()
    token = params["token"]
    group = params["group"]
    semid = params["semid"]
    quarter = params["quarter"]

    return restrequest("getschoolsubmissionstatus",
                       (
                           username,
                           token,
                           group,
                           semid,
                           quarter
                       ),
                       spcall,
                       group,
                       True,
                       jsonify
                       )


@app.route("/report/submission/principal/remind", methods=["PUT"])
@auth.login_required
def remindprincipal():
    '''
            token: $.cookie('token'),
            group: $.cookie('usertype'),
            quarter: $("#cbodlnquarter").val(),
            semid: $.cookie("semid"),
            schoolid: $.cookie("schoolid"),
            facultyid: par_facultyid,
            sectionid:par_sectionid
    :return:
    '''
    username = auth.username()
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    quarter = params["quarter"]
    semid = params["semid"]
    schoolid = params["schoolid"]
    principalid = params["principalid"]


    return restrequest("sendprincipalreminder",
                       (
                           username,
                           token,
                           group,
                           quarter,
                           semid,
                           schoolid,
                           principalid
                       ),
                       spcall,
                       group,
                       True,
                       jsonify)

@app.route("/report/submission/principal/remind/all", methods=["PUT"])
@auth.login_required
def remindprincipalall():
    '''
            token: $.cookie('token'),
            group: $.cookie('usertype'),
            quarter: $("#cbodlnquarter").val(),
            semid: $.cookie("semid"),
            schoolid: $.cookie("schoolid"),
            facultyid: par_facultyid,
            sectionid:par_sectionid
    :return:
    '''
    username = auth.username()
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    quarter = params["quarter"]
    semid = params["semid"]

    return restrequest("sendreminderallprincipals",
                       (
                           username,
                           token,
                           group,
                           quarter,
                           semid
                       ),
                       spcall,
                       group,
                       True,
                       jsonify)

@app.route("/report/mps/division", methods=["POST"])
@auth.login_required
def computempsdivision():
    username = auth.username()
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    semid = params["semid"]
    quarter = params["quarter"]

    return restrequest("computedivisionmps",
                       (
                           username,
                           token,
                           group,
                           quarter,
                           semid
                       ),
                       spcall,
                       group,
                       True,
                       jsonify)

@app.route("/report/ia/division", methods=["POST"])
@auth.login_required
def computeiadivision():
    username = auth.username()
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    semid = params["semid"]
    quarter = params["quarter"]

    return restrequest("computedivisionia",
                       (
                           username,
                           token,
                           group,
                           quarter,
                           semid
                       ),
                       spcall,
                       group,
                       True,
                       jsonify)


@app.route("/report/mps/division", methods=["GET"])
@auth.login_required
def retrievempsdivision():
    params = request.args
    username = auth.username()
    token = params["token"]
    group = params["group"]
    semid = params["semid"]
    quarter = params["quarter"]
    level = params["level"]

    mps = spcall("retrievempsdivision",
                 (
                     username,
                     token,
                     group,
                     quarter,
                     level,
                     semid
                 ), group, True)

    if 'Error' in mps[0][0]:
        return jsonify(
            {
                "status": "error",
                "message": mps[0][0]
            }
        )

    if mps[0][0]["size"] == 0:
        return jsonify(
            {
                "status": "error",
                "message": "Could not find any record!"
            }
        )
    wb = Workbook()
    ws = wb.active
    path = "static/downloads"
    file = 'MPS' + semid + quarter + level + ".xlsx"

    file_location = os.path.join(path, file)
    if os.path.exists(file_location):
        os.remove(file_location)

    ws.append(["MPS Report"])

    records = mps[0][0]["records"]
    subjects = dict()
    subjectcount = dict()

    for record in records:
        if record["subjectid"] in subjects.keys():
            subjects[record["subjectid"]] += record["mps"]
            subjectcount[record["subjectid"]] += 1.0
        else:
            subjects[record["subjectid"]] = record["mps"]
            subjectcount[record["subjectid"]] = 1.0

    ws.append([''] * 4  + subjects.keys())
    mpsaverage = []

    for subject in subjects.keys():
        mpsaverage.append(round(subjects[subject] / subjectcount[subject], 2))


    schoolmps = dict()
    for record in records:
        if not record["schoolid"] in schoolmps.keys():
            schoolmps[record["schoolid"]] = [record["schoolid"], record["name"], record['division'], record['district']] + [''] * len(subjects.keys())
        #for subject in subjects.keys(): + 4
        i = subjects.keys().index(record["subjectid"]) + 4
        schoolmps[record["schoolid"]][i] = record["mps"]


    for school in schoolmps.keys():
        ws.append(schoolmps[school])

    ws.append(['AVERAGE MPS', '', '', ''] + mpsaverage)

    wb.save(path + "/" + file)

    return jsonify({'status':'OK'})


@app.route("/report/ia/division", methods=["GET"])
@auth.login_required
def retrieveiadivision():
    params = request.args
    username = auth.username()
    token = params["token"]
    group = params["group"]
    semid = params["semid"]
    quarter = params["quarter"]
    level = params["level"]

    ia = spcall("retrieveiadivision",
                 (
                     username,
                     token,
                     group,
                     quarter,
                     level,
                     semid
                 ), group, True)

    if 'Error' in ia[0][0]:
        return jsonify(
            {
                "status": "error",
                "message": ia[0][0]
            }
        )


    wb = Workbook()
    ws = wb.active
    path = "static/downloads"
    file = 'IA' + semid + quarter + level + ".xlsx"

    file_location = os.path.join(path, file)
    if os.path.exists(file_location):
        os.remove(file_location)

    c = 0
    for ia in ia[0][0]["iarep"]:
        if len(ia["iasubjects"]) == 0:
            continue
        c = c + 1
        ws.append([ia["schoolid"], ia["schoolname"], ia["division"], ia["district"]])
        ws.append(
            [
                "Item No.", "A", "B", "C", "D",
                "Correct Answer", "Correct Count",
                "Difficulty Index","Interpretation",
                "Discrimination Index", "Evaluation"
            ])
        for subject in ia["iasubjects"]:
            ws.append([subject["subject"]])
            for iares in subject["ia"]:
                ws.append(
                    [
                        iares["itemno"],
                        iares["a"],
                        iares["b"],
                        iares["c"],
                        iares["d"],
                        iares["correctanswer"],
                        iares["correctcount"],
                        iares["difficultyi"],
                        iares["diffinterp"],
                        iares["discriminationi"],
                        iares["ieval"]
                    ]
                )

    wb.save(path + "/" + file)

    if (c == 0):
        return jsonify({'status':"error", "message":"Could not find any record!"})

    return jsonify({'status':'OK'})

@app.route("/exam/coordinators", methods=["GET"])
@auth.login_required
def retrieveexamcoordinators():
    '''
    get_exam_encoders(
          par_username text,
          par_token text,
          par_group text,
          par_quarter text,
          par_semid text,
          par_region text,
          par_level int,
          par_subjectid text
        )
    :return:
    '''
    params = request.args
    username = auth.username()
    token = params["token"]
    group = params["group"]
    quarter = params["quarter"]
    semid = params["semid"]
    region = params["region"]
    level = params["level"]
    subjectid = params["subject"]

    return restrequest(
        "get_exam_encoders",
        (
            username,
            token,
            'adminsuper10subadminmande',
            quarter,
            semid,
            region,
            level,
            subjectid
        ),
        spcall,
        "super10",
        True,
        jsonify
    )

@app.route("/exam/coordinators", methods=["PUT"])
@auth.login_required
def togglecoordinators():
    '''
    back_exam_encoder(
      par_username text,
      par_token text,
      par_group text,
      par_semid text,
      par_subjectid text,
      par_level int,
      par_quarter text,
      par_facultyid text,
      par_region text,
      par_district text)

    remove_exam_encoder(
          par_username text,
          par_token text,
          par_group text,
          par_semid text,
          par_subjectid text,
          par_level int,
          par_quarter text,
          par_facultyid text,
          par_region text,
          par_district text)
    :return:
    '''
    params = request.get_json()
    username = auth.username()
    token = params["token"]
    group = params["group"]
    semid = params["semid"]
    subject = params["subject"]
    level = params["level"]
    quarter = params["quarter"]
    region = params["region"]
    facultyid = params["facultyid"]
    district = params["district"]
    switchv = params["switchv"]

    if switchv == 'true':
        sp = 'remove_exam_encoder'
    else:
        sp = 'back_exam_encoder'

    '''
    return jsonify({"status": "error",
                    "message":
                        {
                            "username": username,
                            "token": token,
                            "group":group,
                            "semid":semid,
                            "subject":subject,
                            "level": level,
                            "quarter":quarter,
                            "facultyid":facultyid,
                            "region":region,
                            "district":district
                        }

                    })
    '''
    return restrequest(sp,
                       (
                           username,
                           token,
                           'adminsuper10subadminmande',
                           semid,
                           subject,
                           level,
                           quarter,
                           facultyid,
                           region,
                           district

                       ),
                       spcall,
                       "super10",
                       True,
                       jsonify)

@app.route("/exam/coordinators/search", methods=["GET"])
@auth.login_required
def searchcoordinators():
    '''
    searchcoordinators(par_username text,
                                              par_token text,
                                              par_group text,
                                              par_region text,
                                              par_searchtext text)
    '''
    params = request.args
    username = auth.username()
    token = params["token"]
    group = params["group"]
    region = params["region"]
    searchtext = params["searchtext"]

    return restrequest(
        "searchcoordinators",
        (
            username,
            token,
            'adminsuper10subadminmande',
            region,
            searchtext
        ),
        spcall,
        'super10',
        True,
        jsonify
    )

@app.route("/exam/coordinators", methods=['POST'])
@auth.login_required
def addcoordinator():
    '''
    initsubjectcoord(
                                              par_username text,
                                              par_token text,
                                              par_group text,
                                              par_semid text,
                                              par_subjectid text,
                                              par_level int,
                                              par_quarter text,
                                              par_facultyid text,
                                              par_region text,
                                              par_district text
        )
    '''
    username = auth.username()
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    semid = params["semid"]
    subjectid = params["subjectid"]
    level = params["level"]
    quarter = params["quarter"]
    facultyid = params["facultyid"]
    region = params["region"]
    district = params["district"]

    return restrequest(
        "initsubjectcoord",
        (
            username,
            token,
            'adminsuper10subadminmande',
            semid,
            subjectid,
            level,
            quarter,
            facultyid,
            region,
            district
        ),
        spcall,
        'super10',
        True,
        jsonify
    )


@app.route("/exam/coordinators/me", methods=['GET'])
@auth.login_required
def getsubjofme():
    '''
    getmycoordsubjects(
                                              par_username text,
                                              par_token text,
                                              par_group text,
                                              par_semid text,
                                              par_quarter text
                                            )
    '''
    username = auth.username()
    params = request.args
    token = params["token"]
    group = params["group"]
    semid = params["semid"]
    quarter = params["quarter"]


    return restrequest("getmycoordsubjects",
                       (
                           username,
                           token,
                           group,
                           semid,
                           quarter
                       ),
                       spcall,
                       'faculty',
                       True,
                       jsonify
                       )




@app.route("/exam/question", methods=['GET'])
@auth.login_required
def retrieveencodedexam():
    username = auth.username()
    params = request.args
    token = params["token"]
    group = params["group"]
    subjectid = params["subjectid"]
    level = params["level"]
    quarter = params["quarter"]
    page = params["page"]

    '''
      retrievequestions(
              par_username text,
              par_token text,
              par_group text,
              par_subectid text,
              par_level int,
              par_quarter text,
              par_page int
            )
    '''
    return restrequest('retrievequestions',
                       (
                           username,
                           token,
                           group,
                           subjectid,
                           level,
                           quarter,
                           page
                       ),
                       spcall,
                       'faculty',
                       True,
                       jsonify
                       )


@app.route("/exam/question", methods=['POST'])
@auth.login_required
def encodedexamquestion():
    '''
     savequestion(
       par_username text,
       par_token text,
       par_group text,
       par_subjectid text,
       par_level int,
       par_question text,
       par_answer text,
       par_a text,
       par_b text,
       par_c text,
       par_d text,
       par_difficultylevel text, --evaluated difficulty level (easy, medium, hard)
       par_itemtime numeric, -- in seconds
       par_points numeric,
       par_topic text,
       par_skillassessed text,
       par_semid text,
       par_schoolid text,
       par_quarter text
    )
    '''
    username = auth.username()
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    subjectid = params["subjectid"]
    level = params["level"]
    question = params["question"]
    answer = params["answer"]
    a = params["a"]
    b = params["b"]
    c = params["c"]
    d = params["d"]
    difficultylevel = params["difficultylevel"]
    itemtime = params["itemtime"]
    points = params["points"]
    topic = params["topic"]
    skillassessed = params["skillassessed"]
    semid = params["semid"]
    schoolid = params["schoolid"]
    quarter = params["quarter"]

    return restrequest(
        "savequestion",
        (
            username,
            token,
            group,
            subjectid,
            level,
            question,
            answer,
            a,
            b,
            c,
            d,
            difficultylevel,
            itemtime,
            points,
            topic,
            skillassessed,
            semid,
            schoolid,
            quarter
        ),
        spcall,
        'faculty',
        True,
        jsonify
    )

@app.route("/exam/question", methods=['PUT'])
@auth.login_required
def editexamquestion():
    '''
    create or replace function updatequestion(
   par_username text,
   par_token text,
   par_group text,
   par_questionid text,
   par_subjectid text,
   par_level int,
   par_question text,
   par_answer text,
   par_a text,
   par_b text,
   par_c text,
   par_d text,
   par_difficultylevel text, --evaluated difficulty level (easy, medium, hard)
   par_itemtime numeric, -- in seconds
   par_points numeric,
   par_topic text,
   par_skillassessed text,
   par_quarter text
)
    :return:
    '''

    username = auth.username()
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    questionid = params["questionid"]
    subjectid = params["subjectid"]
    level = params["level"]
    question = params["question"]
    answer = params["answer"]
    a = params["a"]
    b = params["b"]
    c = params["c"]
    d = params["d"]
    difficultylevel = params["difficultylevel"]
    itemtime = params["itemtime"]
    points = params["points"]
    topic = params["topic"]
    skillassessed = params["skillassessed"]
    semid = params["semid"]
    schoolid = params["schoolid"]
    quarter = params["quarter"]

    return restrequest("updatequestion",
                       (
                           username,
                           token,
                           group,
                           questionid,
                           subjectid,
                           level,
                           question,
                           answer,
                           a,
                           b,
                           c,
                           d,
                           difficultylevel,
                           itemtime,
                           points,
                           topic,
                           skillassessed,
                           quarter
                       ),spcall,
                       "faculty",
                       True,
                       jsonify)


@app.route("/exam", methods=["GET"])
@auth.login_required
def examinfo():
    '''
            par_username text,
            par_token text,
            par_group text,
            par_subjectid text,
            par_level int,
            par_quarter text,
            par_semid text,
            par_schoolid text)
            :return:
    '''
    username = auth.username()
    params = request.args
    token = params["token"]
    group = params["group"]
    subjectid = params["subjectid"]
    level = params["level"]
    quarter = params["quarter"]
    semid = params["semid"]
    schoolid = params["schoolid"]

    return restrequest("examinfo",
                       (
                           username,
                           token,
                           group,
                           subjectid,
                           level,
                           quarter,
                           semid,
                           schoolid
                       ),
                       spcall,
                       "bisor",
                       True,
                       jsonify
                       )



@app.route("/exam", methods=["POST"])
@auth.login_required
def createexam():
    '''
    in_exam(
          par_username text,
  par_token text,
  par_group text,
  par_semid text,
  par_schoolid text,
  par_quarter text,
  par_subjectid text,
  par_level int,
  par_start_time timestamp without time zone,
  par_end_time timestamp without time zone
)
    :return:
    '''
    username = auth.username()
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    semid = params["semid"]
    schoolid = params["schoolid"]
    quarter = params["quarter"]
    subjectid = params["subjectid"]
    level = params["level"]
    start_time = params["starttime"]
    end_time = params["endtime"]

    return restrequest("in_exam",
                       (
                           username,
                           token,
                           group,
                           semid,
                           schoolid,
                           quarter,
                           subjectid,
                           level,
                           start_time,
                           end_time
                       ),
                       spcall,
                       'bisor',
                       True,
                       jsonify
                       )


@app.route("/exam/questions", methods=["GET"])
@auth.login_required
def retrieveexamquestions():
    '''
    retrieveexamquestions(
                                                    par_username text,
                                                    par_token text,
                                                    par_group text,
                                                    par_semid text,
                                                    par_schoolid text,
                                                    par_quarter text,
                                                    par_subjectid text,
                                                    par_level int
                                                )
    :return:
    '''
    username = auth.username()
    params = request.args
    token = params["token"]
    group = params["group"]
    semid = params["semid"]
    schoolid = params["schoolid"]
    quarter = params["quarter"]
    subjectid = params['subjectid']
    level = params["level"]

    '''
    'EXAM' + $("#spansubject").text() +
                                      $("#spanlevel").text() +
                                      $.cookie("quarter") +
                                      $.cookie("schoolid") + 
                                      $.cookie("semid") + '.xlsx'
    '''
    examquestions = restrequest("retrieveexamquestions",
                       (
                           username,
                           token,
                           group,
                           semid,
                           schoolid,
                           quarter,
                           subjectid,
                           level
                       ),
                       spcall,
                       "faculty",
                       True,
                       lambda x:x
                       )

    #return jsonify({"status":"OK", "subjects":examquestions})

    if examquestions["status"] == 'error':
        return jsonify(examquestions)

    if examquestions["totalquestions"] == 0:
        return jsonify({"status":"error", "message":"Exam is not created."})

    wb = Workbook()
    ws = wb.active
    path = "static/downloads"
    file = 'EXAM' + subjectid + level + quarter + schoolid +  semid + '.xlsx'
    file_location = os.path.join(path, file)

    if os.path.exists(file_location):
        os.remove(file_location)

    mergeAlign(ws, ["Regional Exam for " + subjectid + " " + str(level)], "A1:I1", centerCellText)
    ws.append([""])
    ws.append([""])
    ws.append(["Name:"] + [''] * 6 + ["Date:"])
    ws.append(["Grade & Section:"] + [''] * 6 + ["Score:"])
    ws.append([])
    ws.append(["Instruction. Encircle the correct answer."])
    ws.append([])
    #print  examquestions["subjects"]["questions"]
    for question in examquestions["questions"]:
        itemno = question["itemno"]
        questionpar = question["question"]
        ws.append([itemno, cleandat(cleanhtml(questionpar), "&nbsp;", "\n")])
        ws.append(['', 'a', question["a"]])
        ws.append(['', 'b', question["b"]])
        ws.append(['', 'c', question["c"]])
        ws.append(['', 'd', question["d"]])
        ws.append([""])




    wb.save(path + "/" + file)

    return jsonify({"status": "OK"})



@app.route("/exam/publish", methods=["POST"])
@auth.login_required
def enableexam():
    '''
    enableexam(
                                                    par_username text,
                                                    par_token text,
                                                    par_group text,
                                                    par_semid text,
                                                    par_schoolid text,
                                                    par_quarter text,
                                                    par_subjectid text,
                                                    par_level int
                                                )
    :return:
    '''
    params = request.get_json()
    username = auth.username()
    token = params["token"]
    group = params["group"]
    semid = params["semid"]
    schoolid = params["schoolid"]
    quarter = params["quarter"]
    subjectid = params["subjectid"]
    level = params["level"]

    return restrequest("enableexam",
                       (
                           username,
                           token,
                           group,
                           semid,
                           schoolid,
                           quarter,
                           subjectid,
                           level
                       ),
                       spcall,
                       'bisor',
                       True,
                       jsonify
                       )


@app.route("/question/bank/info", methods=["GET"])
@auth.login_required
def retrievequestionabankinfo():
    '''
    questionbankinfo(
                par_username text,
                par_token text,
              par_group text,
              par_subjectid text,
              par_level int,
              par_quarter text,
              par_schoolid text)
    :return:
    '''
    username = auth.username()
    params = request.args
    token = params["token"]
    group = params["group"]
    subjectid = params["subjectid"]
    level = params["level"]
    quarter = params["quarter"]
    schoolid = params["schoolid"]


    return restrequest("questionbankinfo",
                       (
                           username,
                           token,
                           group,
                           subjectid,
                           level,
                           quarter,
                           schoolid
                       ),
                       spcall,
                       "faculty",
                       True,
                       jsonify
                       )


@app.route("/exam/online/itemanalysis", methods=["PUT"])
@auth.login_required
def onlineexamitemanalysis():
    '''
     token: $.cookie("token"),
    group: $.cookie("usertype"),
    subjectid: $("#spansubject").text(),
    level: $("#spanlevel").text(),
    quarter: $.cookie("quarter"),
    semid: $.cookie("semid"),
    schoolid: $.cookie("schoolid")
    :return:
    '''
    username = auth.username()
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    subjectid = params["subjectid"]
    level = params["level"]
    quarter = params["quarter"]
    semid = params["semid"]
    schoolid = params["schoolid"]

    return restrequest("onlineexamitemanalysis",
                       (username,
                        token,
                        group,
                        subjectid,
                        level,
                        quarter,
                        semid,
                        schoolid),
                       spcall,
                       "faculty",
                       True,
                       jsonify
                       )


@app.route("/exam/question/toggle", methods=["PUT"])
@auth.login_required
def togglequestion():
    username = auth.username()
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    questionid = params["questionid"]
    toggle = params["toggle"]

    return restrequest("togglequestion",
                       (
                           username,
                           token,
                           group,
                           questionid,
                           toggle == 'true'
                       ),
                       spcall,
                       'faculty',
                       True,
                       jsonify
                       )

@app.route("/exam/students/me", methods=["GET"])
@auth.login_required
def getstudentexams():
    username = auth.username()
    params = request.args
    token = params["token"]
    group = params["group"]
    quarter = params["quarter"]
    semid = params["semid"]

    return restrequest("myexams",
                       (
                           username,
                           token,
                           group,
                           quarter,
                           semid
                       ),
                       spcall,
                       "students",
                       True,
                       jsonify
                       )


@app.route("/exam/students/me/questions", methods=["GET"])
@auth.login_required
def getmyexamset():
    username = auth.username()
    params = request.args
    token = params["token"]
    group = params["group"]
    subjectid = params["subjectid"]
    level = params["level"]
    quarter = params["quarter"]
    schoolid = params["schoolid"]
    semid = params["semid"]

    return restrequest(
        "getmyexamset",
        (
            username,
            token,
            group,
            subjectid,
            level,
            quarter,
            schoolid,
            semid
        ),
        spcall,
        "students",
        True,
        jsonify
    )

@app.route("/exam/students/me/answer", methods=["POST"])
@auth.login_required
def postoeanswers():
    '''
    poststudentoeanswers(par_username text,
                                                par_token text,
                                                par_group text,
                                                par_sectioncode text,
                                                par_answers text,
                                                par_quarter text,
                                                par_schoolid text)
    :return:
    '''
    username = auth.username()
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    sectioncode = params["sectioncode"]
    answers = params["answers"]
    quarter = params["quarter"]
    schoolid = params["schoolid"]

    return restrequest(
        "poststudentoeanswers",
        (
            username,
            token,
            group,
            sectioncode,
            answers,
            quarter,
            schoolid
        ),
        spcall,
        "students",
        True,
        jsonify
    )


@app.route("/aisa", methods=["GET"])
@auth.login_required
def getschoolaisa():
    params = request.args
    username = auth.username()
    token = params["token"]
    group = params["group"]
    semid = params["semid"]

    return restrequest("getispratings",
                       (
                           username,
                           token,
                           group,
                           semid
                       ),
                       spcall,
                       'ispeval',
                       True,
                       jsonify
                       )

@app.route("/aisa", methods=["POST"])
@auth.login_required
def postschoolaisa():
    params = request.get_json()
    username = auth.username()
    token = params["token"]
    group = params["group"]
    semid = params["semid"]
    ratings = params["ratings"]
    principalid = params["principalid"]


    return restrequest("putisprating",
                       (
                           username,
                           token,
                           group,
                           principalid,
                           semid,
                           ratings
                       ),
                       spcall,
                       group,
                       True,
                       jsonify
                       )

@app.route("/aisa/publish", methods=["PUT"])
@auth.login_required
def publishschoolaisa():
    params = request.get_json()
    username = auth.username()
    token = params["token"]
    group = params["group"]
    semid = params["semid"]
    principalid = params["principalid"]


    return restrequest("publishprincipalevaluation",
                       (
                           username,
                           token,
                           group,
                           principalid,
                           semid
                       ),
                       spcall,
                       group,
                       True,
                       jsonify
                       )

@app.route("/aisa/publish", methods=["GET"])
@auth.login_required
def getprincipalschoolaisa():
    params = request.args
    username = auth.username()
    token = params["token"]
    group = params["group"]
    semid = params["semid"]

    return restrequest("getispratingsprincipal",
                       (
                           username,
                           token,
                           group,
                           semid
                       ),
                       spcall,
                       group,
                       True,
                       jsonify
                       )
@app.route("/forward", methods=["POST"])
@auth.login_required
def forwardtline():
    username = auth.username()
    params = request.get_json()
    schoolid = params["schoolid"]
    semid = params["semid"]
    message = params["message"]
    token = params["token"]
    group = params["group"]
    tltype = params["tltype"]
    vroomid = params["vroomid"]
    name = params["name"]

    CLEANR = re.compile('<.*?>') 
    messageTextOnly = re.sub(CLEANR, '', message)
    msg_type = tltype

    if len(message) == 0:
        return jsonify({"status": "error", "message": "empty message"})

    res = spcall("post_timeline2",
                 (username, token,'',
                  '01/02/2018', message,
                  group, tltype,
                  True, 3, semid, schoolid
                  ), group, True)
    
    receivers = []
    initiatorid = res[0][0]['initiatorid']
    timestamps = res[0][0]['responses'][0]['ts']
    if len(res[0][0]['responses']) > 1:
        for item in res[0][0]['responses']:
            receivers.append(item['receiverid'])
    else:
        receivers.append(res[0][0]['responses'][0]['receiverid'])
   
    poster = f'{name} forwarded a post!'
    channels = vroomid
    usernum = spcall("getpersonidbyusername", (username,),)[0][0]
    notif = pub.notifications(username=username,
                poster=poster, msg_payload=messageTextOnly, type=msg_type, 
                user_type=group, channels=channels, section=None,
                initiator_id=initiatorid, receiver_id=receivers,
                tstamp=timestamps, 
                due_date=None, start_date=None, name=name, 
                action_initiator=usernum)
    

    if 'Error' in res[0][0]:
        return jsonify({"status": "error", "message": res[0][0]})
    else:
        notif.notify()

    return jsonify(res[0][0])



@app.route("/comment", methods=["POST"])
@auth.login_required
def postcomment():
    params = request.get_json()
    username = auth.username()
    token = params["token"]
    group = params["group"]
    initiatorid = params["initiatorid"]
    receiverid = params["receiverid"]
    timelinets = params["timelinets"]
    comment = params["comment"]
    vroomid = params["vroomid"]
    notif_msg = params["notif_msg"]

    messageTextOnly = notif_msg
    msg_type = 'comment'

    if len(cleandat(comment, ' ', '')) == 0:
        return jsonify({'status': 'error', 'message': 'Empty Comment'})
   
 

    result = spcall('post_comments2',
               (
                   username, token,
                   group, initiatorid,
                   receiverid, timelinets,
                   comment
               ),
               group,
               True)[0][0]
    
    
    ts = result['ts']
    usernum = spcall("getpersonidbyusername", (username,),)[0][0]
    person = spcall("getpersonname", (usernum,),)[0][0]
    person = person.split("*")
    commentor = f'{person[1]} {person[0]} has commented on a post!'
    name =f'{person[1]} {person[0]}'
    channels = vroomid
    receiver_id = []
    receiver_id.append(receiverid)
    notif = pub.notifications(username=username,
                poster=commentor, msg_payload=messageTextOnly, 
                type=msg_type, user_type=group, channels=channels,
                initiator_id=initiatorid, receiver_id=receiver_id,
                tstamp=ts,due_date=None, start_date=None, section=None,
                name=name, action_initiator=usernum)
    
    
    # print result

    if 'Error' in result:
        return jsonify({"status": "error", "message": result})
    else:
        notif.notify()
    return jsonify(
        result
    )

    # return jsonify({"status": "error", "message": initiatorid})

@app.route("/comment", methods=["GET"])
@auth.login_required
def getcomment():
    params = request.args
    username = auth.username()
    token = params["token"]
    group = params["group"]
    initiatorid = params["initiatorid"]
    receiverid = params["receiverid"]
    timelinets = params["timelinets"]
    offset = params["offset"]


    result = spcall('get_comments',
               (
                   username, token,
                   group, initiatorid,
                   receiverid, timelinets, offset
               ),
               group,
               True)[0][0]

    #print result

    if 'Error' in result:
        return jsonify({"status": "error", "message": result})

    return jsonify(
        result
    )


@app.route("/reaction", methods=["POST"])
@auth.login_required
def postreaction():
    username = auth.username()
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    initiatorid = params["initiatorid"]
    receiverid = params["receiverid"]
    timelinets = params["timelinets"]
    reaction = params["reaction"]
    vroomid = params["vroomid"]
    personname = params["personname"]


    
    if reaction == '1':
        messageTextOnly = 'Like'
    elif reaction == '2':
        messageTextOnly = 'Happy'
    elif reaction == '3':
        messageTextOnly = 'Sad'
    elif reaction == '4':
        messageTextOnly = 'Angry'
    elif reaction == '5':
        messageTextOnly = 'Surprised'

    msg_type = 'reaction'
    res = spcall(
            "post_reaction2",
            (
                username,
                token,
                group,
                initiatorid,
                receiverid,
                timelinets,
                reaction
            ),
            group,
            True
        )[0][0]

    ts = res['ts']
    usernum = spcall("getpersonidbyusername", (username,),)[0][0]
    poster = f'{personname} reacted on a post!'
    name = personname
    msg = f'{personname} reacted has reacted with {messageTextOnly} on a post'
    channels = vroomid
    receiver_id = []
    receiver_id.append(receiverid)
    notif = pub.notifications(username=username,
                poster=poster, msg_payload=msg, 
                type=msg_type, user_type=group, channels=channels,
                initiator_id=initiatorid, receiver_id=receiver_id,
                tstamp=ts, due_date=None, start_date=None, section=None,
                name=name, action_initiator=usernum)
    

    if 'Error' in res:
        return jsonify({"status": "error", "message": res})
    else:
        notif.notify()
    return jsonify(
        res
    )


@app.route("/report/enrollment/faculty/remind", methods=["PUT"])
@auth.login_required
def remindfacultyenrollment():
    '''
            token: $("#name-rightbadge").data('token'),
            group: $("#name-rightbadge").data('usertype'),
            semid: $("#name-rightbadge").data("semid"),
            schoolid: $("#name-rightbadge").data("schoolid"),
            facultyid: par_facultyid
    :return:
    '''
    username = auth.username()
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    semid = params["semid"]
    schoolid = params["schoolid"]
    facultyid = params["facultyid"]

    return restrequest("sendreminderenrollment2",
                       (
                           username,
                           token,
                           group,
                           semid,
                           schoolid,
                           facultyid
                       ),
                       spcall,
                       group,
                       True,
                       jsonify)


@app.route("/help", methods=['POST'])
@auth.login_required
def savehelpaccess():
    username = auth.username()
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    helptype = params["helptype"]

    res = spcall(
        "accesshelp",
        (
            username,
            token,
            group,
            helptype
        ),
        group,
        True
    )[0][0]

    if 'Error' in res:
        return jsonify({"status": "error", "message": res})

    return jsonify({"status":res})


@app.route("/gradebook/report/levprof/teacher", methods=['GET'])
@auth.login_required
def levprofteacher():
    username = auth.username()
    params = request.args
    token = params["token"]
    group = params["group"]
    offeringid = params["offeringid"]
    quarter = params["quarter"]

    res = spcall(
        "getlevprofteacher",
        (
            offeringid,
            quarter,
            username,
            group,
            token
        ),
        group,
        True
    )[0][0]

    if 'Error' in res:
        return jsonify({"status": "error", "message": res})

    return jsonify(res)

@app.route("/gradebook/report/levprof/principal", methods=['GET'])
@auth.login_required
def levprofprincipal():
    username = auth.username()
    params = request.args
    '''
         token: $("#name-rightbadge").data("token"),
         group: $("#name-rightbadge").data("usertype"),
         subject: $("#cbosubjects").val(),
         level: $("#cbogrdlevel").val(),
         schoolid: $("#name-rightbadge").data("schoolid"),
         quarter: $("#cboquarter").val()
    
    par_username text,
    par_group text,
    par_token text,
    par_subject text,
    par_level int,
    par_schoolid text,
    par_quarter_ text
    
    '''

    token = params["token"]
    group = params["group"]
    subject = params["subject"]
    level = params["level"]
    schoolid = params["schoolid"]
    quarter = params["quarter"]
    semid = params["semid"]


    res = spcall(
        "getlevofproficiencyprincipal",
        (
            username,
            group,
            token,
            subject,
            level,
            schoolid,
            quarter,
            semid
        ),
        group,
        True
    )[0][0]

    if 'Error' in res:
        return jsonify({"status": "error", "message": res})

    return jsonify(res)


@app.route("/gradebook/report/levprof/principal/submit", methods=['POST'])
@auth.login_required
def levprofprincipalsubmit():
    username = auth.username()
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    subject = params["subject"]
    level = params["level"]
    quarter = params["quarter"]
    schoolid = params["schoolid"]
    semid = params["semid"]
    '''
    submitperfprincipal(
                                    par_username text,
                                    par_token text,
                                    par_group text,
                                    par_subject text,
                                    par_level int,
                                    par_quarter text,
                                    par_schoolid text,
                                    par_semid text
    )
    '''
    res = spcall(
        "submitperfprincipal",
        (
            username,
            token,
            group,
            subject,
            level,
            quarter,
            schoolid,
            semid
        ),
        group,
        True
    )[0][0]

    if 'Error' in res:
        return jsonify({"status": "error", "message": res})

    return jsonify(res)

@app.route("/image/upload", methods=['POST'])
@auth.login_required
def imageuploadbutton():
    username = auth.username()
    '''
    params = request.get_json()
    token = params["token"]
    group = params["group"]
    lrn = params["lrn"]
    '''
    #print request.files['file']
    #file = request.files['file']
    #print file
    return jsonify({"status":"error", "message":"hurray"})

@app.route("/dashboard/impression", methods=["GET"])
@auth.login_required
def dashboardimpression():
    #token, group, subject, quarter, semid
    params = request.args

    username = auth.username()
    token = params["token"]
    group = params["group"]
    semid = params["semid"]
    quarter = params["quarter"]

    return restrequest("genimpression",
                        (
                            username,
                            token,
                            group,
                            semid,
                            quarter
                        ),
                        spcall,
                        group,
                        True,
                        jsonify
                       )


@app.route("/dashboard/impression/details", methods=["GET"])
@auth.login_required
def dashboardimpressiondetails():
    params = request.args

    username = auth.username()
    token = params["token"]
    group = params["group"]
    semid = params["semid"]
    quarter = params["quarter"]
    subject = params["subject"]
    gradelevel = params["gradelevel"] #could be all

    return restrequest("getimpressiondetails",
                        (
                            username,
                            token,
                            group,
                            semid,
                            quarter,
                            subject,
                            gradelevel
                        ),
                        spcall,
                        group,
                        True,
                        jsonify
                       )


@app.route("/dashboard/impression/details/district", methods=["GET"])
@auth.login_required
def dashboardimpressiondetailsdistrict():
    params = request.args

    username = auth.username()
    token = params["token"]
    group = params["group"]
    semid = params["semid"]
    quarter = params["quarter"]
    subject = params["subject"]
    gradelevel = params["gradelevel"] #could be all
    district = params["district"]

    return restrequest("getschoolimpressiondetails",
                        (
                            username,
                            token,
                            group,
                            subject,
                            gradelevel,
                            quarter,
                            semid,
                            district
                        ),
                        spcall,
                        group,
                        True,
                        jsonify
                       )


@app.route("/person", methods=['GET'])
@auth.login_required
def searchpersonbcast():
    username = auth.username()
    params = request.args
    token = params["token"]
    searchkey = params["searchkey"]
    result = spcall("searchpersongen", (searchkey, username, token))

    if len(result) == 0:
        return jsonify({
            "status": "OK",
            "size": 0
        })

    if 'Error' in result[0][0]:
        return jsonify({
            'status':'error',
            'message':result[0][0]
        })


    ret = []
    for item in result:
        stu = dict({
            "lrn": item[0],
            "pid": item[1],
            "fullname": item[2],
            "picurl": item[3],
            'type': item[4]
        })
        ret.append(stu)

    return jsonify({
        "status": "OK",
        "size": len(ret),
        "list": ret
    })


@app.route("/message", methods=['POST'])
@auth.login_required
def messagepost():
    params = request.get_json()
    schoolid = params["schoolid"]
    semid = params["semid"]
    message = params["message"]
    token = params["token"]
    group = params["group"]
    recipients = params["persons"]
    username = auth.username()

    if len(message) == 0:
        return jsonify({"status": "error", "message": "empty message"})

    res = spcall("sendmessage",
                 (username, token,message,
                  recipients, 'event', 3,
                  semid, schoolid, group
                  ), group, True)


    if 'Error' in res[0][0]:
        return jsonify({"status": "error", "message": res[0][0]})


    return jsonify(res[0][0])


@app.route("/log", methods=['POST'])
@auth.login_required
def locactivity():
    params = request.get_json()
    message = params["message"]
    token = params["token"]
    group = params["group"]
    username = auth.username()

    if len(message) == 0:
        return jsonify({"status": "error", "message": "empty message"})

    res = spcall("logactivity",
                 (username, token, message, group), group, True)


    if 'Error' in res[0][0]:
        return jsonify({"status": "error", "message": res[0][0]})


    return jsonify(res[0][0])


@app.route("/mail", methods=['GET'])
#@auth.login_required
def sendmail():
    msg = Message("Hello",
                  sender="myeskwela.app@gmail.com",
                  recipients=["orven.llantos@gmail.com"])
    msg.body = "sdfsfsdsfs "
    msg.html = "<h1>hehre</h1>"

    #print(msg)
    mail.send(msg)
    return jsonify({'status':'OK'})

import qrcode

@app.route("/qrcode", methods=["POST"])
@auth.login_required
def createqrcode():
    username = auth.username()
    params = request.get_json()
    group = params["group"]
    name = params["name"]
    res = spcall("checkqrcode", (username,), group, True)[0][0]
            
    status = 'OK'
    message = res    
    if res == 'CR':
        img = qrcode.make(username + ',' + name)
        fnamestore = ('my.e.qr' + digest(username), "png")
        fnamestr = 'my.e.qr' + digest(username) + ".png"
        filename = os.path.join(app.config['UPLOAD_FOLDER'],
                                "%s.%s" %
                                fnamestore)
        res = spcall("ins2qrcode", (username, fnamestr), group, True)[0][0]
        img.save(filename)
        message = fnamestr  
    

    return jsonify({"status": status, "message": message})


@app.route("/offering", methods=["DELETE"])
@auth.login_required
def deleteoffering():
    username = auth.username()
    params = request.get_json()
    offeringid = params["offeringid"]
    group = params["group"]
    level = params["level"]
    semswitch = params["semswitch"]
    res = spcall("removeoffering", (offeringid, username, int(level), str(semswitch)), group, True)[0][0]
    if res != 'OK':
        return jsonify({"status": "error", "message": res})

    return jsonify({"status": 'OK', "message": res})


@app.route("/subjects", methods=["GET"])
@auth.login_required
def encsubjects():
    username = auth.username()
    params = request.args
    group = params["group"]

    res = spcall("fetchencodedsubjects", (), group)[0][0]

    if 'Error' in res:
        return jsonify({"status": "error", "message":res})


    return jsonify({"status": "OK",
                    "result": res
                    })



@app.route("/offering", methods=["POST"])
@auth.login_required
def insertoffering():
    username = auth.username()
    params = request.get_json()
    group = params["group"]
    subjectid = params["subjectid"]
    description = params["description"]
    level = params["level"]
    semid = params["semid"]
    semswitch = params["semswitch"]
    section  = params["section"]
    schoolid = params["schoolid"]

    result = spcall("createofferingbyteacher",
                    (username, subjectid,
                     description, level,
                     semid, str(semswitch), section, schoolid),
                    group, True)[0][0]

    if 'Error' in result:
        return jsonify({'status': 'error', 'message': result})

    return jsonify({'status':'OK', 'message': result})


@app.route("/notifsetdeadline", methods=["POST"])
@auth.login_required
def notif_setdeadline():
    params = request.get_json()
    username = auth.username()
    chan = params["channels"]
    group = params["group"]
    initiatorid = params["initiatorid"]
    
    timestamps = params["ts"]
    name = params["name"]
    date = params["date"]
    dltype = params["dltype"]
    quarter = params["quarter"]
    channels = chan.split()


    date_object = datetime.strptime(date, "%m-%d-%Y")
    clean_date = date_object.strftime("%b %d %Y")

    messageTextOnly = f'''{name} posted a deadline!\n
    Type of deadline is {dltype.upper()}
    For Quarter {quarter}
    Deadline is on {clean_date.upper()}
    '''


    msg_type = 'deadline'

    usernum = spcall("getpersonidbyusername", (username,),)[0][0]
    poster = f'{name} has posted a deadline'

    channels = [(''.join(i.split(','))) for i in channels ]
    
    if group == 'admin':
        receivers = ['']

    else:
        receivers = receivers.split()
        receivers = [(''.join(i.split(','))) for i in receivers ]

    


    notif = pub.notifications(username=username,
            poster=poster, msg_payload=messageTextOnly, 
            type=msg_type, user_type=group, channels = channels,
            initiator_id=initiatorid, receiver_id=receivers, tstamp=timestamps,
            due_date=date, 
            start_date=None, section=None,
            name=name, action_initiator=usernum
        )
        
    notif.notify()

    return jsonify({"status": "OK"})
    


@app.route("/newnotifcount", methods=["GET"])
@auth.login_required
def getnewnotif_count():
    params = request.args
    chan = params["channels"]
    personid = params["personid"]
    group = params["group"]
    channels = chan.split()
    kid_id = params["kid_id"]
    kids = kid_id.split()

    channels = [(''.join(i.split(','))) for i in channels ]

    # channels = ['a934fae687b6d918841b', 'myeskwela-testchan']
    # print(f'\n\nchannels: {channels}\ntype: {type(channels)}\n\n')
    res = spcall("getnewnotifcount",(True, channels, personid, group, kids),)[0][0]

    print(res)

    return jsonify({'status':'OK', 'count': res})

@app.route("/readnewnotif", methods=["POST"])
@auth.login_required
def readnewnotif():
    params = request.get_json()
    chan = params["channels"]
    personid = params["personid"]
    group = params["group"]

    channels = chan.split()
    kid_id = params["kid_id"]
    kids = kid_id.split()
    channels = [(''.join(i.split(','))) for i in channels ]
    # channels = ['a934fae687b6d918841b', 'myeskwela-testchan']
    # print(f'\n\nchannels: {channels}\ntype: {type(channels)}\n\n')
    # res = spcall("readnewnotification",(channels, personid), group, True)[0][0]
    res = spcall("getnotification",(channels, personid, group, kids),)[0][0]
    notifs = res["notifs"]
    # print(notifs)
    for item in notifs:
        # print(f'\nnotif id: {item["notif_id"]}\ninitiatorid:{item["initiatorid"]}\n\n')
        seen_result = spcall("see_newnotif", (item["notif_id"], personid), group, True)[0][0]
    print(seen_result)
    return jsonify({'status':seen_result})

@app.route("/readnotif", methods=["POST"])
@auth.login_required
def readnotif():
    params = request.get_json()
    personid = params["personid"]
    group = params["group"]
    notif_id = params["notifid"]

    
    
    res = spcall("read_notif", (notif_id, personid), group, True)[0][0]
    
    return jsonify({'status':res})

@app.route("/getnotif", methods=["GET"])
@auth.login_required
def getnotif():
    params = request.args
    chan = params["channels"]
    personid = params["personid"]
    group = params["group"]
    channels = chan.split()
    kid_id = params["kid_id"]
    kids = kid_id.split()
    # channels = ['a934fae687b6d918841b', 'XXX', 'F2018CARDOMAGTANGGOL-1']
    # print(f'\n\nchannels: {channels}\ntype: {type(channels)}\n\n')
    channels = [(''.join(i.split(','))) for i in channels ]
    res = spcall("getnotification",(channels, personid, group, kids),)[0][0]
    # print(res)

    return jsonify({'status':'OK', 'notifs': res["notifs"] })

@app.route("/getpost", methods=["GET"])
@auth.login_required
def getpost():
    params = request.args
    intiatorid = params["initiatorid"]
    receiverid = params["receiverid"]
    ts = params["ts"]

    timelines = spcall("notif_poppost",(auth.username(), intiatorid, receiverid, ts))[0][0]

    # print(timelines)
    if timelines["status"] == "Error":
        return jsonify({"status": "error", "message": timelines["item"][0]})

    tsize = timelines["size"]

    if tsize == 0:
        return jsonify({"status": "ok",
                        "size": 0})

    timelines = timelines["post"]
    pic = fileServer()
    timelinelists = []
    for timeline in timelines:
        ##this is for showing the feelings with its count
        # recordentryid = ""
        # recentfeeling = ""
        # if timeline[2] == 'grade' and group == 'students':
        # owner = timeline[0].split('$')[2]
        # ts = str(timeline[1])
        # feelings = spcall("showfeelings", (), group)
        # user = spcall("getpersonidbyusername", (auth.username(),), group)[0][0]

        # recordentryid = spcall("getrecordentryidbyts", (str(timeline[3]),))[0][0]
        # recordentryid = getrecordentryid(timeline[0].split('$')[0].split('msg:')[1], timeline[0].split('$')[1].split(':')[1], user, auth.username(), str(timeline[3]))
        # return jsonify({"FF":timeline[0].split('$')[0].split('msg:')[1]})
        # entryid = spcall("getentryidbyrecordentryid", (recordentryid,), group)[0]
        # recentfeeling = spcall("recentfeeling", (user, entryid,), group)
        # return jsonify({"FF":recentfeeling})

        timelineitem = dict()
        timelineitem["body"] = timeline[u"body"].split('$')[0].split('msg:')[1]
        timelineitem["pic"] = pic.getImgSrcUrl(timeline[u"pic"])
        # timelineitem["pic"] = timeline[u"pic"]
        timelineitem["tstamp"] = str(timeline[u'tstamp'])
        timelineitem["tltype"] = timeline[u'tltype']
        timelineitem["owner"] = timeline[u"body"].split('$')[2]
        # timelineitem["feelings"] = feelings
        # timelineitem["recentfeeling"] = recentfeeling
        # timelineitem["recordentryid"] = recordentryid
        timelineitem["initiatorid"] = timeline[u"initiatorid"]
        timelineitem["receiverid"] = timeline[u"receiverid"]
        timelineitem["tlts"] = timeline["tlts"]
        timelineitem["reactions"] = dict({
            'okc':timeline["okc"],
            'happyc': timeline['happyc'],
            'sadc': timeline["sadc"],
            'angryc': timeline["angryc"],
            'surprisedc': timeline["surprisedc"]
        })                           #timeline[u"comments"]
        timelineitem["comments"] = spcall("get_comments_relaxed", (timeline[u"initiatorid"], timeline[u"receiverid"], timeline["tlts"], 0, auth.username()), "super", True)[0][0]
        timelinelists.append(timelineitem)

    
    return jsonify(
        {
            "status": "ok",
            "size": tsize,
            "timelines": timelinelists 
        }
    )



@app.route("/notifreminders", methods=["POST"])
@auth.login_required
def notif_reminders():
    params = request.get_json()
    username = auth.username()

    channels = params["channels"]
    group = params["group"]
    initiatorid = params["initiatorid"]
    receiver = params["receiver"]

    reminder_type = params["reminder_type"]
    subject = params["subject"]
    section = params["section"]
    
    timestamps = params["ts"]
    name = params["name"]
    

    msg_type = 'reminder'

    usernum = spcall("getpersonidbyusername", (username,),)[0][0]
    if reminder_type == 'ALL':
        poster = f'{name} posted a Gentle Reminder'
        messageTextOnly = f'''{name} posted a Gentle Reminder for You \nReminding you Regarding the GRADE and ITEM ANALYSIS\nFor ALL SUBJECTS'''
    else: 
        poster = f'{name} posted a Gentle {reminder_type.upper()} Reminder'
        messageTextOnly = f'{name} posted a Gentle {reminder_type.upper()} Reminder for You\n For the subject {subject.upper()} section {section.upper()}'


    receivers = [receiver]
    
    print(messageTextOnly)

    notif = pub.notifications(username=username,
            poster=poster, msg_payload=messageTextOnly, 
            type=msg_type, user_type=group, channels = channels,
            initiator_id=initiatorid, receiver_id=receivers, tstamp=timestamps,
            due_date=None, 
            start_date=None, section=None,
            name=name, action_initiator=usernum
        )
        
    notif.notify()

    return jsonify({"status": "OK"})



@app.route("/addPhoneNum", methods=["POST"])
@auth.login_required
def addPhoneNum():
    params = request.get_json()
    username = auth.username()
    phoneNum = params["phoneNum"]
    personid = params["personid"]
    schoolid = params["schoolid"]
    group = params["group"]

    print(f'''
    Phone Number: {phoneNum}
    Person ID: {personid}
    School ID: {schoolid}
    User Type: {group}
    ''')


    return {"status": "OK"}


#last

# @app.route("/exam/<string:username>", methods=["POST"])
# # @auth.login_required
# def create_exam(username):
#     group = spcall("getuser_group", (username,))[0][0]
#     # return jsonify({"res": 'helllo'})
#     params = request.get_json()
#     # return jsonify({"res": params})
#     subjectid = str(params['subjectid'])
#     duration = str(params['duration'])
#     duedate = str(params['duedate'])
#     title = str(params['title'])
#
#     # add checking for the exam
#     if '' in [subjectid, duration, duedate, title]:
#         return jsonify({"status": "error", 'message': 'Exam not created. Fill up all' + \
#                                                       ' the information.'})
#
#     exam_creator = str(spcall("getpersonidbyusername", (username,))[0][0])
#     # return jsonify({"res":exam_creator})
#     exam_id = str(digest(subjectid + title + exam_creator))
#     if group == "faculty" or group == "admin":
#         date_created = str(datetime.datetime.now())
#         result = spcall("add_exam", (exam_id, subjectid, date_created, duration, duedate, title, exam_creator), group,
#                         True)
#         return jsonify({"status": 'ok', "exam_id": result[0][0]})
#     else:
#         return jsonify({"status": "error", "message": "Unauthorized access."})
#         # if exam_scope == "division":
#         #     schools = spcall("getschoolsbydivision",(division,))
#         #     for school in schools:
#         #         sections = spcall("getsectionbyschoolperyearlevel", (school[0], yearlevel))
#         #         for section in sections:
#         #             row = []
#         #             exam_id = digest(school[0] + academiclevel + yearlevel + section[0] + subjectid + title + exam_creator)
#         #             result = spcall("add_exam", (exam_id, schoolid, academic_level,
#         #                                   year_level, section, subjectid,
#         #                                   datetime.datetime.now(), duration, due_date, title,
#         #                                   examiner, division), group, True)
#         #             row.append({'schoolid':schoolid, 'section':section, 'examid':exam_id})
#         #             exam_id_list.append(row)
#
#         # elif exam_scope == "schools":
#         #     sections = spcall("getsectionbyschoolperyearlevel", (school[0], yearlevel))
#         #     for section in sections:
#         #         row = []
#         #         exam_id = digest(school[0] + academiclevel + yearlevel + section[0] + subjectid + title + exam_creator)
#         #         result = spcall("add_exam", (exam_id, schoolid, academic_level,
#         #                                   year_level, section, subjectid,
#         #                                   datetime.datetime.now(), duration, due_date, title,
#         #                                   examiner, division), group, True)
#         #         row.append({'schoolid':schoolid, 'section':section, 'examid':exam_id})
#         #         exam_id_list.append(row)
#
#         # else:
#         #     exam_id = digest(schoolid + academiclevel + yearlevel + section + subjectid + title + exam_creator)
#
#         #     result = spcall("add_exam", (exam_id, schoolid, academic_level,
#         #                                   year_level, section, subjectid,
#         #                                   datetime.datetime.now(), duration, due_date, title,
#         #                                   examiner, division), group, True)
#         #     row = []
#         #     row.append({'schoolid':schoolid, 'section':section, 'examid':exam_id})
#         #     exam_id_list.append(row)
#
#
# @app.route("/exam", methods=["GET"])
# @auth.login_required
# def get_exams(subject_id):
#     group = spcall("getuser_group", (auth.username(),))[0][0]
#     if group == "faculty":
#         exams = spcall("get_all_exams", (auth.username(),))
#     elif group == "students":
#         exams = spcall("get_student_exams", (auth.username(),))
#     elif group == "parents":
#         exams = spcall("get_child_exams", (auth.username(),))
#
#     return jsonify({"status": "ok", "exams": exams})
#
#
# @app.route("/exam/<string:exam_id>", methods=["GET"])
# @auth.login_required
# def get_exam_content(exam_id):
#     group = spcall("getuser_group", (auth.username(),))[0][0]
#     params = request.get_json()
#
#     lrn = params["lrn"]
#     is_locked = params["is_locked"]
#
#     if group == "faculty":
#         questions = ("get_all_questions", (exam_id,), group)
#         return jsonify({"status": "ok", "exam_details": {"questions": questions}})
#
#     elif group == "students":
#         if is_locked:
#             exam_score = spcall("getexamscorestudent", (exam_id, lrn))[0][0]
#
#             return jsonify({"status": "ok", "exam_details": {"exam_score": exam_score}})
#         else:
#             questions = spcall("getquestionsstudent", (exam_id,))
#
#             return jsonify({"status": "ok", "exam_details": {"questions": questions}})
#
#     else:
#         exam_score = spcall("getexamscorestudent", (exam_id, lrn))[0][0]
#         return jsonify({"status": "ok", "exam_details": {"exam_score": exam_score}})
#
#
# @app.route("/exam/<string:exam_id>", methods=["POST"])
# @auth.login_required
# def edit_exam_information(exam_id):
#     group = spcall("getuser_group", (auth.username(),))[0][0]
#
#     params = request.get_json()
#
#     schoolid = params['schoolid']
#     academiclevel = params['academiclevel']
#     yearlevel = params['yearlevel']
#     section = params['section']
#     subjectid = params['subjectid']
#     duration = params['duration']
#     duedate = params['duedate']
#     title = params['title']
#
#     if group == "faculty":
#         result = spcall("edit_information", (exam_id, schoolid, academic_level,
#                                              year_level, section, subjectid,
#                                              datetime.datetime.now(), duration, due_date, title,
#                                              examiner), group, True)[0][0]
#
#         return jsonify({"status": "ok", "result": result})
#     else:
#         return jsonify({"status": "error", "message": "Unauthorized access."})
#
#
# @app.route("/exam/<string:exam_id>", methods=["DELETE"])
# @auth.login_required
# def delete_exam(exam_id):
#     group = spcall("getuser_group", (auth.username(),))[0][0]
#
#     if group == "faculty":
#         result = spcall("delete_exam", (exam_id,), group, True)[0][0]
#
#         return jsonify({"status": "ok", "message": result})
#
#     else:
#         return jsonify({"status": "error", "message": "Unauthorized access."})
#
#
# @app.route("/exam/<string:lrn>/<string:exam_id>", methods=["POST"])
# @auth.login_required
# def submit_exam(lrn, exam_id):
#     group = spcall("getuser_group", (auth.username(),))[0][0]
#
#     if group == "students":
#         score = calculate_score(exam_id, lrn)
#
#         result = spcall("submit_exam", (exam_id, score, lrn))[0][0]
#
#         return jsonify({"status": "ok", "message": result})
#     else:
#         return jsonify({"status": "error", "message": "Unauthorized"})


# def calculate_score(exam_id, username):
#     """calculate the score"""
#
#     ans = spcall("get_answers", (exam_id, lrn))
#
#     score = 0
#     for a in ans:
#         tmp = a
#         stringed = map(str, a)
#         if stringed[0] == 'None':
#             score = score
#         else:
#             s = int(tmp[2])
#             if (stringed[0] == stringed[1]):
#                 score = score + int(s)
#     return score


# @app.route("/question/<string:username>/<string:examid>", methods=["POST"])
# # @auth.login_required
# def create_question(username, examid):
#     group = spcall("getuser_group", (username,))[0][0]
#     params = request.get_json()
#     # return jsonify({"res": params})
#     question_text = params["question_text"]
#     choices = params["choices"]
#     answer = params["answer"]
#     points = params["points"]
#     mode = params["mode"]
#     filelink = params["filelink"]
#
#     question_id = digest(str(question_text) + str(choices) + str(answer))
#
#     if group == "faculty":
#         qstn_id = spcall("create_question",
#                          (question_id, question_text, choices, answer, points, mode, filelink),
#                          group, True)[0][0]
#         spcall("add_question_to_exam", (examid, qstn_id))
#
#         return jsonify({"status": "ok", "question_id": qstn_id})
#
#     else:
#         return jsonify({"status": "error", "message": "Unauthorized access."})
#
#
# @app.route("/question/<string:question_id>", methods=["DELETE"])
# @auth.login_required
# def delete_question(question_id):
#     group = spcall("getuser_group", (auth.username(),))[0][0]
#
#     if group == "faculty":
#         result = spcall("delete_question", (question_id,))[0][0]
#
#         return jsonify({"status": "ok", "message": result})
#     else:
#         return jsonify({"status": "error", "message": "Unauthorized access."})
#
#
# @app.route("/question/<string:exam_id>/<string:question_id>/", methods=["POST"])
# @auth.login_required
# def answer_question(exam_id, question_id):
#     group = spcall("getuser_group", (auth.username(),))[0][0]
#
#     params = request.get_json()
#     answer = params["answer"]
#     lrn = params["lrn"]
#
#     if group == "students":
#         result = spcall("answer_question", (question_id, exam_id, answer, lrn))[0][0]
#
#         return jsonify({"status": "ok", "message": result})
#     else:
#         return jsonify({"status": "error", "message": "Unauthorized access."})
#
#
# @app.route("/ExamTaken/<string:exam_id>")
# # @auth.login_required
# def add_taker(exam_id, username):
#     group = spcall("getuser_group", (auth.username(),))[0][0]
#
#     if group == "students":
#         result = spcall("add_taker", (exam_id, username), True)[0][0]
#
#         return jsonify({"status": "ok", "message": result})
#     else:
#         return jsonify({"status": "error", "message": "Unauthorized access."})


@app.after_request
def add_cors(resp):
    resp.headers['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
    resp.headers['Access-Control-Allow-Credentials'] = True
    resp.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS, GET, PUT, DELETE'
    resp.headers['Access-Control-Allow-Headers'] = request.headers.get('Access-Control-Request-Headers',
                                                                             'Authorization')
    resp.headers['X-Frame-Options'] = 'SAMEORIGIN'
    resp.headers['X-XSS-Protection'] = '1; mode=block'
    resp.headers['X-Content-Type-Options'] = 'nosniff'
    # set low for debugging

    if app.debug:
        resp.headers["Access-Control-Max-Age"] = '1'
    return resp


@auth.error_handler
def invalid_cred(status):
    return make_response(jsonify({'status': 'error', 'message': str(status) + ": Unauthorized "}), '201')



@app.errorhandler(400)  # Bad Request
def unauthorized(error):
    return jsonify({'status': 'error', 'message': 'Bad Request'})


@app.errorhandler(401)  # Unauthorized
def unauthorized(error):
    return jsonify({'status': 'error', 'message': 'Unauthorized'})


#@app.errorhandler(402)  # Payment Required
#def payment_required(error):
#    return jsonify({'status': 'error', 'message': 'Payment Required'})


@app.errorhandler(403)  # Forbidden
def forbidden(error):
    return jsonify({'status': 'error', 'message': 'Forbidden'})


@app.errorhandler(404)  # not found
def not_found(error):
    return jsonify({'status': 'error', "message": 'Not Found'})


@app.errorhandler(405)  # not allowed
def not_allowed(error):
    return jsonify({'status': 'error', "message": 'Not Allowed'})


@app.errorhandler(409)  # not found
def conflict(error):
    return jsonify({'status': 'error', 'message': 'Conflict'})


def formatres(result):
    if 'Error' in result[0][0] or \
            'function' in result[0][0] or \
            'error' in result[0][0]:

        status = 'Error'
        message = result[0][0]
    else:
        status = 'ok'
        message = 'ok'
    return {'status': status,
            'message': message,
            'item': result[0]}


if __name__ == "__main__":
    app.run(debug=True, threaded=True)
