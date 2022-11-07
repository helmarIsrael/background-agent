from datetime import datetime
import re
from openpyxl.styles import Alignment
import os

def age(when, on=None):
    if on is None:
        on = datetime.today()
    was_earlier = (on.month, on.day) < (when.month, when.day)
    return on.year - when.year - (was_earlier)


def autoswitch():
    yearnow = datetime.today()

    if yearnow.month <= 6:
        now = datetime.strptime("06/01/" + str(yearnow.year), '%m/%d/%Y')
    else:
        now = yearnow
    return now

def cleandat(d, pattern, rep):
    ret = d
    if len(d) > 0:
        ret = d.replace(pattern, rep)
    return ret



def restrequest(spname, spparams, spcall, spgroup, spcommit, jsonify):
    #for use in SPs that returns JSON, or singleton

    res = spcall(spname, spparams, spgroup, spcommit)[0][0]
    if 'Error' in str(res):
        return jsonify({
            'status': "error",
            'message': str(res).split('CONTEXT:')[0].split('DETAIL:')[0].split('LINE')[0]
        })

    return jsonify(res)



#for setup: postgre is either owner or superuser
def sqlAlconn(ptype):
    #this assumes that a DATABASE_URL is properly set-up at the
    #localhost machine as well as the cloud machine
    #sample DATABASE_URL = postgresql://username:password@ipaddress:ipport/dbname
    #_osconf = os.environ['DATABASE_URL'].split('@')[1].split("/")[0].split(":")
    #_dbadd =  _osconf[0]#"localhost"#osconf[0] #'sl-us-south-1-portal.27.dblayer.com'#_osconf[0] #
    #_dbport = _osconf[1]#_osconf[1] #'5432'#'47445'#'5433'#

    _connstr = {"super": { 'user':'postgre',
                           'password':'UflY9ib43U'
                         },
                "faculty":{'user':'teacher',
                            'password':'T5XP9MbHoR'
                          },
                "admin":  {'user':'trator',
                           'password':'k2oDBalkTs'
                          },
                "principal": {'user': 'trator',
                          'password': 'k2oDBalkTs'
                          },
                "students": {'user':'aaral',
                             'password':'YxYKWg48l7'
                            },
                "staff":   {'user':'assist',
                           'password':'cEBAdKWFcL'
                           },
                "registrar":
                           {'user':'regis',
                          'password':'gNNVaY6MKZ'
                           },
                "parents":{
                          'user':'roots',
                          'password':'FrXnD4UGCF'
                          },
                "super10":{
                           'user':'super10dent',
                           'password': 'uz7En2vAmE'
                           },
                "bisor":
                    {
                        'user': 'subisor',
                        'password': 'q0owSP3FLa'


                },
                "ispeval":
                {
                        'user': 'ispbisor',
                        'password': 'mtFq5EEZca'
                },
                "subadmin":
                    {
                        'user': 'subadminagent',
                        'password': 'SwR18ci7fE'
                    },
                "mande":
                    {
                        'user': 'mandemore',
                        'password': 'qC7NpLsdqr'
                    },

                "datamart":{
                          'user':'postgres',
                          'password':'12345'
                }
                }

    if ptype == 'datamart':
      _dbname = 'datamart1'
    else:
      _dbname = 'my2' #local:'my2'

    _dbadd = 'localhost'#'sl-us-south-1-portal.27.dblayer.com'
    _dbport = '5432'

    #_dbadd = 'localhost'
    #_dbport = '5433'

    return "postgresql://"+ _connstr[ptype]["user"] + \
            ":" + _connstr[ptype]["password"] + "@"+ \
            _dbadd + ":" + _dbport + "/" + _dbname
'''
def constr(ptype):
    _dbadd = 'myeskwela.postgres.database.azure.com'
    _dbport = '5432'
    _connstr = {"super":  'user=myeskwela2 ' + \
                          'dbname=myeskwela ' + \
                          'password=letmein ' + \
                          'host='+ _dbadd + " " + \
                          'port='+_dbport,
                "faculty":'user=teacher ' + \
                          'dbname=myeskwela ' + \
                          'password=esdh2398uis ' + \
                          'host='+ _dbadd + " " + \
                          'port='+_dbport,
                "admin":  'user=trator ' + \
                          'dbname=myeskwela ' + \
                          'password=sn82fnnw89 ' + \
                          'host='+ _dbadd + " " + \
                          'port='+_dbport,
                "students":
                          'user=aaral ' + \
                          'dbname=myeskwela ' + \
                          'password=32950jwerhr ' + \
                          'host='+ _dbadd + " " + \
                          'port='+_dbport,
                "staff":  'user=assist ' + \
                          'dbname=myeskwela ' + \
                          'password=wfj2309ufh19 ' + \
                          'host='+ _dbadd + " " + \
                          'port='+_dbport,
                "registrar":
                          'user=regis ' + \
                          'dbname=myeskwela ' + \
                          'password=0-832wer1351 ' + \
                          'host='+ _dbadd + " " + \
                          'port='+_dbport,
                "parents":
                          'user=roots ' + \
                          'dbname=myeskwela ' + \
                          'password=wnf58-12jfg38 ' + \
                          'host='+ _dbadd + " " + \
                          'port='+_dbport,
                "bisor":  'user=subisor ' + \
                          'dbname=myeskwela ' + \
                          'password=q0owSP3FLa ' + \
                          'host='+ _dbadd + " " + \
                          'port='+_dbport}
    return _connstr[ptype]
'''
#print age(datetime.strptime("12/1/1978", '%m/%d/%Y'), autoswitch())

#def age(dateString):
#  #code originally from http://stackoverflow.com/a/10119380
#    yearnow = datetime.today()

#    if yearnow.month <= 6:
#        now = datetime.strptime("06/01/" + str(yearnow.year), '%m/%d/%Y')
#    else:
#        now = yearnow



#    yearNow = now.year
#    monthNow = now.month
#    dateNow = now.day

#    dob = datetime.strptime(dateString, '%m/%d/%Y')

#    yearDob = dob.year
#    monthDob = dob.month
#    dateDob = dob.day

#    yearAge = yearNow - yearDob

#    if monthNow >= monthDob:
#        monthAge = monthNow - monthDob
#    else:
#        yearAge = yearAge - 1
#        monthAge = 12 + monthNow - monthDob

#    if dateNow >= dateDob:
#        dateAge = dateNow - dateDob
#    else:
#        monthAge = monthAge - 1
#        dateAge = 31 + dateNow - dateDob
#        if monthAge < 0:
#            monthAge = 11
#            yearAge = yearAge - 1
#    return yearAge

#print sqlAlconn("registrar")


def checkpassword(password):
    errorm = ''
    if len(password) < 8:
        errorm += ' Must be at least 8 characters.'

    if len(re.compile(r'[0-9]').findall(password)) == 0:
        errorm += ' Must contain at least one digit.'

    if len(re.compile(r'[a-z]').findall(password)) == 0:
        errorm += ' Must contain at least one small character.'

    if len(re.compile(r'[A-Z]').findall(password)) == 0:
        errorm += ' Must contain at least one capital character.'

    return errorm


def updatepassword(par_username):
    password = spcall('random_string',(10,))[0][0]
    res = spcall('updatepassword', (par_username, password), 'super', True)
    if res[0][0] == 'OK':
        return par_username + '->' + password
    else:
        return res[0][0]

def centerCellText(ws,address):
    currentCell = ws.cell(address)
    currentCell.alignment = Alignment(horizontal='center')

def rightCellText(ws,address):
    currentCell = ws.cell(address)
    currentCell.alignment = Alignment(horizontal='right')

def leftCellText(ws,address):
    currentCell = ws.cell(address)
    currentCell.alignment = Alignment(horizontal='left')

def appendBlanks(ws, n):
    for i in range(n):
        ws.append([''])

def mergeAlign(ws, row,range, alignfcn):
    ws.append(row)
    ws.merge_cells(range)
    alignfcn(ws, range.split(":")[0])

def mergeAlignOnly(ws, range, alignfcn):
    ws.merge_cells(range)
    alignfcn(ws, range.split(":")[0])

#thanks to: https://stackoverflow.com/a/48698261
def letter_range(start, stop="{", step=1):
    """Yield a range of lowercase letters."""
    for ord_ in range(ord(start.lower()), ord(stop.lower()), step):
        yield chr(ord_)


def finalize(res):
    if 'Error' in res[0][0]:
        return {"status": "error", "message": res[0][0]}


    return res[0][0]

def cleanhtml(raw_html): #thanks to: https://stackoverflow.com/a/12982689
  cleanr = re.compile('<.*?>')
  cleantext = re.sub(cleanr, '', raw_html)
  return cleantext

def errordict(errmsg):
    return {
            'status': 'error',
            'msg': str(errmsg)
        }

def lengthesize(v,l):
    if len(v) <= l:
        return v
    else:
        return v[0:l]

def forcemax(token):
    return lengthesize(token, 128)
